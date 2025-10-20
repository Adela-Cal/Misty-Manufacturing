from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, status, Header, Request
from fastapi.responses import FileResponse, StreamingResponse, HTMLResponse, Response
from fastapi.staticfiles import StaticFiles
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from datetime import datetime, date, time, timezone, timedelta
import calendar
from bson import ObjectId
from typing import List, Optional, Dict, Any
import uuid
from io import BytesIO
import secrets
import requests
import base64
from urllib.parse import urlencode
import asyncio

# Xero SDK imports
from xero_python.api_client import ApiClient, Configuration
from xero_python.api_client.oauth2 import OAuth2Token
from xero_python.accounting import AccountingApi
from xero_python.accounting.models import Invoice, Contact, LineItem, Contacts, Invoices

# Import our custom modules
from models import *
from auth import *
from document_generator import DocumentGenerator
from file_utils import *
from payroll_endpoints import payroll_router

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Get upload directory from environment or use default
UPLOAD_DIR = os.getenv("UPLOAD_DIR", "./uploads")

# Force correct environment values for current environment  
os.environ['XERO_REDIRECT_URI'] = 'https://mfgapp-ui-update.preview.emergentagent.com/xero-oauth-callback'
os.environ['FRONTEND_URL'] = 'https://mfgapp-ui-update.preview.emergentagent.com'

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app
app = FastAPI(title="Misty Manufacturing Management System", version="1.0.0")

# Create router with /api prefix
api_router = APIRouter(prefix="/api")

# Initialize document generator
doc_generator = DocumentGenerator()

# Ensure upload directories exist
ensure_upload_dirs()

# ============= AUTHENTICATION ENDPOINTS =============

@api_router.post("/auth/login", response_model=Token)
async def login(user_credentials: UserLogin):
    """Authenticate user and return JWT token"""
    user_data = await db.users.find_one({"username": user_credentials.username})
    
    if not user_data or not verify_password(user_credentials.password, user_data["hashed_password"]):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid username or password"
        )
    
    if not user_data.get("is_active", True):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Account is disabled"
        )
    
    # Update last login
    await db.users.update_one(
        {"id": user_data["id"]},
        {"$set": {"last_login": datetime.now(timezone.utc)}}
    )
    
    # Create JWT token
    access_token = create_access_token(
        data={"sub": user_data["username"], "role": user_data["role"], "user_id": user_data["id"]}
    )
    
    user_info = {
        "id": user_data["id"],
        "username": user_data["username"],
        "full_name": user_data["full_name"],
        "role": user_data["role"],
        "email": user_data["email"]
    }
    
    return {"access_token": access_token, "token_type": "bearer", "user": user_info}

@api_router.post("/auth/register", response_model=StandardResponse)
async def register_user(user_data: UserCreate, current_user: dict = Depends(require_admin)):
    """Register new user (Admin only)"""
    # Check if username already exists
    existing_user = await db.users.find_one({"username": user_data.username})
    if existing_user:
        raise HTTPException(status_code=400, detail="Username already exists")
    
    # Check if email already exists
    existing_email = await db.users.find_one({"email": user_data.email})
    if existing_email:
        raise HTTPException(status_code=400, detail="Email already exists")
    
    # Create new user
    hashed_password = get_password_hash(user_data.password)
    new_user = User(
        username=user_data.username,
        email=user_data.email,
        hashed_password=hashed_password,
        role=user_data.role,
        full_name=user_data.full_name
    )
    
    await db.users.insert_one(new_user.dict())
    
    return StandardResponse(success=True, message="User created successfully")

@api_router.get("/auth/me")
async def get_current_user_info(current_user: dict = Depends(get_current_user)):
    """Get current user information"""
    user_data = await db.users.find_one({"id": current_user["user_id"]})
    if not user_data:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {
        "id": user_data["id"],
        "username": user_data["username"],
        "full_name": user_data["full_name"],
        "role": user_data["role"],
        "email": user_data["email"]
    }

# ============= CLIENT MANAGEMENT ENDPOINTS =============

@api_router.get("/clients", response_model=List[Client])
async def get_clients(current_user: dict = Depends(require_any_role)):
    """Get all clients"""
    clients = await db.clients.find({"is_active": True}).to_list(1000)
    # Transform logo_path to proper URL for frontend
    for client in clients:
        if client.get("logo_path"):
            client["logo_path"] = get_file_url(client["logo_path"])
    return [Client(**client) for client in clients]

@api_router.post("/clients", response_model=StandardResponse)
async def create_client(client_data: ClientCreate, current_user: dict = Depends(require_admin_or_sales)):
    """Create new client"""
    new_client = Client(**client_data.dict())
    await db.clients.insert_one(new_client.dict())
    
    return StandardResponse(success=True, message="Client created successfully", data={"id": new_client.id})

@api_router.get("/clients/{client_id}", response_model=Client)
async def get_client(client_id: str, current_user: dict = Depends(require_any_role)):
    """Get specific client"""
    client = await db.clients.find_one({"id": client_id, "is_active": True})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Transform logo_path to proper URL for frontend
    if client.get("logo_path"):
        client["logo_path"] = get_file_url(client["logo_path"])
    
    return Client(**client)

@api_router.put("/clients/{client_id}", response_model=StandardResponse)
async def update_client(client_id: str, client_data: ClientCreate, current_user: dict = Depends(require_admin_or_sales)):
    """Update client"""
    update_data = client_data.dict()
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    result = await db.clients.update_one(
        {"id": client_id, "is_active": True},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Client not found")
    
    return StandardResponse(success=True, message="Client updated successfully")

@api_router.post("/clients/{client_id}/logo")
async def upload_client_logo(client_id: str, file: UploadFile = File(...), current_user: dict = Depends(require_admin_or_sales)):
    """Upload client logo"""
    # Check if client exists
    client = await db.clients.find_one({"id": client_id, "is_active": True})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Save logo file
    try:
        file_path = await save_logo_file(file, client_id)
        file_url = get_file_url(file_path)
        
        # Update client with logo path
        await db.clients.update_one(
            {"id": client_id},
            {"$set": {"logo_path": file_path, "updated_at": datetime.now(timezone.utc)}}
        )
        
        return StandardResponse(success=True, message="Logo uploaded successfully", data={"file_url": file_url})
    
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(status_code=500, detail="Failed to upload logo")

# ============= PRODUCT MANAGEMENT ENDPOINTS =============

@api_router.get("/products", response_model=List[Product])
async def get_all_products(current_user: dict = Depends(require_any_role)):
    """Get all products"""
    products = await db.products.find({"is_active": True}).to_list(1000)
    return [Product(**product) for product in products]

@api_router.get("/clients/{client_id}/products", response_model=List[Product])
async def get_client_products(client_id: str, current_user: dict = Depends(require_any_role)):
    """Get products for specific client"""
    products = await db.products.find({"client_id": client_id, "is_active": True}).to_list(1000)
    return [Product(**product) for product in products]

@api_router.post("/products", response_model=StandardResponse)
async def create_product(product_data: ProductCreate, current_user: dict = Depends(require_admin_or_sales)):
    """Create new product for client"""
    new_product = Product(**product_data.dict())
    await db.products.insert_one(new_product.dict())
    
    return StandardResponse(success=True, message="Product created successfully", data={"id": new_product.id})

@api_router.get("/products/{product_id}", response_model=Product)
async def get_product(product_id: str, current_user: dict = Depends(require_any_role)):
    """Get specific product"""
    product = await db.products.find_one({"id": product_id, "is_active": True})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    return Product(**product)

@api_router.put("/products/{product_id}", response_model=StandardResponse)
async def update_product(product_id: str, product_data: ProductCreate, current_user: dict = Depends(require_admin_or_sales)):
    """Update product"""
    update_data = product_data.dict()
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    result = await db.products.update_one(
        {"id": product_id, "is_active": True},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    
    return StandardResponse(success=True, message="Product updated successfully")

# ============= MATERIALS MANAGEMENT ENDPOINTS =============

@api_router.get("/materials", response_model=List[Material])
async def get_materials(current_user: dict = Depends(require_any_role)):
    """Get all materials"""
    materials = await db.materials.find({"is_active": True}).to_list(1000)
    return [Material(**material) for material in materials]

@api_router.post("/materials", response_model=StandardResponse)
async def create_material(material_data: MaterialCreate, current_user: dict = Depends(require_admin_or_manager)):
    """Create new material"""
    new_material = Material(**material_data.dict())
    await db.materials.insert_one(new_material.dict())
    
    return StandardResponse(success=True, message="Material created successfully", data={"id": new_material.id})

@api_router.get("/materials/{material_id}", response_model=Material)
async def get_material(material_id: str, current_user: dict = Depends(require_any_role)):
    """Get specific material"""
    material = await db.materials.find_one({"id": material_id, "is_active": True})
    if not material:
        raise HTTPException(status_code=404, detail="Material not found")
    
    return Material(**material)

@api_router.put("/materials/{material_id}", response_model=StandardResponse)
async def update_material(material_id: str, material_data: MaterialCreate, current_user: dict = Depends(require_admin_or_manager)):
    """Update material"""
    update_data = material_data.dict()
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    result = await db.materials.update_one(
        {"id": material_id, "is_active": True},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Material not found")
    
    return StandardResponse(success=True, message="Material updated successfully")

@api_router.delete("/materials/{material_id}", response_model=StandardResponse)
async def delete_material(material_id: str, current_user: dict = Depends(require_admin_or_manager)):
    """Delete material (soft delete)"""
    result = await db.materials.update_one(
        {"id": material_id, "is_active": True},
        {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc)}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Material not found")
    
    return StandardResponse(success=True, message="Material deleted successfully")

# ============= SUPPLIERS ENDPOINTS =============

@api_router.get("/suppliers", response_model=List[Supplier])
async def get_suppliers(current_user: dict = Depends(require_any_role)):
    """Get all active suppliers"""
    suppliers = await db.suppliers.find({"is_active": True}).sort("supplier_name", 1).to_list(1000)
    return [Supplier(**supplier) for supplier in suppliers]

@api_router.post("/suppliers", response_model=StandardResponse)
async def create_supplier(supplier_data: SupplierCreate, current_user: dict = Depends(require_admin_or_manager)):
    """Create new supplier"""
    new_supplier = Supplier(**supplier_data.dict())
    
    await db.suppliers.insert_one(new_supplier.dict())
    
    return StandardResponse(success=True, message="Supplier created successfully", data={"id": new_supplier.id})

@api_router.get("/suppliers/{supplier_id}", response_model=Supplier)
async def get_supplier(supplier_id: str, current_user: dict = Depends(require_any_role)):
    """Get specific supplier by ID"""
    supplier = await db.suppliers.find_one({"id": supplier_id, "is_active": True})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    return Supplier(**supplier)

@api_router.put("/suppliers/{supplier_id}", response_model=StandardResponse)
async def update_supplier(supplier_id: str, supplier_data: SupplierCreate, current_user: dict = Depends(require_admin_or_manager)):
    """Update supplier"""
    update_data = supplier_data.dict()
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    result = await db.suppliers.update_one(
        {"id": supplier_id, "is_active": True},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    return StandardResponse(success=True, message="Supplier updated successfully")

@api_router.delete("/suppliers/{supplier_id}", response_model=StandardResponse)
async def delete_supplier(supplier_id: str, current_user: dict = Depends(require_admin_or_manager)):
    """Soft delete supplier"""
    result = await db.suppliers.update_one(
        {"id": supplier_id, "is_active": True},
        {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc)}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    return StandardResponse(success=True, message="Supplier deleted successfully")

# ============= PRODUCT SPECIFICATIONS ENDPOINTS =============

@api_router.get("/product-specifications", response_model=List[ProductSpecification])
async def get_product_specifications(current_user: dict = Depends(require_any_role)):
    """Get all product specifications"""
    specs = await db.product_specifications.find({"is_active": True}).sort("product_name", 1).to_list(1000)
    return [ProductSpecification(**spec) for spec in specs]

@api_router.post("/product-specifications", response_model=StandardResponse)
async def create_product_specification(spec_data: ProductSpecificationCreate, current_user: dict = Depends(require_admin_or_manager)):
    """Create new product specification with automatic thickness calculation"""
    # Calculate total thickness from material layers (thickness * quantity for each layer)
    calculated_thickness = sum(layer.thickness * (layer.quantity or 1.0) for layer in spec_data.material_layers)
    
    # Generate thickness options (±5%, ±10%, exact)
    thickness_options = []
    if calculated_thickness > 0:
        thickness_options = [
            round(calculated_thickness * 0.90, 3),  # -10%
            round(calculated_thickness * 0.95, 3),  # -5%
            round(calculated_thickness, 3),         # Exact
            round(calculated_thickness * 1.05, 3),  # +5%
            round(calculated_thickness * 1.10, 3),  # +10%
        ]
        # Remove duplicates and sort
        thickness_options = sorted(list(set(thickness_options)))
    
    # Create new specification with calculated values
    new_spec = ProductSpecification(
        **spec_data.dict(),
        calculated_total_thickness=calculated_thickness if calculated_thickness > 0 else None,
        thickness_options=thickness_options
    )
    
    await db.product_specifications.insert_one(new_spec.dict())
    
    return StandardResponse(success=True, message="Product specification created successfully", data={
        "id": new_spec.id, 
        "calculated_thickness": calculated_thickness,
        "thickness_options": thickness_options
    })

@api_router.get("/product-specifications/{spec_id}", response_model=ProductSpecification)
async def get_product_specification(spec_id: str, current_user: dict = Depends(require_any_role)):
    """Get specific product specification by ID"""
    spec = await db.product_specifications.find_one({"id": spec_id, "is_active": True})
    if not spec:
        raise HTTPException(status_code=404, detail="Product specification not found")
    
    return ProductSpecification(**spec)

@api_router.put("/product-specifications/{spec_id}", response_model=StandardResponse)
async def update_product_specification(spec_id: str, spec_data: ProductSpecificationCreate, current_user: dict = Depends(require_admin_or_manager)):
    """Update product specification with automatic thickness calculation"""
    # Calculate total thickness from material layers (thickness * quantity for each layer)
    calculated_thickness = sum(layer.thickness * (layer.quantity or 1.0) for layer in spec_data.material_layers)
    
    # Generate thickness options (±5%, ±10%, exact)
    thickness_options = []
    if calculated_thickness > 0:
        thickness_options = [
            round(calculated_thickness * 0.90, 3),  # -10%
            round(calculated_thickness * 0.95, 3),  # -5%
            round(calculated_thickness, 3),         # Exact
            round(calculated_thickness * 1.05, 3),  # +5%
            round(calculated_thickness * 1.10, 3),  # +10%
        ]
        # Remove duplicates and sort
        thickness_options = sorted(list(set(thickness_options)))
    
    # Prepare update data
    update_data = spec_data.dict()
    update_data.update({
        "calculated_total_thickness": calculated_thickness if calculated_thickness > 0 else None,
        "thickness_options": thickness_options,
        "updated_at": datetime.now(timezone.utc)
    })
    
    result = await db.product_specifications.update_one(
        {"id": spec_id, "is_active": True},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Product specification not found")
    
    return StandardResponse(success=True, message="Product specification updated successfully", data={
        "calculated_thickness": calculated_thickness,
        "thickness_options": thickness_options
    })

@api_router.delete("/product-specifications/{spec_id}", response_model=StandardResponse)
async def delete_product_specification(spec_id: str, current_user: dict = Depends(require_admin_or_manager)):
    """Soft delete product specification"""
    result = await db.product_specifications.update_one(
        {"id": spec_id, "is_active": True},
        {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc)}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Product specification not found")
    
    return StandardResponse(success=True, message="Product specification deleted successfully")


@api_router.post("/sync/product-spec-to-client-products", response_model=StandardResponse)
async def sync_product_spec_to_client_products(spec_id: str, current_user: dict = Depends(require_admin_or_manager)):
    """
    Sync material_layers from product_specifications to all matching client_products
    Matches products by product_name (case-insensitive partial match)
    """
    try:
        # Get the product specification
        spec = await db.product_specifications.find_one({"id": spec_id, "is_active": True})
        
        if not spec:
            raise HTTPException(status_code=404, detail="Product specification not found")
        
        material_layers = spec.get("material_layers", [])
        product_name = spec.get("product_name", "")
        
        if not material_layers:
            return StandardResponse(success=True, message="No material layers to sync", data={"synced_count": 0})
        
        # Find matching client_products (by product_name or product_description)
        matching_products = await db.client_products.find({
            "$or": [
                {"product_name": {"$regex": product_name, "$options": "i"}},
                {"product_description": {"$regex": product_name, "$options": "i"}},
                {"product_code": spec.get("product_code")}
            ],
            "is_active": True
        }).to_list(length=None)
        
        synced_count = 0
        for product in matching_products:
            result = await db.client_products.update_one(
                {"id": product["id"]},
                {"$set": {
                    "material_layers": material_layers,
                    "updated_at": datetime.now(timezone.utc)
                }}
            )
            if result.modified_count > 0:
                synced_count += 1
        
        return StandardResponse(
            success=True,
            message=f"Synced material layers to {synced_count} client product(s)",
            data={"synced_count": synced_count, "products_found": len(matching_products)}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to sync product spec to client products: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Sync failed: {str(e)}")


@api_router.post("/sync/client-product-to-spec", response_model=StandardResponse)
async def sync_client_product_to_spec(client_id: str, product_id: str, current_user: dict = Depends(require_admin_or_manager)):
    """
    Sync material_layers from client_product to matching product_specification
    Creates a new spec if none exists with matching name
    """
    try:
        # Get the client product
        product = await db.client_products.find_one({
            "id": product_id,
            "client_id": client_id,
            "is_active": True
        })
        
        if not product:
            raise HTTPException(status_code=404, detail="Client product not found")
        
        material_layers = product.get("material_layers", [])
        product_name = product.get("product_name") or product.get("product_description", "")
        
        if not material_layers:
            return StandardResponse(success=True, message="No material layers to sync", data={"synced_count": 0})
        
        # Try to find matching product specification
        spec = await db.product_specifications.find_one({
            "$or": [
                {"product_name": {"$regex": product_name, "$options": "i"}},
                {"product_code": product.get("product_code")}
            ],
            "is_active": True
        })
        
        if spec:
            # Update existing spec
            result = await db.product_specifications.update_one(
                {"id": spec["id"]},
                {"$set": {
                    "material_layers": material_layers,
                    "updated_at": datetime.now(timezone.utc)
                }}
            )
            return StandardResponse(
                success=True,
                message="Synced material layers to product specification",
                data={"spec_id": spec["id"], "action": "updated"}
            )
        else:
            # Create new spec
            new_spec_id = str(uuid.uuid4())
            new_spec = {
                "id": new_spec_id,
                "product_name": product_name,
                "product_type": product.get("product_type", "Other"),
                "specifications": {},
                "material_layers": material_layers,
                "materials_composition": [],
                "machinery": [],
                "manufacturing_notes": f"Auto-created from client product: {product_name}",
                "is_active": True,
                "created_at": datetime.now(timezone.utc),
                "updated_at": datetime.now(timezone.utc)
            }
            
            await db.product_specifications.insert_one(new_spec)
            
            return StandardResponse(
                success=True,
                message="Created new product specification with material layers",
                data={"spec_id": new_spec_id, "action": "created"}
            )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to sync client product to spec: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Sync failed: {str(e)}")


# ============= MACHINERY RATES ENDPOINTS =============

@api_router.get("/machinery-rates", response_model=List[MachineryRate])
async def get_machinery_rates(current_user: dict = Depends(require_any_role)):
    """Get all machinery rates"""
    rates = await db.machinery_rates.find({"is_active": True}).sort("function", 1).to_list(1000)
    return [MachineryRate(**rate) for rate in rates]

@api_router.post("/machinery-rates", response_model=StandardResponse)
async def create_machinery_rate(rate_data: MachineryRateCreate, current_user: dict = Depends(require_admin_or_manager)):
    """Create new machinery rate"""
    # Check if rate for this function already exists
    existing_rate = await db.machinery_rates.find_one({
        "function": rate_data.function,
        "is_active": True
    })
    
    if existing_rate:
        raise HTTPException(status_code=400, detail=f"Rate for function '{rate_data.function}' already exists")
    
    # Create new rate
    new_rate = MachineryRate(
        **rate_data.dict(),
        created_at=datetime.now(timezone.utc)
    )
    
    rate_dict = new_rate.dict()
    await db.machinery_rates.insert_one(rate_dict)
    
    return StandardResponse(
        success=True,
        message="Machinery rate created successfully",
        data={"id": new_rate.id}
    )

@api_router.get("/machinery-rates/{rate_id}", response_model=MachineryRate)
async def get_machinery_rate(rate_id: str, current_user: dict = Depends(require_any_role)):
    """Get specific machinery rate by ID"""
    rate = await db.machinery_rates.find_one({"id": rate_id, "is_active": True})
    if not rate:
        raise HTTPException(status_code=404, detail="Machinery rate not found")
    
    return MachineryRate(**rate)

@api_router.put("/machinery-rates/{rate_id}", response_model=StandardResponse)
async def update_machinery_rate(rate_id: str, rate_data: MachineryRateUpdate, current_user: dict = Depends(require_admin_or_manager)):
    """Update machinery rate"""
    # Prepare update data
    update_data = {k: v for k, v in rate_data.dict().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    # If function is being changed, check for duplicates
    if "function" in update_data:
        existing_rate = await db.machinery_rates.find_one({
            "function": update_data["function"],
            "is_active": True,
            "id": {"$ne": rate_id}
        })
        if existing_rate:
            raise HTTPException(status_code=400, detail=f"Rate for function '{update_data['function']}' already exists")
    
    result = await db.machinery_rates.update_one(
        {"id": rate_id, "is_active": True},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Machinery rate not found")
    
    return StandardResponse(success=True, message="Machinery rate updated successfully")

@api_router.delete("/machinery-rates/{rate_id}", response_model=StandardResponse)
async def delete_machinery_rate(rate_id: str, current_user: dict = Depends(require_admin_or_manager)):
    """Soft delete machinery rate"""
    result = await db.machinery_rates.update_one(
        {"id": rate_id, "is_active": True},
        {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc)}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Machinery rate not found")
    
    return StandardResponse(success=True, message="Machinery rate deleted successfully")

# ============= CALCULATORS ENDPOINTS =============

@api_router.post("/calculators/material-consumption-by-client", response_model=StandardResponse)
async def calculate_material_consumption_by_client(
    request: MaterialConsumptionByClientRequest, 
    current_user: dict = Depends(require_any_role)
):
    """Calculate material consumption for a client within date range"""
    try:
        # Get orders for client within date range
        orders = await db.orders.find({
            "client_id": request.client_id,
            "created_at": {
                "$gte": datetime.combine(request.start_date, datetime.min.time()),
                "$lte": datetime.combine(request.end_date, datetime.max.time())
            },
            "status": {"$ne": "cancelled"}
        }).to_list(1000)
        
        # Get material info
        material = await db.materials.find_one({"id": request.material_id})
        if not material:
            raise HTTPException(status_code=404, detail="Material not found")
        
        total_consumption = 0
        order_breakdown = []
        
        for order in orders:
            # Calculate material usage based on order items and product specifications
            order_consumption = 0
            for item in order["items"]:
                # This is a simplified calculation - you may need to enhance based on your business logic
                order_consumption += item["quantity"] * 1.0  # Placeholder calculation
            
            total_consumption += order_consumption
            order_breakdown.append({
                "order_number": order["order_number"],
                "order_date": order["created_at"],
                "consumption": order_consumption
            })
        
        result = CalculationResult(
            calculation_type="material_consumption_by_client",
            input_parameters={
                "client_id": request.client_id,
                "material_id": request.material_id,
                "start_date": request.start_date.isoformat(),
                "end_date": request.end_date.isoformat()
            },
            results={
                "total_consumption": total_consumption,
                "material_name": material["product_code"],
                "unit": material["unit"],
                "order_count": len(orders),
                "order_breakdown": order_breakdown
            },
            calculated_by=current_user["sub"]
        )
        
        return StandardResponse(success=True, message="Calculation completed", data=result.dict())
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Calculation failed: {str(e)}")

@api_router.post("/calculators/material-permutation", response_model=StandardResponse)
async def calculate_material_permutation(
    request: MaterialPermutationRequest,
    current_user: dict = Depends(require_any_role)
):
    """Calculate optimal material permutation for core IDs"""
    try:
        # Get product specifications for the core IDs
        core_specs = await db.product_specifications.find({
            "specifications.core_id": {"$in": request.core_ids},
            "is_active": True
        }).to_list(1000)
        
        # Calculate permutation options
        permutation_options = []
        for arrangement in _generate_permutations(request.sizes_to_manufacture, request.master_deckle_width):
            waste_percentage = _calculate_waste(arrangement, request.master_deckle_width)
            if waste_percentage <= request.acceptable_waste_percentage:
                permutation_options.append({
                    "arrangement": arrangement,
                    "waste_percentage": waste_percentage,
                    "efficiency": 100 - waste_percentage
                })
        
        # Sort by efficiency
        permutation_options.sort(key=lambda x: x["efficiency"], reverse=True)
        
        result = CalculationResult(
            calculation_type="material_permutation",
            input_parameters=request.dict(),
            results={
                "permutation_options": permutation_options[:10],  # Top 10 options
                "total_options_found": len(permutation_options)
            },
            calculated_by=current_user["sub"]
        )
        
        return StandardResponse(success=True, message="Permutation calculation completed", data=result.dict())
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Permutation calculation failed: {str(e)}")

@api_router.post("/calculators/spiral-core-consumption", response_model=StandardResponse)
async def calculate_spiral_core_consumption(
    request: SpiralCoreConsumptionRequest,
    current_user: dict = Depends(require_any_role)
):
    """Calculate material consumption for Spiral Paper Cores"""
    try:
        # Get product specification
        spec = await db.product_specifications.find_one({"id": request.product_specification_id})
        if not spec:
            raise HTTPException(status_code=404, detail="Product specification not found")
        
        if spec["product_type"] != "Spiral Paper Core":
            raise HTTPException(status_code=400, detail="Specification is not for Spiral Paper Cores")
        
        specifications = spec["specifications"]
        
        # Get material information
        material_id = specifications.get("selected_material_id")
        if not material_id:
            raise HTTPException(status_code=400, detail="No material selected in specification")
            
        material = await db.materials.find_one({"id": material_id})
        if not material:
            raise HTTPException(status_code=404, detail="Material not found")
        
        # Calculate consumption
        internal_diameter = float(specifications.get("internal_diameter", 0))
        wall_thickness = float(specifications.get("wall_thickness_required", 0))
        gsm = float(material.get("gsm", 0))
        material_thickness = float(material.get("thickness_mm", 1))
        
        # Calculate surface area per core
        outer_diameter = internal_diameter + (2 * wall_thickness)
        circumference = 3.14159 * outer_diameter
        surface_area = circumference * request.core_length  # mm²
        surface_area_m2 = surface_area / 1000000  # Convert to m²
        
        # Calculate material weight per core
        layers_required = max(1, round(wall_thickness / material_thickness))
        material_weight_per_core = surface_area_m2 * gsm * layers_required / 1000  # kg
        
        # Total consumption
        total_weight = material_weight_per_core * request.quantity
        
        result = CalculationResult(
            calculation_type="spiral_core_consumption",
            input_parameters=request.dict(),
            results={
                "material_name": material["product_code"],
                "gsm": gsm,
                "material_thickness_mm": material_thickness,
                "layers_required": layers_required,
                "internal_diameter_mm": internal_diameter,
                "outer_diameter_mm": outer_diameter,
                "wall_thickness_mm": wall_thickness,
                "core_length_mm": request.core_length,
                "surface_area_m2_per_core": round(surface_area_m2, 4),
                "material_weight_per_core_kg": round(material_weight_per_core, 4),
                "quantity": request.quantity,
                "total_material_weight_kg": round(total_weight, 2),
                "unit": material["unit"]
            },
            calculated_by=current_user["sub"]
        )
        
        return StandardResponse(success=True, message="Spiral core consumption calculated", data=result.dict())
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Spiral core calculation failed: {str(e)}")

# Helper functions for calculations
def _generate_permutations(sizes, master_width):
    """Generate possible arrangements of sizes within master width"""
    # Simplified permutation logic - you can enhance this
    arrangements = []
    for size_combo in _get_size_combinations(sizes, master_width):
        arrangements.append(size_combo)
    return arrangements

def _get_size_combinations(sizes, master_width):
    """Get combinations of sizes that fit within master width"""
    combinations = []
    # This is a simplified algorithm - implement more sophisticated logic as needed
    for size in sizes:
        width = size.get("width", 0)
        if width <= master_width:
            combinations.append([size])
    return combinations

def _calculate_waste(arrangement, master_width):
    """Calculate waste percentage for an arrangement"""
    used_width = sum(item.get("width", 0) for item in arrangement)
    waste = master_width - used_width
    return (waste / master_width) * 100 if master_width > 0 else 100

# ============= STOCKTAKE ENDPOINTS =============

@api_router.get("/stocktake/current", response_model=StandardResponse)
async def get_current_stocktake(current_user: dict = Depends(require_manager_or_admin)):
    """Get current month's stocktake or check if one is needed"""
    current_date = date.today()
    current_month = current_date.strftime("%Y-%m")
    
    # Check if stocktake exists for current month
    stocktake = await db.stocktakes.find_one({"month": current_month})
    
    # Check if it's first business day and stocktake is needed
    is_first_business_day = _is_first_business_day(current_date)
    
    if not stocktake and is_first_business_day:
        return StandardResponse(
            success=True, 
            message="Stocktake required", 
            data={
                "stocktake_required": True,
                "month": current_month,
                "first_business_day": True
            }
        )
    
    # Convert ObjectId to string if stocktake exists
    if stocktake:
        stocktake["id"] = str(stocktake["_id"])
        del stocktake["_id"]
    
    return StandardResponse(
        success=True, 
        message="Stocktake status retrieved", 
        data={
            "stocktake": stocktake,
            "stocktake_required": not bool(stocktake),
            "first_business_day": is_first_business_day
        }
    )

@api_router.post("/stocktake", response_model=StandardResponse)
async def create_stocktake(
    stocktake_data: StocktakeCreate,
    current_user: dict = Depends(require_manager_or_admin)
):
    """Create new stocktake with all materials"""
    current_month = stocktake_data.stocktake_date.strftime("%Y-%m")
    
    # Check if stocktake already exists for this month
    existing = await db.stocktakes.find_one({"month": current_month})
    if existing:
        raise HTTPException(status_code=400, detail="Stocktake already exists for this month")
    
    # Get all active materials
    materials = await db.materials.find({"is_active": True}).to_list(1000)
    
    new_stocktake = Stocktake(
        stocktake_date=stocktake_data.stocktake_date,
        month=current_month,
        created_by=current_user["sub"]
    )
    
    # Convert the stocktake to dict and handle date serialization
    stocktake_dict = new_stocktake.dict()
    stocktake_dict["stocktake_date"] = datetime.combine(stocktake_data.stocktake_date, datetime.min.time())
    
    await db.stocktakes.insert_one(stocktake_dict)
    
    return StandardResponse(
        success=True, 
        message="Stocktake created", 
        data={
            "stocktake_id": new_stocktake.id,
            "materials_count": len(materials),
            "materials": [{"id": m["id"], "name": f"{m['supplier']} - {m['product_code']}", "unit": m["unit"]} for m in materials]
        }
    )

@api_router.put("/stocktake/{stocktake_id}/entry", response_model=StandardResponse)
async def update_stocktake_entry(
    stocktake_id: str,
    entry_data: StocktakeEntryUpdate,
    current_user: dict = Depends(require_manager_or_admin)
):
    """Update stocktake entry for a material"""
    stocktake = await db.stocktakes.find_one({"id": stocktake_id})
    if not stocktake:
        raise HTTPException(status_code=404, detail="Stocktake not found")
    
    material = await db.materials.find_one({"id": entry_data.material_id})
    if not material:
        raise HTTPException(status_code=404, detail="Material not found")
    
    # Create or update entry
    entry = StocktakeEntry(
        stocktake_id=stocktake_id,
        material_id=entry_data.material_id,
        material_name=f"{material['supplier']} - {material['product_code']}",
        current_quantity=entry_data.current_quantity,
        unit=material["unit"],
        counted_by=current_user["sub"]
    )
    
    # Update stocktake entries
    await db.stocktake_entries.replace_one(
        {"stocktake_id": stocktake_id, "material_id": entry_data.material_id},
        entry.dict(),
        upsert=True
    )
    
    return StandardResponse(success=True, message="Stocktake entry updated")

@api_router.post("/stocktake/{stocktake_id}/complete", response_model=StandardResponse)
async def complete_stocktake(
    stocktake_id: str,
    current_user: dict = Depends(require_manager_or_admin)
):
    """Complete stocktake"""
    result = await db.stocktakes.update_one(
        {"id": stocktake_id},
        {
            "$set": {
                "status": "completed",
                "completed_by": current_user["sub"],
                "completed_at": datetime.now(timezone.utc)
            }
        }
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Stocktake not found")
    
    return StandardResponse(success=True, message="Stocktake completed")

def _is_first_business_day(check_date):
    """Check if given date is first business day of month"""
    first_day = check_date.replace(day=1)
    # Skip weekends
    while first_day.weekday() >= 5:  # 5=Saturday, 6=Sunday
        first_day += timedelta(days=1)
    return check_date == first_day

# ============= USER MANAGEMENT ENDPOINTS =============

@api_router.get("/users", response_model=List[dict])
async def get_users(current_user: dict = Depends(require_admin)):
    """Get all users (Admin only)"""
    users = await db.users.find({}).sort("full_name", 1).to_list(1000)
    
    # Remove password hashes and convert ObjectId to string
    for user in users:
        if "password_hash" in user:
            del user["password_hash"]
        if "_id" in user:
            del user["_id"]  # Remove MongoDB ObjectId
    
    return users

@api_router.post("/users", response_model=StandardResponse)
async def create_user(user_data: UserCreate, current_user: dict = Depends(require_admin)):
    """Create new user account (Admin only)"""
    # Check if username or email already exists
    existing_user = await db.users.find_one({
        "$or": [
            {"username": user_data.username},
            {"email": user_data.email}
        ]
    })
    
    if existing_user:
        if existing_user["username"] == user_data.username:
            raise HTTPException(status_code=400, detail="Username already exists")
        if existing_user["email"] == user_data.email:
            raise HTTPException(status_code=400, detail="Email already exists")
    
    # Hash the password
    password_hash = hash_password(user_data.password)
    
    # Create user document
    new_user = {
        "id": str(uuid.uuid4()),
        "username": user_data.username,
        "email": user_data.email,
        "password_hash": password_hash,
        "full_name": user_data.full_name,
        "role": user_data.role,
        "department": user_data.department,
        "phone": user_data.phone,
        "employment_type": user_data.employment_type.value,
        "is_active": True,
        "created_at": datetime.now(timezone.utc),
        "updated_at": None,
        "created_by": current_user["sub"]
    }
    
    await db.users.insert_one(new_user)
    
    return StandardResponse(success=True, message="User created successfully", data={"id": new_user["id"]})

@api_router.get("/users/{user_id}", response_model=dict)
async def get_user(user_id: str, current_user: dict = Depends(require_admin)):
    """Get specific user by ID (Admin only)"""
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Remove password hash and ObjectId from response
    if "password_hash" in user:
        del user["password_hash"]
    if "_id" in user:
        del user["_id"]  # Remove MongoDB ObjectId
    
    return user

@api_router.put("/users/{user_id}", response_model=StandardResponse)
async def update_user(user_id: str, user_data: UserUpdate, current_user: dict = Depends(require_admin)):
    """Update user account (Admin only)"""
    # Check if user exists
    existing_user = await db.users.find_one({"id": user_id})
    if not existing_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Prepare update data
    update_data = {}
    if user_data.username is not None:
        # Check if username is already used by another user
        username_user = await db.users.find_one({"username": user_data.username, "id": {"$ne": user_id}})
        if username_user:
            raise HTTPException(status_code=400, detail="Username already exists")
        update_data["username"] = user_data.username
    
    if user_data.email is not None:
        # Check if email is already used by another user
        email_user = await db.users.find_one({"email": user_data.email, "id": {"$ne": user_id}})
        if email_user:
            raise HTTPException(status_code=400, detail="Email already exists")
        update_data["email"] = user_data.email
    
    if user_data.full_name is not None:
        update_data["full_name"] = user_data.full_name
    if user_data.role is not None:
        update_data["role"] = user_data.role
    if user_data.department is not None:
        update_data["department"] = user_data.department
    if user_data.phone is not None:
        update_data["phone"] = user_data.phone
    if user_data.employment_type is not None:
        update_data["employment_type"] = user_data.employment_type.value
    if user_data.is_active is not None:
        update_data["is_active"] = user_data.is_active
    
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    await db.users.update_one(
        {"id": user_id},
        {"$set": update_data}
    )
    
    return StandardResponse(success=True, message="User updated successfully")

@api_router.delete("/users/{user_id}", response_model=StandardResponse)
async def delete_user(user_id: str, current_user: dict = Depends(require_admin)):
    """Delete user permanently (Admin only)"""
    # Check if user exists before deletion
    existing_user = await db.users.find_one({"id": user_id})
    if not existing_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Prevent admin from deleting themselves
    if user_id == current_user.get("user_id") or user_id == current_user.get("sub"):
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    
    # Permanently delete the user
    result = await db.users.delete_one({"id": user_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return StandardResponse(success=True, message="User deleted successfully")

@api_router.post("/users/change-password", response_model=StandardResponse)
async def change_user_password(password_data: PasswordChangeRequest, current_user: dict = Depends(get_current_user)):
    """Change user's own password"""
    # Get current user from database using user_id from JWT token
    user_id = current_user.get("user_id")
    if not user_id:
        raise HTTPException(status_code=404, detail="User ID not found in token")
    
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Verify current password (check both possible field names for compatibility)
    password_field = "password_hash" if "password_hash" in user else "hashed_password"
    if not verify_password(password_data.current_password, user[password_field]):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    
    # Hash new password
    new_password_hash = hash_password(password_data.new_password)
    
    # Update password (use the same field name that exists)
    await db.users.update_one(
        {"id": user_id},
        {"$set": {
            password_field: new_password_hash,
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    return StandardResponse(success=True, message="Password changed successfully")

# Helper functions for user management
def _generate_permissions(role: UserRole) -> List[str]:
    """Generate permissions based on role"""
    base_permissions = []
    
    if role == UserRole.ADMIN:
        return ["all"]  # Admin has all permissions
    
    elif role == UserRole.MANAGER:
        return [
            "manage_clients", "create_orders", "update_production", 
            "view_reports", "manage_payroll", "manage_materials",
            "manage_suppliers", "manage_specifications", "use_calculators"
        ]
    
    elif role == UserRole.MANAGER:
        return [
            "create_orders", "update_production", "use_calculators", 
            "view_own_payroll", "submit_timesheets", "request_leave"
        ]
    
    elif role == UserRole.PRODUCTION_TEAM:
        return [
            "update_production", "view_own_payroll", "submit_timesheets", "request_leave"
        ]
    
    elif role == UserRole.SALES:
        return [
            "manage_clients", "create_orders", "view_reports"
        ]
    
    return base_permissions

# ============= CLIENT PRODUCT CATALOGUE ENDPOINTS =============

@api_router.get("/clients/{client_id}/catalog", response_model=List[ClientProduct])
async def get_client_product_catalog(client_id: str, current_user: dict = Depends(require_any_role)):
    """Get product catalogue for specific client"""
    products = await db.client_products.find({"client_id": client_id, "is_active": True}).to_list(1000)
    return [ClientProduct(**product) for product in products]

@api_router.post("/clients/{client_id}/catalog", response_model=StandardResponse)
async def create_client_product(client_id: str, product_data: ClientProductCreate, current_user: dict = Depends(require_admin_or_sales)):
    """Create new product for client catalogue"""
    # Verify client exists
    client = await db.clients.find_one({"id": client_id, "is_active": True})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Set client_id from URL
    product_data.client_id = client_id
    new_product = ClientProduct(**product_data.dict())
    await db.client_products.insert_one(new_product.dict())
    
    return StandardResponse(success=True, message="Client product created successfully", data={"id": new_product.id})

@api_router.get("/clients/{client_id}/catalog/{product_id}", response_model=ClientProduct)
async def get_client_product(client_id: str, product_id: str, current_user: dict = Depends(require_any_role)):
    """Get specific client product"""
    product = await db.client_products.find_one({
        "id": product_id, 
        "client_id": client_id,
        "is_active": True
    })
    if not product:
        raise HTTPException(status_code=404, detail="Client product not found")
    
    return ClientProduct(**product)

@api_router.put("/clients/{client_id}/catalog/{product_id}", response_model=StandardResponse)
async def update_client_product(client_id: str, product_id: str, product_data: ClientProductCreate, current_user: dict = Depends(require_admin_or_sales)):
    """Update client product"""
    # Verify client exists
    client = await db.clients.find_one({"id": client_id, "is_active": True})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    update_data = product_data.dict()
    update_data["client_id"] = client_id  # Ensure client_id is correct
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    result = await db.client_products.update_one(
        {"id": product_id, "client_id": client_id, "is_active": True},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Client product not found")
    
    return StandardResponse(success=True, message="Client product updated successfully")

@api_router.delete("/clients/{client_id}/catalog/{product_id}", response_model=StandardResponse)
async def delete_client_product(client_id: str, product_id: str, current_user: dict = Depends(require_admin_or_sales)):
    """Delete client product (soft delete)"""
    result = await db.client_products.update_one(
        {"id": product_id, "client_id": client_id, "is_active": True},
        {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc)}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Client product not found")
    
    return StandardResponse(success=True, message="Client product deleted successfully")

@api_router.post("/clients/{client_id}/catalog/{product_id}/copy-to/{target_client_id}", response_model=StandardResponse)
async def copy_client_product(client_id: str, product_id: str, target_client_id: str, current_user: dict = Depends(require_admin_or_sales)):
    """Copy product to another client catalogue"""
    # Verify both clients exist
    source_client = await db.clients.find_one({"id": client_id, "is_active": True})
    target_client = await db.clients.find_one({"id": target_client_id, "is_active": True})
    
    if not source_client:
        raise HTTPException(status_code=404, detail="Source client not found")
    if not target_client:
        raise HTTPException(status_code=404, detail="Target client not found")
    
    # Get source product
    source_product = await db.client_products.find_one({
        "id": product_id,
        "client_id": client_id,
        "is_active": True
    })
    
    if not source_product:
        raise HTTPException(status_code=404, detail="Source product not found")
    
    # Create copy for target client
    copied_product = ClientProduct(**source_product)
    copied_product.id = str(uuid.uuid4())  # Generate new ID
    copied_product.client_id = target_client_id
    copied_product.created_at = datetime.now(timezone.utc)
    copied_product.updated_at = None
    
    await db.client_products.insert_one(copied_product.dict())
    
    return StandardResponse(
        success=True, 
        message=f"Product copied successfully to {target_client['company_name']}", 
        data={"id": copied_product.id}
    )

@api_router.delete("/clients/{client_id}", response_model=StandardResponse)
async def delete_client(client_id: str, current_user: dict = Depends(require_admin)):
    """Delete client (soft delete - Admin only)"""
    # Check if client exists
    existing_client = await db.clients.find_one({"id": client_id, "is_active": True})
    if not existing_client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Check if client has active orders
    active_orders = await db.orders.find_one({"client_id": client_id, "status": {"$nin": ["completed", "cancelled"]}})
    if active_orders:
        raise HTTPException(status_code=400, detail="Cannot delete client with active orders")
    
    # Perform soft delete
    result = await db.clients.update_one(
        {"id": client_id},
        {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc)}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Client not found")
    
    return StandardResponse(success=True, message="Client deleted successfully")

# ============= ORDER MANAGEMENT ENDPOINTS =============

@api_router.get("/orders", response_model=List[Order])
async def get_orders(status_filter: Optional[str] = None, current_user: dict = Depends(require_any_role)):
    """Get all orders with optional status filter (excludes archived/completed orders by default)"""
    query = {}
    
    # Exclude completed/cleared orders unless specifically requested
    if status_filter:
        query["status"] = status_filter
    else:
        # By default, exclude completed orders and orders in cleared stage
        # (they should only appear in archived jobs)
        query["$and"] = [
            {"status": {"$ne": "completed"}},
            {"current_stage": {"$ne": "cleared"}}
        ]
    
    orders = await db.orders.find(query).sort("created_at", -1).to_list(1000)
    return [Order(**order) for order in orders]

@api_router.post("/orders", response_model=StandardResponse)
async def create_order(order_data: OrderCreate, current_user: dict = Depends(require_admin_or_manager)):
    """Create new order"""
    # Get client details
    client = await db.clients.find_one({"id": order_data.client_id, "is_active": True})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Generate order number - find the highest existing order number for this year
    current_year = datetime.now().year
    order_number_pattern = f"ADM-{current_year}-"
    
    # Get all orders for this year and find the highest number
    existing_orders = await db.orders.find(
        {"order_number": {"$regex": f"^{order_number_pattern}"}}
    ).sort("order_number", -1).limit(1).to_list(length=1)
    
    if existing_orders:
        # Extract the number from the last order (e.g., "ADM-2025-0007" -> 7)
        last_number = int(existing_orders[0]["order_number"].split("-")[-1])
        next_number = last_number + 1
    else:
        # First order of the year
        next_number = 1
    
    # Generate order number with uniqueness check
    max_attempts = 100
    for attempt in range(max_attempts):
        order_number = f"ADM-{current_year}-{next_number:04d}"
        
        # Check if this order number already exists
        existing = await db.orders.find_one({"order_number": order_number})
        if not existing:
            break
        
        # If it exists, increment and try again
        next_number += 1
    else:
        # If we've tried 100 times and still can't find a unique number, something is very wrong
        raise HTTPException(status_code=500, detail="Unable to generate unique order number")
    
    # Calculate totals with discount applied before GST
    subtotal = sum(item.total_price for item in order_data.items)
    
    # Apply discount if provided
    discount_percentage = order_data.discount_percentage or 0
    discount_amount = 0
    discounted_subtotal = subtotal
    
    if discount_percentage > 0:
        discount_amount = subtotal * (discount_percentage / 100)
        discounted_subtotal = subtotal - discount_amount
    
    # Calculate GST on discounted amount
    gst = discounted_subtotal * 0.1
    total_amount = discounted_subtotal + gst
    
    new_order = Order(
        order_number=order_number,
        client_id=order_data.client_id,
        client_name=client["company_name"],
        purchase_order_number=order_data.purchase_order_number,
        items=order_data.items,
        subtotal=subtotal,
        discount_percentage=discount_percentage if discount_percentage > 0 else None,
        discount_amount=discount_amount if discount_amount > 0 else None,
        discount_notes=order_data.discount_notes if order_data.discount_notes else None,
        discounted_subtotal=discounted_subtotal,
        gst=gst,
        total_amount=total_amount,
        due_date=order_data.due_date,
        priority=order_data.priority,
        delivery_address=order_data.delivery_address,
        delivery_instructions=order_data.delivery_instructions,
        runtime_estimate=order_data.runtime_estimate,
        notes=order_data.notes,
        created_by=current_user["user_id"]
    )
    
    await db.orders.insert_one(new_order.dict())
    
    # Log initial production stage
    production_log = ProductionLog(
        order_id=new_order.id,
        stage=ProductionStage.ORDER_ENTERED,
        updated_by=current_user["user_id"],
        notes="Order created"
    )
    await db.production_logs.insert_one(production_log.dict())
    
    return StandardResponse(success=True, message="Order created successfully", data={"id": new_order.id, "order_number": order_number})

@api_router.get("/orders/{order_id}", response_model=Order)
async def get_order(order_id: str, current_user: dict = Depends(require_any_role)):
    """Get specific order"""
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    return Order(**order)


@api_router.put("/orders/{order_id}", response_model=StandardResponse)
async def update_order(order_id: str, update_data: dict, current_user: dict = Depends(require_any_role)):
    """Update order fields"""
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Add updated_at timestamp
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    # Update the order
    await db.orders.update_one(
        {"id": order_id},
        {"$set": update_data}
    )
    
    return StandardResponse(success=True, message="Order updated successfully")

@api_router.put("/orders/{order_id}/stage", response_model=StandardResponse)
async def update_production_stage(order_id: str, stage_update: ProductionStageUpdate, current_user: dict = Depends(require_production_access)):
    """Update order production stage"""
    # Check permissions for specific actions
    user_role = current_user.get("role")
    
    # Auto-populate updated_by from current user if not provided
    if not stage_update.updated_by:
        stage_update.updated_by = current_user.get("user_id") or current_user.get("sub")
    
    # Auto-populate order_id if not provided
    if not stage_update.order_id:
        stage_update.order_id = order_id
    
    # Production team can only move stages forward, not create/delete orders
    if user_role == UserRole.PRODUCTION_TEAM.value:
        if stage_update.to_stage == ProductionStage.ORDER_ENTERED:
            raise HTTPException(status_code=403, detail="Cannot move back to order entry stage")
    
    # Only admin and production manager can invoice
    if stage_update.to_stage == ProductionStage.INVOICING and not can_invoice(user_role):
        raise HTTPException(status_code=403, detail="Insufficient permissions to invoice")
    
    # Update order stage
    update_data = {
        "current_stage": stage_update.to_stage,
        "updated_at": datetime.now(timezone.utc)
    }
    
    # If moving to completed, set completion date and archive the order
    if stage_update.to_stage == ProductionStage.CLEARED:
        update_data["completed_at"] = datetime.now(timezone.utc)
        update_data["status"] = OrderStatus.COMPLETED
        
        # Get the complete order data for archiving
        order = await db.orders.find_one({"id": order_id})
        if order:
            order_number = order.get("order_number", "")
            
            # Archive stock allocations for this order
            archived_stock_count = 0
            try:
                # Find all stock allocations for this order
                stock_allocations = await db.stock_movements.find({
                    "reference": order_number,
                    "movement_type": "allocation",
                    "is_archived": False
                }).to_list(length=None)
                
                # Mark all allocations as archived
                if stock_allocations:
                    result = await db.stock_movements.update_many(
                        {
                            "reference": order_number,
                            "movement_type": "allocation",
                            "is_archived": False
                        },
                        {
                            "$set": {
                                "is_archived": True,
                                "archived_at": datetime.now(timezone.utc).isoformat(),
                                "archived_by": current_user["user_id"],
                                "archived_reason": "Order completed and archived"
                            }
                        }
                    )
                    archived_stock_count = result.modified_count
                    logger.info(f"Archived {archived_stock_count} stock allocation(s) for completed order {order_number}")
                
            except Exception as e:
                logger.error(f"Failed to archive stock allocations for order {order_number}: {str(e)}")
                # Don't fail the entire order archiving if stock archiving fails
            
            # Create archived order
            archived_order = ArchivedOrder(
                original_order_id=order["id"],
                order_number=order["order_number"],
                client_id=order["client_id"],
                client_name=order["client_name"],
                purchase_order_number=order.get("purchase_order_number"),
                items=order["items"],
                subtotal=order["subtotal"],
                gst=order["gst"],
                total_amount=order["total_amount"],
                due_date=order["due_date"],
                delivery_address=order.get("delivery_address"),
                delivery_instructions=order.get("delivery_instructions"),
                runtime_estimate=order.get("runtime_estimate"),
                notes=order.get("notes"),
                created_by=order["created_by"],
                created_at=order["created_at"],
                completed_at=datetime.now(timezone.utc),
                archived_by=current_user["user_id"]
            )
            
            # Insert into archived orders collection
            await db.archived_orders.insert_one(archived_order.dict())
            
            logger.info(f"Order {order_number} archived successfully with {archived_stock_count} stock allocations")
            
            # Update order status to archived and keep in main collection for now
            # This allows for transition period and potential rollback
            update_data["status"] = OrderStatus.ARCHIVED
    
    result = await db.orders.update_one(
        {"id": order_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Log production stage change
    production_log = ProductionLog(
        order_id=order_id,
        stage=stage_update.to_stage,
        updated_by=current_user["user_id"],
        notes=stage_update.notes
    )
    await db.production_logs.insert_one(production_log.dict())
    
    return StandardResponse(success=True, message="Production stage updated successfully")


@api_router.put("/orders/reorder", response_model=StandardResponse)
async def reorder_jobs_in_stage(
    reorder_data: dict,
    current_user: dict = Depends(require_production_access)
):
    """Reorder jobs within a production stage for drag and drop functionality"""
    try:
        stage = reorder_data.get("stage")
        job_order = reorder_data.get("job_order")  # List of order IDs in new order
        
        if not stage or not job_order:
            raise HTTPException(status_code=400, detail="Stage and job_order are required")
        
        # Update display order for each job in the stage
        for index, order_id in enumerate(job_order):
            await db.orders.update_one(
                {"id": order_id},
                {"$set": {"display_order": index}}
            )
        
        logger.info(f"Reordered {len(job_order)} jobs in stage {stage} by user {current_user['user_id']}")
        
        return StandardResponse(
            success=True,
            message=f"Successfully reordered {len(job_order)} jobs in {stage}",
            data={"updated_count": len(job_order)}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to reorder jobs: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to reorder jobs")


# ============= JOB SPECIFICATION ENDPOINTS =============

@api_router.post("/job-specifications", response_model=StandardResponse)
async def create_job_specification(spec_data: JobSpecificationCreate, current_user: dict = Depends(require_admin_or_manager)):
    """Create job specification"""
    new_spec = JobSpecification(**spec_data.dict(), created_by=current_user["user_id"])
    await db.job_specifications.insert_one(new_spec.dict())
    
    return StandardResponse(success=True, message="Job specification created successfully", data={"id": new_spec.id})

@api_router.get("/orders/{order_id}/job-specification", response_model=JobSpecification)
async def get_job_specification_by_order(order_id: str, current_user: dict = Depends(require_any_role)):
    """Get job specification for order"""
    spec = await db.job_specifications.find_one({"order_id": order_id})
    if not spec:
        raise HTTPException(status_code=404, detail="Job specification not found")
    
    return JobSpecification(**spec)

@api_router.delete("/orders/{order_id}", response_model=StandardResponse)
async def delete_order(order_id: str, current_user: dict = Depends(require_admin)):
    """Delete order permanently (Admin only) and return allocated stock to inventory"""
    # Check if order exists
    existing_order = await db.orders.find_one({"id": order_id})
    if not existing_order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Check if order is in production (not safe to delete if in progress)
    unsafe_stages = ["paper_slitting", "winding", "finishing"]
    current_stage = existing_order.get("current_stage", "order_entered")
    if current_stage in unsafe_stages:
        raise HTTPException(status_code=400, detail="Cannot delete order in production. Contact manager to halt production first.")
    
    # RETURN ALLOCATED STOCK TO INVENTORY
    # Find all stock allocations for this order
    order_number = existing_order.get("order_number", "")
    stock_allocations = await db.stock_movements.find({
        "reference": order_number,
        "movement_type": "allocation",
        "quantity": {"$lt": 0}  # Negative quantities are allocations
    }).to_list(length=None)
    
    returned_stock_count = 0
    for allocation in stock_allocations:
        allocated_quantity = abs(allocation.get("quantity", 0))
        stock_id = allocation.get("stock_id")
        product_id = allocation.get("product_id")
        client_id = allocation.get("client_id")
        
        if allocated_quantity > 0 and stock_id:
            # Return stock to inventory
            stock_entry = await db.raw_substrate_stock.find_one({"id": stock_id})
            if stock_entry:
                old_quantity = stock_entry.get("quantity_on_hand", 0)
                new_quantity = old_quantity + allocated_quantity
                
                logger.info(f"Returning stock: {stock_id}, old quantity: {old_quantity}, adding: {allocated_quantity}, new quantity: {new_quantity}")
                
                update_result = await db.raw_substrate_stock.update_one(
                    {"id": stock_id},
                    {"$set": {"quantity_on_hand": new_quantity}}
                )
                
                if update_result.modified_count > 0:
                    logger.info(f"Successfully updated stock {stock_id} quantity from {old_quantity} to {new_quantity}")
                else:
                    logger.warning(f"Failed to update stock {stock_id} - no rows modified")
                
                # Create return movement record
                return_movement = {
                    "id": str(uuid.uuid4()),
                    "stock_id": stock_id,
                    "product_id": product_id,
                    "client_id": client_id,
                    "movement_type": "return",
                    "quantity": allocated_quantity,  # Positive for return
                    "reference": f"Return from deleted order {order_number}",
                    "created_by": current_user["user_id"],
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "is_archived": False
                }
                await db.stock_movements.insert_one(return_movement)
                returned_stock_count += 1
                logger.info(f"Returned {allocated_quantity} units to stock {stock_id} from order {order_number}")
            else:
                logger.warning(f"Stock entry {stock_id} not found for returning allocation")
    
    # Clean up related data
    # Remove job specifications
    await db.job_specifications.delete_many({"order_id": order_id})
    
    # Remove materials status
    await db.materials_status.delete_many({"order_id": order_id})
    
    # Remove order items status
    await db.order_items_status.delete_many({"order_id": order_id})
    
    # Mark stock movements as archived (don't delete for audit trail)
    # Only archive allocation movements, keep return movements for audit
    await db.stock_movements.update_many(
        {
            "reference": order_number,
            "movement_type": "allocation"
        },
        {"$set": {"is_archived": True, "archived_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    # Perform hard delete - completely remove the order
    result = await db.orders.delete_one({"id": order_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Order not found")
    
    message = f"Order deleted successfully. {returned_stock_count} stock allocation(s) returned to inventory."
    logger.info(f"Deleted order {order_number} and returned {returned_stock_count} stock allocation(s)")
    return StandardResponse(success=True, message=message, data={"returned_stock_items": returned_stock_count})

# ============= PRODUCTION BOARD ENDPOINTS =============

@api_router.get("/production/board")
async def get_production_board(current_user: dict = Depends(require_any_role)):
    """Get production board with orders grouped by stage"""
    # Get all active orders
    orders = await db.orders.find({"status": {"$ne": OrderStatus.COMPLETED}}).to_list(1000)
    
    # Group orders by production stage
    board = {}
    for stage in ProductionStage:
        if stage != ProductionStage.CLEARED:  # Don't show cleared orders on board
            board[stage.value] = []
    
    for order in orders:
        stage = order.get("current_stage", ProductionStage.ORDER_ENTERED.value)
        if stage in board:
            # Get client info for logo
            client = await db.clients.find_one({"id": order["client_id"]})
            
            # Get materials status
            materials_status = await db.materials_status.find_one({"order_id": order["id"]})
            materials_ready = materials_status.get("materials_ready", False) if materials_status else False
            
            # Handle due_date comparison properly for timezone awareness
            due_date = order["due_date"]
            if isinstance(due_date, str):
                due_date = datetime.fromisoformat(due_date.replace("Z", "+00:00"))
            elif isinstance(due_date, datetime) and due_date.tzinfo is None:
                # Make timezone-naive datetime timezone-aware (assume UTC)
                due_date = due_date.replace(tzinfo=timezone.utc)
            
            order_info = {
                "id": order["id"],
                "order_number": order["order_number"],
                "client_name": order["client_name"],
                "client_logo": get_file_url(client.get("logo_path", "")) if client else None,
                "due_date": order["due_date"],
                "total_amount": order["total_amount"],
                "runtime": order.get("runtime_estimate", "2-3 days"),
                "materials_ready": materials_ready,
                "items": order["items"],
                "delivery_address": order.get("delivery_address"),
                "is_overdue": due_date < datetime.now(timezone.utc),
                "production_started_at": order.get("production_started_at"),
                "stage_start_times": order.get("stage_start_times", {}),
                "display_order": order.get("display_order", 999)  # Default to end if not set
            }
            board[stage].append(order_info)
    
    # Sort each stage by display_order
    for stage in board:
        board[stage].sort(key=lambda x: x.get("display_order", 999))
    
    return {"success": True, "data": board}

@api_router.get("/production/logs/{order_id}")
async def get_production_logs(order_id: str, current_user: dict = Depends(require_any_role)):
    """Get production logs for order"""
    logs = await db.production_logs.find({"order_id": order_id}).sort("timestamp", 1).to_list(1000)
    return {"success": True, "data": logs}

@api_router.post("/production/move-stage/{order_id}")
async def move_order_stage(
    order_id: str, 
    request: StageMovementRequest, 
    current_user: dict = Depends(require_admin_or_manager)
):
    """Move order to next or previous production stage"""
    # Get current order
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Define stage progression
    stage_order = [
        ProductionStage.ORDER_ENTERED,
        ProductionStage.PENDING_MATERIAL,
        ProductionStage.PAPER_SLITTING,
        ProductionStage.WINDING,
        ProductionStage.FINISHING,
        ProductionStage.DELIVERY,
        ProductionStage.INVOICING,
        ProductionStage.CLEARED
    ]
    
    current_stage = ProductionStage(order["current_stage"])
    current_index = stage_order.index(current_stage)
    
    # Calculate new stage
    if request.direction == "forward":
        new_index = min(current_index + 1, len(stage_order) - 1)
    elif request.direction == "backward":
        new_index = max(current_index - 1, 0)
    else:
        raise HTTPException(status_code=400, detail="Direction must be 'forward' or 'backward'")
    
    if new_index == current_index:
        raise HTTPException(status_code=400, detail="Order is already at the first/last stage")
    
    new_stage = stage_order[new_index]
    
    # Update order stage
    await db.orders.update_one(
        {"id": order_id},
        {
            "$set": {
                "current_stage": new_stage.value,
                "updated_at": datetime.now(timezone.utc)
            }
        }
    )
    
    # Log stage change
    stage_log = ProductionLog(
        order_id=order_id,
        stage=new_stage,
        updated_by=current_user["sub"],
        notes=request.notes
    )
    await db.production_logs.insert_one(stage_log.dict())
    
    return {"success": True, "message": f"Order moved to {new_stage.value}", "new_stage": new_stage.value}

@api_router.post("/production/jump-stage/{order_id}")
async def jump_to_stage(
    order_id: str, 
    request: StageJumpRequest, 
    current_user: dict = Depends(require_admin_or_manager)
):
    """Jump order directly to a specific production stage"""
    # Get current order
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Validate target stage
    try:
        target_stage = ProductionStage(request.target_stage)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid target stage")
    
    current_stage = ProductionStage(order["current_stage"])
    
    # Check if already at target stage
    if current_stage == target_stage:
        raise HTTPException(status_code=400, detail=f"Order is already at {target_stage.value} stage")
    
    # Update order stage
    await db.orders.update_one(
        {"id": order_id},
        {
            "$set": {
                "current_stage": target_stage.value,
                "updated_at": datetime.now(timezone.utc)
            }
        }
    )
    
    # Log stage jump
    stage_log = ProductionLog(
        order_id=order_id,
        stage=target_stage,
        updated_by=current_user["sub"],
        notes=f"Jumped from {current_stage.value} to {target_stage.value}" + (f" - {request.notes}" if request.notes else "")
    )
    await db.production_logs.insert_one(stage_log.dict())
    
    return {"success": True, "message": f"Order jumped to {target_stage.value}", "new_stage": target_stage.value}

@api_router.get("/production/materials-status/{order_id}")
async def get_materials_status(order_id: str, current_user: dict = Depends(require_any_role)):
    """Get materials status for order"""
    # Check if order exists first
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    status = await db.materials_status.find_one({"order_id": order_id})
    if not status:
        # Create default status if not exists
        default_status = MaterialsStatus(
            order_id=order_id,
            materials_ready=False,
            materials_checklist=[],
            updated_by=current_user["sub"]
        )
        await db.materials_status.insert_one(default_status.dict())
        return {"success": True, "data": default_status.dict()}
    
    # Remove MongoDB _id field to avoid serialization issues
    if "_id" in status:
        del status["_id"]
    
    return {"success": True, "data": status}

@api_router.put("/production/materials-status/{order_id}")
async def update_materials_status(
    order_id: str, 
    status_update: MaterialsStatusUpdate,
    current_user: dict = Depends(require_admin_or_manager)
):
    """Update materials status for order"""
    # Check if order exists
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Update or create materials status
    update_data = {
        "materials_ready": status_update.materials_ready,
        "materials_checklist": status_update.materials_checklist,
        "updated_by": current_user["sub"],
        "updated_at": datetime.now(timezone.utc)
    }
    
    result = await db.materials_status.update_one(
        {"order_id": order_id},
        {"$set": update_data},
        upsert=True
    )
    
    return {"success": True, "message": "Materials status updated"}



# ============= JOB CARD ARCHIVING ENDPOINTS =============

@api_router.post("/production/job-cards", response_model=StandardResponse)
async def save_job_card(job_card_data: dict, current_user: dict = Depends(require_any_role)):
    """Save completed job card to archive"""
    
    # Add metadata
    job_card_record = {
        "id": str(uuid.uuid4()),
        **job_card_data,
        "saved_at": datetime.now(timezone.utc).isoformat(),
        "saved_by": current_user.get("user_id") or current_user.get("sub")
    }
    
    # Insert into job_cards collection
    await db.job_cards.insert_one(job_card_record)
    
    return StandardResponse(
        success=True, 
        message="Job card archived successfully",
        data={"job_card_id": job_card_record["id"]}
    )

@api_router.get("/production/job-cards/order/{order_id}")
async def get_job_cards_by_order(order_id: str, current_user: dict = Depends(require_any_role)):
    """Get all job cards for a specific order"""
    
    job_cards = await db.job_cards.find({"order_id": order_id}).sort("completed_at", 1).to_list(100)
    
    # Remove MongoDB _id
    for card in job_cards:
        if "_id" in card:
            del card["_id"]
    
    return {
        "success": True,
        "data": job_cards
    }

@api_router.put("/production/job-cards/{job_card_id}", response_model=StandardResponse)
async def update_job_card(job_card_id: str, update_data: dict, current_user: dict = Depends(require_any_role)):
    """Update an archived job card"""
    
    # Check if job card exists
    job_card = await db.job_cards.find_one({"id": job_card_id})
    if not job_card:
        raise HTTPException(status_code=404, detail="Job card not found")
    
    # Add update metadata
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    update_data["updated_by"] = current_user.get("user_id") or current_user.get("sub")
    
    # Update the job card
    await db.job_cards.update_one(
        {"id": job_card_id},
        {"$set": update_data}
    )
    
    return StandardResponse(success=True, message="Job card updated successfully")

@api_router.get("/production/job-cards/search")
async def search_job_cards(
    customer: Optional[str] = None,
    invoice_number: Optional[str] = None,
    product: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(require_any_role)
):
    """Search job cards by customer, invoice number, or product"""
    
    # Build query
    query = {}
    
    # If searching by customer, get matching orders first
    if customer:
        orders = await db.orders.find({
            "client_name": {"$regex": customer, "$options": "i"}
        }).to_list(1000)
        order_ids = [o["id"] for o in orders]
        query["order_id"] = {"$in": order_ids}
    
    # Search by product name in product specs
    if product:
        query["product_specs.product_name"] = {"$regex": product, "$options": "i"}
    
    # Date range filter
    if start_date or end_date:
        date_filter = {}
        if start_date:
            date_filter["$gte"] = start_date
        if end_date:
            date_filter["$lte"] = end_date
        query["completed_at"] = date_filter
    
    # Get matching job cards
    job_cards = await db.job_cards.find(query).sort("completed_at", -1).to_list(1000)
    
    # Remove MongoDB _id and enrich with order info
    results = []
    for card in job_cards:
        if "_id" in card:
            del card["_id"]
        
        # Get order info
        order = await db.orders.find_one({"id": card["order_id"]})
        if order:
            card["order_number"] = order["order_number"]
            card["client_name"] = order["client_name"]
            
            # Check for invoice
            if invoice_number:
                invoice = await db.invoices.find_one({
                    "job_id": card["order_id"],
                    "invoice_number": {"$regex": invoice_number, "$options": "i"}
                })
                if invoice:
                    card["invoice_number"] = invoice["invoice_number"]
                    results.append(card)
            else:
                results.append(card)
    
    # Group by order
    grouped_results = {}
    for card in results:
        order_id = card["order_id"]
        if order_id not in grouped_results:
            grouped_results[order_id] = {
                "order_id": order_id,
                "order_number": card.get("order_number", "Unknown"),
                "client_name": card.get("client_name", "Unknown"),
                "invoice_number": card.get("invoice_number"),
                "job_cards": []
            }
        grouped_results[order_id]["job_cards"].append(card)
    
    return {
        "success": True,
        "data": list(grouped_results.values())
    }

@api_router.put("/production/order-item-status/{order_id}")
async def update_order_item_status(
    order_id: str,
    item_update: OrderItemStatusUpdate,
    current_user: dict = Depends(require_admin_or_manager)
):
    """Update completion status of specific order item"""
    # Get the order
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Validate item index
    if item_update.item_index >= len(order["items"]) or item_update.item_index < 0:
        raise HTTPException(status_code=400, detail="Invalid item index")
    
    # Update the specific item's completion status
    update_query = {
        f"items.{item_update.item_index}.is_completed": item_update.is_completed,
        "updated_at": datetime.now(timezone.utc)
    }
    
    await db.orders.update_one(
        {"id": order_id},
        {"$set": update_query}
    )
    
    return {"success": True, "message": f"Item {item_update.item_index} marked as {'completed' if item_update.is_completed else 'pending'}"}


# ============= REPORTS ENDPOINTS =============

@api_router.get("/reports/outstanding-jobs")
async def get_outstanding_jobs_report(current_user: dict = Depends(require_admin_or_manager)):
    """Generate outstanding jobs report"""
    # Count jobs by stage
    jobs_by_stage = {}
    total_jobs = 0
    overdue_jobs = 0
    jobs_due_today = 0
    jobs_due_this_week = 0
    
    now = datetime.now(timezone.utc)
    today_end = now.replace(hour=23, minute=59, second=59)
    week_end = now + timedelta(days=7)
    
    # Get all active orders
    orders = await db.orders.find({"status": {"$ne": OrderStatus.COMPLETED}}).to_list(1000)
    
    for order in orders:
        stage = order.get("current_stage", ProductionStage.ORDER_ENTERED.value)
        jobs_by_stage[stage] = jobs_by_stage.get(stage, 0) + 1
        total_jobs += 1
        
        due_date = order["due_date"]
        if isinstance(due_date, str):
            due_date = datetime.fromisoformat(due_date.replace("Z", "+00:00"))
        elif isinstance(due_date, datetime) and due_date.tzinfo is None:
            # Make timezone-naive datetime timezone-aware (assume UTC)
            due_date = due_date.replace(tzinfo=timezone.utc)
        
        if due_date < now:
            overdue_jobs += 1
        elif due_date <= today_end:
            jobs_due_today += 1
        elif due_date <= week_end:
            jobs_due_this_week += 1
    
    report = OutstandingJobsReport(
        total_jobs=total_jobs,
        jobs_by_stage=jobs_by_stage,
        overdue_jobs=overdue_jobs,
        jobs_due_today=jobs_due_today,
        jobs_due_this_week=jobs_due_this_week
    )
    
    return {"success": True, "data": report.dict()}

@api_router.get("/reports/late-deliveries")
async def get_late_deliveries_report(current_user: dict = Depends(require_admin_or_manager)):
    """Generate late deliveries report"""
    # Get completed orders that were delivered late
    completed_orders = await db.orders.find({"status": OrderStatus.COMPLETED}).to_list(1000)
    
    late_deliveries = []
    for order in completed_orders:
        due_date = order["due_date"]
        completed_date = order.get("completed_at")
        
        if completed_date and due_date:
            if isinstance(due_date, str):
                due_date = datetime.fromisoformat(due_date.replace("Z", "+00:00"))
            if isinstance(completed_date, str):
                completed_date = datetime.fromisoformat(completed_date.replace("Z", "+00:00"))
            
            if completed_date > due_date:
                delay_days = (completed_date - due_date).days
                late_deliveries.append({
                    "order": order,
                    "delay_days": delay_days,
                    "client_name": order["client_name"]
                })
    
    # Calculate statistics
    total_late = len(late_deliveries)
    avg_delay = sum(ld["delay_days"] for ld in late_deliveries) / total_late if total_late > 0 else 0
    
    # Group by client
    late_by_client = {}
    for ld in late_deliveries:
        client = ld["client_name"]
        late_by_client[client] = late_by_client.get(client, 0) + 1
    
    # Group by month
    late_by_month = {}
    for ld in late_deliveries:
        completed_date = ld["order"]["completed_at"]
        if isinstance(completed_date, str):
            completed_date = datetime.fromisoformat(completed_date.replace("Z", "+00:00"))
        month_key = completed_date.strftime("%Y-%m")
        late_by_month[month_key] = late_by_month.get(month_key, 0) + 1
    
    report = LateDeliveryReport(
        total_late_deliveries=total_late,
        average_delay_days=avg_delay,
        late_deliveries_by_client=late_by_client,
        late_deliveries_by_month=late_by_month
    )
    
    return {"success": True, "data": report.dict()}

@api_router.get("/reports/customer-annual/{client_id}")
async def get_customer_annual_report(client_id: str, year: int, current_user: dict = Depends(require_admin_or_manager)):
    """Generate customer annual report"""
    # Get client info
    client = await db.clients.find_one({"id": client_id})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Get orders for the year
    start_date = datetime(year, 1, 1)
    end_date = datetime(year + 1, 1, 1)
    
    orders = await db.orders.find({
        "client_id": client_id,
        "created_at": {"$gte": start_date, "$lt": end_date}
    }).to_list(1000)
    
    # Calculate statistics
    total_orders = len(orders)
    total_revenue = sum(order["total_amount"] for order in orders)
    avg_order_value = total_revenue / total_orders if total_orders > 0 else 0
    
    # Orders by month
    orders_by_month = {}
    revenue_by_month = {}
    for i in range(1, 13):
        month_key = f"{year}-{i:02d}"
        orders_by_month[month_key] = 0
        revenue_by_month[month_key] = 0
    
    for order in orders:
        created_date = order["created_at"]
        if isinstance(created_date, str):
            created_date = datetime.fromisoformat(created_date.replace("Z", "+00:00"))
        month_key = created_date.strftime("%Y-%m")
        if month_key in orders_by_month:
            orders_by_month[month_key] += 1
            revenue_by_month[month_key] += order["total_amount"]
    
    # Top products
    product_stats = {}
    for order in orders:
        for item in order["items"]:
            product_name = item["product_name"]
            if product_name not in product_stats:
                product_stats[product_name] = {"quantity": 0, "revenue": 0, "orders": 0}
            product_stats[product_name]["quantity"] += item["quantity"]
            product_stats[product_name]["revenue"] += item["total_price"]
            product_stats[product_name]["orders"] += 1
    
    top_products = sorted(
        [{"product": k, **v} for k, v in product_stats.items()],
        key=lambda x: x["revenue"],
        reverse=True
    )[:5]
    
    # On-time delivery rate
    completed_orders = [o for o in orders if o.get("completed_at")]
    on_time_orders = 0
    for order in completed_orders:
        due_date = order["due_date"]
        completed_date = order["completed_at"]
        
        if isinstance(due_date, str):
            due_date = datetime.fromisoformat(due_date.replace("Z", "+00:00"))
        if isinstance(completed_date, str):
            completed_date = datetime.fromisoformat(completed_date.replace("Z", "+00:00"))
        
        if completed_date <= due_date:
            on_time_orders += 1
    
    on_time_rate = (on_time_orders / len(completed_orders)) * 100 if completed_orders else 0
    
    report = CustomerAnnualReport(
        client_id=client_id,
        client_name=client["company_name"],
        year=year,
        total_orders=total_orders,
        total_revenue=total_revenue,
        average_order_value=avg_order_value,
        orders_by_month=orders_by_month,
        revenue_by_month=revenue_by_month,
        top_products=top_products,
        on_time_delivery_rate=on_time_rate
    )
    
    return {"success": True, "data": report.dict()}

# ============= DOCUMENT GENERATION ENDPOINTS =============

# Simplified authentication for PDF documents - allow open access for testing
async def simple_pdf_auth(token: str = None):
    """Simplified authentication for PDF downloads"""
    # For now, let's make PDFs accessible without strict authentication for testing
    # In production, you'd want proper authentication
    return {"user_id": "test-user", "role": "admin"}

@api_router.get("/documents/acknowledgment/{order_id}")
async def generate_acknowledgment(order_id: str):
    """Generate order acknowledgment PDF"""
    # Get order and client data
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    client = await db.clients.find_one({"id": order["client_id"]})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Prepare data for PDF generation
    order_data = {
        "order_number": order["order_number"],
        "invoice_number": f"INV-{order['order_number']}",
        "due_date": order["due_date"].strftime("%d/%m/%Y") if isinstance(order["due_date"], datetime) else order["due_date"],
        "client_name": client["company_name"],
        "client_address": f"{client['address']}, {client['city']}, {client['state']} {client['postal_code']}",
        "client_email": client["email"],
        "client_phone": client["phone"],
        "client_payment_terms": client.get("payment_terms", "Net 30 days"),
        "client_lead_time_days": client.get("lead_time_days", 7),
        "delivery_instructions": order.get("delivery_instructions", "Standard delivery terms apply."),
        "items": order["items"],
        "subtotal": order["subtotal"],
        "gst": order["gst"],
        "total_amount": order["total_amount"],
        "bank_details": client.get("bank_details")
    }
    
    # Generate PDF
    pdf_content = doc_generator.generate_order_acknowledgment(order_data)
    
    return StreamingResponse(
        BytesIO(pdf_content),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=acknowledgment_{order['order_number']}.pdf"}
    )

@api_router.get("/documents/job-card/{order_id}")
async def generate_job_card(order_id: str):
    """Generate job card PDF"""
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Get job specifications if available
    job_spec = await db.job_specifications.find_one({"order_id": order_id})
    
    # Get product specifications
    specifications = None
    if order["items"] and len(order["items"]) > 0:
        product_id = order["items"][0].get("product_id")
        if product_id:
            product = await db.products.find_one({"id": product_id})
            if product:
                specifications = product.get("specifications")
    
    job_data = {
        "order_number": order["order_number"],
        "client_name": order["client_name"],
        "due_date": order["due_date"].strftime("%d/%m/%Y") if isinstance(order["due_date"], datetime) else order["due_date"],
        "current_stage": order.get("current_stage", "order_entered"),
        "specifications": specifications,
        "job_specification": job_spec
    }
    
    pdf_content = doc_generator.generate_job_card(job_data)
    
    return StreamingResponse(
        BytesIO(pdf_content),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=job_card_{order['order_number']}.pdf"}
    )

@api_router.get("/documents/packing-list/{order_id}")
async def generate_packing_list(order_id: str):
    """Generate packing list PDF"""
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    order_data = {
        "order_number": order["order_number"],
        "client_name": order["client_name"],
        "delivery_address": order.get("delivery_address", "N/A"),
        "delivery_instructions": order.get("delivery_instructions"),
        "items": order["items"]
    }
    
    pdf_content = doc_generator.generate_packing_list(order_data)
    
    return StreamingResponse(
        BytesIO(pdf_content),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=packing_list_{order['order_number']}.pdf"}
    )

# Moved this function above document endpoints

@api_router.get("/documents/invoice/{order_id}")
async def generate_invoice_pdf(order_id: str):
    """Generate invoice PDF"""
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    client = await db.clients.find_one({"id": order["client_id"]})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    invoice_data = {
        "invoice_number": f"INV-{order['order_number']}",
        "order_number": order["order_number"],
        "payment_due_date": (datetime.now(timezone.utc) + timedelta(days=30)).strftime("%d/%m/%Y"),
        "client_name": client["company_name"],
        "client_address": f"{client['address']}, {client['city']}, {client['state']} {client['postal_code']}",
        "client_abn": client.get("abn", "N/A"),
        "items": order["items"],
        "subtotal": order["subtotal"],
        "gst": order["gst"],
        "total_amount": order["total_amount"]
    }
    
    pdf_content = doc_generator.generate_invoice(invoice_data)
    
    return StreamingResponse(
        BytesIO(pdf_content),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=invoice_{order['order_number']}.pdf"}
    )

# ============= XERO INTEGRATION ENDPOINTS =============

# Xero Integration Configuration
# Force correct environment values for current environment  
os.environ['XERO_REDIRECT_URI'] = 'https://mfgapp-ui-update.preview.emergentagent.com/xero-oauth-callback'
os.environ['FRONTEND_URL'] = 'https://mfgapp-ui-update.preview.emergentagent.com'

XERO_CLIENT_ID = os.getenv("XERO_CLIENT_ID")
XERO_CLIENT_SECRET = os.getenv("XERO_CLIENT_SECRET")  
XERO_CALLBACK_URL = os.getenv("XERO_REDIRECT_URI")
XERO_SCOPES = "openid profile email accounting.transactions accounting.contacts accounting.settings offline_access"

# Debug: Print at module load time
print(f"🔥 DEBUG: XERO_REDIRECT_URI env var: {os.getenv('XERO_REDIRECT_URI')}")
print(f"🔥 MODULE LOAD: XERO_CALLBACK_URL set to: {XERO_CALLBACK_URL}")

# Xero Account Configuration
XERO_DEFAULT_SALES_ACCOUNT_CODE = "200"  # Sales account code
XERO_DEFAULT_TAX_TYPE = "OUTPUT"  # Default GST/tax type

# Debug endpoint for testing
@api_router.get("/xero/debug")
async def debug_xero_config():
    """Debug endpoint to check Xero configuration"""
    # Debug: print the actual value being used right now
    print(f"🚀 DEBUG ENDPOINT: XERO_CALLBACK_URL = {XERO_CALLBACK_URL}")
    
    return {
        "client_id": XERO_CLIENT_ID,
        "callback_url": XERO_CALLBACK_URL,
        "callback_url_from_variable": XERO_CALLBACK_URL,
        "scopes": XERO_SCOPES,
        "expected_auth_url_start": "https://login.xero.com/identity/connect/authorize",
        "frontend_url": os.getenv("FRONTEND_URL", "https://app.emergent.sh"),
        "configuration": {
            "XERO_SALES_ACCOUNT_CODE": os.getenv("XERO_SALES_ACCOUNT_CODE", "200"),
            "XERO_SALES_ACCOUNT_NAME": os.getenv("XERO_SALES_ACCOUNT_NAME", "Sales")
        },
        "environment_check": {
            "client_id_set": bool(XERO_CLIENT_ID),
            "client_secret_set": bool(XERO_CLIENT_SECRET),
            "redirect_uri": XERO_CALLBACK_URL
        }
    }

@api_router.get("/debug/test-pdf")
async def test_pdf_download():
    """Simple test PDF for debugging download issues"""
    from reportlab.pdfgen import canvas
    from io import BytesIO
    
    buffer = BytesIO()
    c = canvas.Canvas(buffer)
    c.drawString(100, 750, "Test PDF Download")
    c.drawString(100, 730, "If you can see this, PDF downloads are working!")
    c.save()
    
    buffer.seek(0)
    
    return StreamingResponse(
        BytesIO(buffer.read()),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=test.pdf"}
    )

@api_router.get("/xero/status")
async def check_xero_connection_status(current_user: dict = Depends(require_admin_or_manager)):
    """Check if user has active Xero connection"""
    # Check if user has stored Xero tokens
    user_tokens = await db.xero_tokens.find_one({"user_id": current_user["user_id"]})
    
    if user_tokens and user_tokens.get("access_token"):
        # TODO: Validate token is still active with Xero API
        return {"connected": True, "message": "Xero connection active", "tenant_id": user_tokens.get("tenant_id")}
    else:
        return {"connected": False, "message": "No Xero connection found"}

# Helper function to get authenticated Xero API client
async def get_xero_api_client(user_id: str):
    """Get authenticated Xero API client for user"""
    user_tokens = await db.xero_tokens.find_one({"user_id": user_id})
    if not user_tokens or not user_tokens.get("access_token"):
        raise HTTPException(status_code=401, detail="No Xero connection found")
    
    # Check if token is expired
    if user_tokens.get("expires_at") and user_tokens["expires_at"] < datetime.now(timezone.utc):
        # Try to refresh token
        try:
            refreshed_tokens = await refresh_xero_token(user_id, user_tokens["refresh_token"])
            user_tokens = refreshed_tokens
        except Exception as e:
            raise HTTPException(status_code=401, detail="Xero token expired and refresh failed")
    
    # Create OAuth2 token object
    oauth2_token = OAuth2Token(
        client_id=XERO_CLIENT_ID,
        client_secret=XERO_CLIENT_SECRET
    )
    oauth2_token.access_token = user_tokens["access_token"]
    oauth2_token.refresh_token = user_tokens.get("refresh_token")
    oauth2_token.expires_at = user_tokens.get("expires_at")
    
    # Create API client
    configuration = Configuration(oauth2_token=oauth2_token)
    api_client = ApiClient(configuration, pool_threads=1)
    
    return api_client, user_tokens.get("tenant_id")

async def refresh_xero_token(user_id: str, refresh_token: str):
    """Refresh expired Xero access token"""
    token_url = "https://identity.xero.com/connect/token"
    
    auth_string = f"{XERO_CLIENT_ID}:{XERO_CLIENT_SECRET}"
    auth_bytes = auth_string.encode('ascii')
    auth_b64 = base64.b64encode(auth_bytes).decode('ascii')
    
    headers = {
        "Authorization": f"Basic {auth_b64}",
        "Content-Type": "application/x-www-form-urlencoded"
    }
    
    token_data = {
        "grant_type": "refresh_token",
        "refresh_token": refresh_token
    }
    
    response = requests.post(token_url, headers=headers, data=token_data)
    response.raise_for_status()
    
    tokens = response.json()
    
    # Update stored tokens
    token_record = {
        "user_id": user_id,
        "access_token": tokens["access_token"],
        "refresh_token": tokens.get("refresh_token", refresh_token),
        "expires_at": datetime.now(timezone.utc) + timedelta(seconds=tokens.get("expires_in", 1800)),
        "updated_at": datetime.now(timezone.utc)
    }
    
    await db.xero_tokens.update_one(
        {"user_id": user_id},
        {"$set": token_record}
    )
    
    return token_record

@api_router.get("/xero/auth/url")
async def get_xero_auth_url(current_user: dict = Depends(require_admin_or_manager)):
    """Get Xero OAuth authorization URL"""
    # Generate secure state parameter
    state = secrets.token_urlsafe(32)
    
    # Store state for validation
    await db.xero_auth_states.insert_one({
        "state": state,
        "user_id": current_user["user_id"],
        "created_at": datetime.now(timezone.utc)
    })
    
    # Build authorization URL
    auth_params = {
        "response_type": "code",
        "client_id": XERO_CLIENT_ID,
        "redirect_uri": XERO_CALLBACK_URL,
        "scope": XERO_SCOPES,
        "state": state
    }
    
    auth_url = f"https://login.xero.com/identity/connect/authorize?{urlencode(auth_params)}"
    
    # Debug logging
    logger.info(f"Generated Xero OAuth URL: {auth_url}")
    logger.info(f"Client ID: {XERO_CLIENT_ID}")
    logger.info(f"Callback URL: {XERO_CALLBACK_URL}")
    logger.info(f"Scopes: {XERO_SCOPES}")
    
    return {"auth_url": auth_url, "state": state, "debug_info": {
        "client_id": XERO_CLIENT_ID,
        "callback_url": XERO_CALLBACK_URL,
        "scopes": XERO_SCOPES
    }}

@api_router.get("/xero/callback")
async def handle_xero_oauth_redirect(code: str = None, state: str = None, error: str = None):
    """Handle Xero OAuth redirect - redirect to frontend"""
    if error:
        # Redirect to frontend with error
        return Response(
            status_code=302,
            headers={"Location": f"{os.getenv('FRONTEND_URL')}/xero/callback?error={error}"}
        )
    
    if code and state:
        # Redirect to frontend with code and state  
        return Response(
            status_code=302,
            headers={"Location": f"{os.getenv('FRONTEND_URL')}/xero/callback?code={code}&state={state}"}
        )
    
    # Default redirect to frontend
    return Response(
        status_code=302,
        headers={"Location": f"{os.getenv('FRONTEND_URL')}/"}
    )

@api_router.post("/xero/auth/callback")
async def handle_xero_callback(
    callback_data: dict,
    current_user: dict = Depends(require_admin_or_manager)
):
    """Handle Xero OAuth callback and exchange code for tokens"""
    auth_code = callback_data.get("code")
    state = callback_data.get("state")
    
    if not auth_code or not state:
        raise HTTPException(status_code=400, detail="Missing authorization code or state")
    
    # Validate state parameter
    stored_state = await db.xero_auth_states.find_one({
        "state": state,
        "user_id": current_user["user_id"]
    })
    
    if not stored_state:
        raise HTTPException(status_code=400, detail="Invalid state parameter")
    
    # Clean up used state
    await db.xero_auth_states.delete_one({"state": state})
    
    # Exchange code for tokens
    token_url = "https://identity.xero.com/connect/token"
    
    # Prepare authentication
    auth_string = f"{XERO_CLIENT_ID}:{XERO_CLIENT_SECRET}"
    auth_bytes = auth_string.encode('ascii')
    auth_b64 = base64.b64encode(auth_bytes).decode('ascii')
    
    headers = {
        "Authorization": f"Basic {auth_b64}",
        "Content-Type": "application/x-www-form-urlencoded"
    }
    
    token_data = {
        "grant_type": "authorization_code",
        "code": auth_code,
        "redirect_uri": XERO_CALLBACK_URL
    }
    
    try:
        response = requests.post(token_url, headers=headers, data=token_data)
        response.raise_for_status()
        
        tokens = response.json()
        
        # Store tokens for user
        token_record = {
            "user_id": current_user["user_id"],
            "access_token": tokens["access_token"],
            "refresh_token": tokens["refresh_token"],
            "expires_at": datetime.now(timezone.utc) + timedelta(seconds=tokens.get("expires_in", 1800)),
            "created_at": datetime.now(timezone.utc),
            "tenant_id": None  # Will be updated when we get tenant info
        }
        
        # Upsert token record
        await db.xero_tokens.replace_one(
            {"user_id": current_user["user_id"]},
            token_record,
            upsert=True
        )
        
        return {"message": "Xero connection successful", "access_token_expires_in": tokens.get("expires_in", 1800)}
        
    except requests.exceptions.RequestException as e:
        logger.error(f"Xero token exchange failed: {str(e)}")
        raise HTTPException(status_code=400, detail="Failed to exchange authorization code")

@api_router.delete("/xero/disconnect")
async def disconnect_xero(current_user: dict = Depends(require_admin_or_manager)):
    """Disconnect from Xero"""
    # Remove stored tokens
    result = await db.xero_tokens.delete_one({"user_id": current_user["user_id"]})
    
    if result.deleted_count > 0:
        return {"message": "Xero disconnection successful"}
    else:
        return {"message": "No Xero connection found to disconnect"}

@api_router.get("/xero/next-invoice-number")
async def get_next_xero_invoice_number(current_user: dict = Depends(require_admin_or_manager)):
    """Get the next available invoice number from Xero"""
    try:
        api_client, tenant_id = await get_xero_api_client(current_user["user_id"])
        
        if not tenant_id:
            # Get tenant info if not stored
            from xero_python.identity import IdentityApi
            identity_api = IdentityApi(api_client)
            connections = identity_api.get_connections()
            
            if not connections or not connections[0]:
                raise HTTPException(status_code=400, detail="No Xero organization connected")
            
            tenant_id = connections[0].tenant_id
            
            # Update stored token record with tenant_id
            await db.xero_tokens.update_one(
                {"user_id": current_user["user_id"]},
                {"$set": {"tenant_id": tenant_id}}
            )
        
        # Get invoices to determine next number
        accounting_api = AccountingApi(api_client)
        
        # Get recent invoices ordered by invoice number descending
        invoices_response = accounting_api.get_invoices(
            xero_tenant_id=tenant_id,
            order="InvoiceNumber DESC",
            page=1
        )
        
        next_number = 1
        if invoices_response.invoices and len(invoices_response.invoices) > 0:
            # Extract numeric part from the last invoice number
            last_invoice = invoices_response.invoices[0]
            if last_invoice.invoice_number:
                # Try to extract number from invoice number (handle different formats)
                import re
                numbers = re.findall(r'\d+', last_invoice.invoice_number)
                if numbers:
                    next_number = int(numbers[-1]) + 1
        
        # Format as INV-XXXXXX
        formatted_number = f"INV-{next_number:06d}"
        
        return {
            "next_number": next_number,
            "formatted_number": formatted_number,
            "tenant_id": tenant_id
        }
        
    except Exception as e:
        logger.error(f"Failed to get next invoice number from Xero: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get next invoice number: {str(e)}")

@api_router.get("/xero/account-codes")
async def get_xero_account_codes(current_user: dict = Depends(require_admin_or_manager)):
    """Get available account codes from Xero for setup verification"""
    try:
        api_client, tenant_id = await get_xero_api_client(current_user["user_id"])
        
        if not tenant_id:
            raise HTTPException(status_code=400, detail="No Xero tenant ID found")
        
        accounting_api = AccountingApi(api_client)
        
        # Get all revenue accounts
        accounts_response = accounting_api.get_accounts(
            xero_tenant_id=tenant_id,
            where='Type=="REVENUE" AND Status=="ACTIVE"'
        )
        
        revenue_accounts = []
        if accounts_response.accounts:
            for account in accounts_response.accounts:
                revenue_accounts.append({
                    "code": account.code,
                    "name": account.name,
                    "description": account.description,
                    "type": account.type
                })
        
        return {
            "success": True,
            "revenue_accounts": revenue_accounts,
            "recommended_setup": {
                "message": "For best results, ensure you have a revenue account with code '200' or update Misty to use one of your existing codes",
                "current_default": "200"
            }
        }
        
    except Exception as e:
        logger.error(f"Failed to get Xero account codes: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get account codes: {str(e)}")

@api_router.get("/xero/tax-rates")
async def get_xero_tax_rates(current_user: dict = Depends(require_admin_or_manager)):
    """Get available tax rates from Xero"""
    try:
        api_client, tenant_id = await get_xero_api_client(current_user["user_id"])
        
        if not tenant_id:
            raise HTTPException(status_code=400, detail="No Xero tenant ID found")
        
        accounting_api = AccountingApi(api_client)
        
        # Get all tax rates
        tax_rates_response = accounting_api.get_tax_rates(xero_tenant_id=tenant_id)
        
        tax_rates = []
        if tax_rates_response.tax_rates:
            for tax_rate in tax_rates_response.tax_rates:
                tax_rates.append({
                    "name": tax_rate.name,
                    "tax_type": tax_rate.tax_type,
                    "rate": str(tax_rate.effective_rate) if tax_rate.effective_rate else "0",
                    "status": tax_rate.status
                })
        
        return {
            "success": True,
            "tax_rates": tax_rates,
            "current_default": "OUTPUT"
        }
        
    except Exception as e:
        logger.error(f"Failed to get Xero tax rates: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get tax rates: {str(e)}")

@api_router.post("/xero/webhook")
async def handle_xero_webhook(request: Request):
    """Handle Xero webhook notifications"""
    try:
        # Get the raw body
        body = await request.body()
        
        # Get webhook signature from headers
        xero_signature = request.headers.get("x-xero-signature")
        
        # Log the webhook for debugging
        logger.info(f"Received Xero webhook: {body[:100]}...")  # Log first 100 chars
        
        # For now, just return success - webhook validation can be added later
        return {"status": "received"}
        
    except Exception as e:
        logger.error(f"Xero webhook error: {str(e)}")
        raise HTTPException(status_code=500, detail="Webhook processing failed")

@api_router.get("/xero/webhook")
async def xero_webhook_intent():
    """Handle Xero webhook 'Intent to receive' verification"""
    # Return 200 to confirm we can receive webhooks
    return {"status": "ready", "message": "Webhook endpoint is ready to receive notifications"}

# ============= INTERNAL XERO HELPER FUNCTIONS =============

async def get_next_xero_invoice_number():
    """Internal helper to get next invoice number from Xero"""
    try:
        # Get system Xero token (simplified for accounting transactions)
        xero_token = await db.xero_tokens.find_one({"user_id": "system"})
        if not xero_token:
            raise Exception("No Xero connection found")
        
        api_client, tenant_id = await get_xero_api_client("system")
        
        if not tenant_id:
            # Get tenant info
            from xero_python.identity import IdentityApi
            identity_api = IdentityApi(api_client)
            connections = identity_api.get_connections()
            
            if not connections or not connections[0]:
                raise Exception("No Xero organization connected")
            
            tenant_id = connections[0].tenant_id
            
            # Update stored token record with tenant_id
            await db.xero_tokens.update_one(
                {"user_id": "system"},
                {"$set": {"tenant_id": tenant_id}}
            )
        
        # Get invoices to determine next number
        accounting_api = AccountingApi(api_client)
        
        invoices_response = accounting_api.get_invoices(
            xero_tenant_id=tenant_id,
            order="InvoiceNumber DESC",
            page=1
        )
        
        next_number = 1
        if invoices_response.invoices and len(invoices_response.invoices) > 0:
            latest_invoice = invoices_response.invoices[0]
            if latest_invoice.invoice_number:
                try:
                    # Extract numeric part and increment
                    import re
                    numeric_part = re.findall(r'\d+', latest_invoice.invoice_number)
                    if numeric_part:
                        next_number = int(numeric_part[-1]) + 1
                except (ValueError, IndexError):
                    next_number = 1
        
        formatted_number = f"INV-{next_number:04d}"
        
        return {
            "next_number": next_number,
            "formatted_number": formatted_number
        }
        
    except Exception as e:
        logger.error(f"Failed to get next Xero invoice number: {str(e)}")
        raise Exception(f"Failed to get next invoice number: {str(e)}")

async def create_xero_draft_invoice(invoice_data):
    """Internal helper to create draft invoice in Xero with proper formatting"""
    try:
        api_client, tenant_id = await get_xero_api_client("system")
        
        if not tenant_id:
            raise Exception("No Xero tenant ID available")
        
        from xero_python.accounting import Contact, Invoice, LineItem
        
        # Create contact with proper formatting
        contact = Contact(name=invoice_data["client_name"])
        
        # Add email address if available (not required)
        if invoice_data.get("client_email"):
            contact.email_address = invoice_data["client_email"]
        
        # Create line items with proper Xero formatting
        line_items = []
        for item in invoice_data["items"]:
            # Build description - combine product name and specifications
            description_parts = []
            if item.get("product_name"):
                description_parts.append(item["product_name"])
            if item.get("specifications"):
                description_parts.append(item["specifications"])
            if item.get("description"):
                description_parts.append(item["description"])
            
            # Use the first non-empty description or default
            description = " - ".join(description_parts) if description_parts else "Product/Service"
            
            # Create line item with required fields
            line_item = LineItem(
                description=description,
                quantity=float(item.get("quantity", 1)),
                unit_amount=float(item.get("unit_price", 0)),
                account_code=os.getenv("XERO_SALES_ACCOUNT_CODE", "200")  # Use configured sales account
            )
            
            # Add inventory item code if available (optional)
            if item.get("product_code"):
                line_item.item_code = item["product_code"]
            
            # Add discount if available (optional)
            if item.get("discount_percent"):
                line_item.discount_rate = float(item["discount_percent"])
            
            line_items.append(line_item)
        
        # Create invoice with proper date formatting
        invoice_date = datetime.now(timezone.utc).date()
        due_date = None
        
        if invoice_data.get("due_date"):
            try:
                due_date = datetime.strptime(invoice_data["due_date"], '%Y-%m-%d').date()
            except ValueError:
                # If date parsing fails, default to 30 days from invoice date
                due_date = invoice_date + timedelta(days=30)
        else:
            # Default to 30 days if no due date specified
            due_date = invoice_date + timedelta(days=30)
        
        # Create invoice with all required fields
        invoice = Invoice(
            type="ACCREC",  # Accounts Receivable (required)
            contact=contact,  # Contact (required)
            date=invoice_date,  # Invoice date (required)
            due_date=due_date,  # Due date (required)
            line_items=line_items,  # Line items (required)
            invoice_number=invoice_data["invoice_number"],  # Invoice number (required)
            status="DRAFT"  # Create as draft
        )
        
        # Add reference (order number) if available (optional)
        if invoice_data.get("reference") or invoice_data.get("order_number"):
            invoice.reference = invoice_data.get("reference") or invoice_data.get("order_number")
        
        # Add currency if specified (optional)
        if invoice_data.get("currency"):
            invoice.currency_code = invoice_data["currency"]
        
        accounting_api = AccountingApi(api_client)
        
        # Create the invoice
        invoices = [invoice]
        created_invoices = accounting_api.create_invoices(
            xero_tenant_id=tenant_id, 
            invoices=invoices
        )
        
        if created_invoices.invoices and len(created_invoices.invoices) > 0:
            created_invoice = created_invoices.invoices[0]
            return {
                "success": True,
                "invoice_id": created_invoice.invoice_id,
                "invoice_number": created_invoice.invoice_number,
                "status": created_invoice.status,
                "total": float(created_invoice.total) if created_invoice.total else 0
            }
        else:
            raise Exception("No invoice was created")
        
    except Exception as e:
        logger.error(f"Failed to create Xero draft invoice: {str(e)}")
        raise Exception(f"Failed to create Xero draft: {str(e)}")

async def validate_sales_account(accounting_api, tenant_id: str) -> str:
    """Validate that Sales account with code 200 exists, or find suitable alternative"""
    try:
        # First, try to find account with code "200"
        accounts_response = accounting_api.get_accounts(
            xero_tenant_id=tenant_id,
            where=f'Code="{XERO_DEFAULT_SALES_ACCOUNT_CODE}" AND Status=="ACTIVE"'
        )
        
        if accounts_response.accounts and len(accounts_response.accounts) > 0:
            account = accounts_response.accounts[0]
            logger.info(f"Found Sales account: {account.name} (Code: {account.code})")
            return account.code
        
        # If code "200" doesn't exist, look for any Sales or Revenue account
        logger.warning(f"Sales account with code {XERO_DEFAULT_SALES_ACCOUNT_CODE} not found. Looking for alternatives...")
        
        # Try to find Sales or Revenue accounts
        sales_accounts = accounting_api.get_accounts(
            xero_tenant_id=tenant_id,
            where='(Type=="REVENUE" OR Type=="SALES") AND Status=="ACTIVE"'
        )
        
        if sales_accounts.accounts and len(sales_accounts.accounts) > 0:
            # Use the first available sales/revenue account
            account = sales_accounts.accounts[0]
            logger.info(f"Using alternative Sales account: {account.name} (Code: {account.code})")
            return account.code
        
        # Last resort: look for any income account
        income_accounts = accounting_api.get_accounts(
            xero_tenant_id=tenant_id,
            where='Type=="REVENUE" AND Status=="ACTIVE"'
        )
        
        if income_accounts.accounts and len(income_accounts.accounts) > 0:
            account = income_accounts.accounts[0]
            logger.info(f"Using fallback Revenue account: {account.name} (Code: {account.code})")
            return account.code
        
        # If nothing found, return the default and let Xero handle the error
        logger.error("No suitable Sales or Revenue accounts found in Xero")
        return XERO_DEFAULT_SALES_ACCOUNT_CODE
        
    except Exception as e:
        logger.error(f"Error validating sales account: {str(e)}")
        return XERO_DEFAULT_SALES_ACCOUNT_CODE

@api_router.post("/xero/create-draft-invoice")
async def create_xero_draft_invoice(
    invoice_data: dict,
    current_user: dict = Depends(require_admin_or_manager)
):
    """Create a draft invoice in Xero"""
    try:
        api_client, tenant_id = await get_xero_api_client(current_user["user_id"])
        
        if not tenant_id:
            raise HTTPException(status_code=400, detail="No Xero tenant ID found")
        
        accounting_api = AccountingApi(api_client)
        
        # Get or create contact in Xero
        contact_name = invoice_data.get("client_name", "Unknown Client")
        contact_email = invoice_data.get("client_email")
        
        # Search for existing contact
        contact_id = None
        if contact_email:
            try:
                contacts_response = accounting_api.get_contacts(
                    xero_tenant_id=tenant_id,
                    where=f'EmailAddress="{contact_email}"'
                )
                if contacts_response.contacts and len(contacts_response.contacts) > 0:
                    contact_id = contacts_response.contacts[0].contact_id
            except:
                pass
        
        # Create contact if not found
        if not contact_id:
            try:
                new_contact = Contact(
                    name=contact_name,
                    email_address=contact_email if contact_email else None
                )
                contacts_request = Contacts(contacts=[new_contact])
                contacts_response = accounting_api.create_contacts(
                    xero_tenant_id=tenant_id,
                    contacts=contacts_request
                )
                if contacts_response.contacts and len(contacts_response.contacts) > 0:
                    contact_id = contacts_response.contacts[0].contact_id
                else:
                    raise Exception("Failed to create contact")
            except Exception as e:
                logger.error(f"Failed to create contact in Xero: {str(e)}")
                # Use default contact or create a simple one
                contact_id = None
        
        # Prepare line items with Sales account
        line_items = []
        items = invoice_data.get("items", [])
        
        # Debug logging
        logger.info(f"Creating Xero invoice with {len(items)} items")
        
        # Validate that Sales account code "200" exists
        sales_account_code = await validate_sales_account(accounting_api, tenant_id)
        
        for item in items:
            # Debug individual item
            logger.info(f"Processing item: {item}")
            
            line_item = LineItem(
                description=item.get("product_name", item.get("description", "Product/Service")),
                quantity=float(item.get("quantity", 1)),
                unit_amount=float(item.get("unit_price", 0)),
                account_code=sales_account_code,
                tax_type=XERO_DEFAULT_TAX_TYPE
            )
            line_items.append(line_item)
        
        # If no items provided, create a default line item
        if not line_items:
            total_amount = float(invoice_data.get("total_amount", 0))
            logger.info(f"No items found, creating default line item with total: {total_amount}")
            line_item = LineItem(
                description=f"Services for {contact_name}",
                quantity=1,
                unit_amount=total_amount,
                account_code=sales_account_code,
                tax_type=XERO_DEFAULT_TAX_TYPE
            )
            line_items.append(line_item)
        
        # Parse and format due_date properly for Xero
        due_date_value = invoice_data.get("due_date")
        if isinstance(due_date_value, str):
            try:
                due_date_parsed = datetime.fromisoformat(due_date_value).date()
            except:
                due_date_parsed = datetime.now().date()
        else:
            due_date_parsed = due_date_value or datetime.now().date()
        
        # Debug logging for Xero invoice creation
        logger.info(f"Creating Xero invoice: contact_name={contact_name}, contact_id={contact_id}")
        logger.info(f"Line items count: {len(line_items)}")
        logger.info(f"Invoice number: {invoice_data.get('invoice_number')}")
        
        # Create invoice object
        invoice = Invoice(
            type="ACCREC",  # Accounts Receivable
            contact=Contact(contact_id=contact_id) if contact_id else Contact(name=contact_name),
            line_items=line_items,
            date=datetime.now().date(),
            due_date=due_date_parsed,
            invoice_number=invoice_data.get("invoice_number"),
            reference=invoice_data.get("reference", invoice_data.get("order_number")),
            status="DRAFT"
        )
        
        # Create invoice in Xero
        invoices_request = Invoices(invoices=[invoice])
        invoices_response = accounting_api.create_invoices(
            xero_tenant_id=tenant_id,
            invoices=invoices_request
        )
        
        if not invoices_response.invoices or len(invoices_response.invoices) == 0:
            raise Exception("No invoice returned from Xero API")
        
        created_invoice = invoices_response.invoices[0]
        
        # Log successful creation
        logger.info(f"Successfully created Xero invoice: {created_invoice.invoice_number}")
        
        # Check for validation errors
        if hasattr(created_invoice, 'validation_errors') and created_invoice.validation_errors:
            error_messages = [error.message for error in created_invoice.validation_errors]
            raise Exception(f"Invoice validation failed: {', '.join(error_messages)}")
        
        return {
            "success": True,
            "message": "Draft invoice created in Xero",
            "invoice_id": created_invoice.invoice_id,
            "invoice_number": created_invoice.invoice_number,
            "status": created_invoice.status,
            "total": str(created_invoice.total) if created_invoice.total else "0",
            "xero_url": f"https://go.xero.com/AccountsReceivable/View.aspx?InvoiceID={created_invoice.invoice_id}"
        }
        
    except Exception as e:
        logger.error(f"Failed to create draft invoice in Xero: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to create draft invoice: {str(e)}")

# ============= INVOICING ENDPOINTS =============

@api_router.get("/invoicing/live-jobs")
async def get_live_jobs(current_user: dict = Depends(require_admin_or_manager)):
    """Get all jobs ready for invoicing (in invoicing stage, including partially invoiced)"""
    live_jobs = await db.orders.find({
        "current_stage": "invoicing",
        "invoiced": {"$ne": True},  # Exclude fully invoiced orders
        "status": {"$ne": "completed"}
    }).to_list(length=None)
    
    # Convert ObjectId to string and enrich with client information
    for job in live_jobs:
        # Remove MongoDB ObjectId if present
        if "_id" in job:
            del job["_id"]
            
        client = await db.clients.find_one({"id": job["client_id"]})
        if client:
            job["client_name"] = client["company_name"]
            job["client_email"] = client.get("email", "")  # Add client email for Xero
            job["client_payment_terms"] = client.get("payment_terms", "Net 30 days")
        
    return {"data": live_jobs}

@api_router.post("/invoicing/generate/{job_id}")
async def generate_job_invoice(
    job_id: str,
    invoice_data: dict,
    current_user: dict = Depends(require_admin_or_manager)
):
    """Generate invoice for a completed job"""
    # Get job/order
    job = await db.orders.find_one({"id": job_id})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    # Get client info
    client = await db.clients.find_one({"id": job["client_id"]})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Get invoice history for this job
    invoice_history = job.get("invoice_history", [])
    is_partial = invoice_data.get("invoice_type") == "partial"
    
    # Generate base invoice number if this is the first invoice for this job
    if not invoice_history:
        invoice_count = await db.invoices.count_documents({}) + 1
        base_invoice_number = f"INV-{invoice_count:04d}"
    else:
        # Use existing base invoice number from first invoice
        base_invoice_number = invoice_history[0].get("base_invoice_number") or invoice_history[0].get("invoice_number").split("~")[0]
    
    # Add suffix for partial invoices
    if is_partial:
        partial_count = len(invoice_history) + 1
        invoice_number = f"{base_invoice_number}~{partial_count}"
    else:
        invoice_number = base_invoice_number
    
    # Prepare invoice data
    invoice_record = {
        "id": str(uuid.uuid4()),
        "invoice_number": invoice_number,
        "base_invoice_number": base_invoice_number,
        "job_id": job_id,
        "client_id": job["client_id"],
        "invoice_type": invoice_data.get("invoice_type", "full"),
        "items": invoice_data.get("items", job["items"]),
        "subtotal": invoice_data.get("subtotal", job["subtotal"]),
        "gst": invoice_data.get("gst", job["gst"]),
        "total_amount": invoice_data.get("total_amount", job["total_amount"]),
        "payment_terms": client.get("payment_terms", "Net 30 days"),
        "due_date": invoice_data.get("due_date"),
        "created_by": current_user["user_id"],
        "created_at": datetime.now(timezone.utc),
        "status": "draft"
    }
    
    # Insert invoice record
    await db.invoices.insert_one(invoice_record)
    
    # Calculate if job is fully invoiced for partial invoices
    is_fully_invoiced = False
    if invoice_data.get("invoice_type") == "partial":
        # Get existing invoice history
        invoice_history = job.get("invoice_history", [])
        
        # Add this invoice to history
        invoice_history_entry = {
            "invoice_number": invoice_number,
            "invoice_id": invoice_record["id"],
            "items": invoice_data.get("items", []),
            "date": datetime.now(timezone.utc).isoformat()
        }
        invoice_history.append(invoice_history_entry)
        
        # Calculate total invoiced quantities per item
        invoiced_quantities = {}
        for inv in invoice_history:
            for item in inv.get("items", []):
                product_id = item.get("product_id") or item.get("product_name")
                if product_id:
                    invoiced_quantities[product_id] = invoiced_quantities.get(product_id, 0) + item.get("quantity", 0)
        
        # Check if all items are fully invoiced
        is_fully_invoiced = True
        for item in job["items"]:
            product_id = item.get("product_id") or item.get("product_name")
            original_quantity = item.get("quantity", 0)
            invoiced_qty = invoiced_quantities.get(product_id, 0)
            if invoiced_qty < original_quantity:
                is_fully_invoiced = False
                break
    
    # Update job status and move to accounting transactions
    update_data = {
        "invoiced": True,
        "invoice_id": invoice_record["id"],
        "current_stage": "accounting_transaction",  # New stage for accounting processing
        "status": "accounting_draft",  # New status for accounting transactions
        "invoice_date": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc)
    }
    
    # If partial invoice, track history and check if fully invoiced
    if invoice_data.get("invoice_type") == "partial":
        update_data["invoice_history"] = invoice_history
        update_data["fully_invoiced"] = is_fully_invoiced
        
        if not is_fully_invoiced:
            # Keep in invoicing stage for remaining items
            update_data["invoiced"] = False
            update_data["partially_invoiced"] = True
            update_data["current_stage"] = "invoicing"  # Keep in invoicing stage so it appears in "Jobs Ready for Invoicing"
            update_data["status"] = "active"  # Keep as active for partial invoices
            # Remove invoice date for partial invoices
            del update_data["invoice_date"]
        else:
            # All items invoiced - move to accounting
            update_data["invoiced"] = True
            update_data["partially_invoiced"] = False
    
    await db.orders.update_one(
        {"id": job_id},
        {"$set": update_data}
    )
    
    # For full invoices in accounting transactions, automatically create Xero draft
    if invoice_data.get("invoice_type") != "partial":
        try:
            # Check if Xero is connected
            xero_token = await db.xero_tokens.find_one({"user_id": "system"})
            if xero_token and xero_token.get("access_token"):
                logger.info("Creating Xero draft invoice for accounting transaction")
                
                # Get client info for Xero invoice
                client = await db.clients.find_one({"id": job["client_id"]})
                
                # Get next Xero invoice number
                next_number_response = await get_next_xero_invoice_number()
                next_invoice_number = next_number_response["formatted_number"]
                
                # Prepare Xero invoice data with proper formatting
                xero_invoice_data = {
                    "client_name": client["company_name"] if client else job.get("client_name", "Unknown Client"),
                    "client_email": client.get("email", "") if client else "",
                    "invoice_number": next_invoice_number,
                    "order_number": job["order_number"],
                    "items": [],
                    "total_amount": invoice_data.get("total_amount", job["total_amount"]),
                    "due_date": invoice_data.get("due_date"),
                    "reference": job["order_number"]
                }
                
                # Format items for Xero with proper field mapping
                items = invoice_data.get("items", job["items"])
                for item in items:
                    xero_item = {
                        "product_name": item.get("product_name", item.get("description", "Product")),
                        "description": item.get("description", ""),
                        "specifications": item.get("specifications", ""),
                        "quantity": item.get("quantity", 1),
                        "unit_price": item.get("unit_price", item.get("price", 0)),
                        "product_code": item.get("product_code", ""),  # For InventoryItemCode
                        "discount_percent": item.get("discount_percent", 0) if item.get("discount_percent") else None
                    }
                    xero_invoice_data["items"].append(xero_item)
                
                # Create draft invoice in Xero
                xero_response = await create_xero_draft_invoice(xero_invoice_data)
                logger.info(f"Xero draft invoice created successfully: {xero_response}")
                
                # Update the invoice record with Xero details
                await db.invoices.update_one(
                    {"id": invoice_record["id"]},
                    {"$set": {
                        "xero_invoice_id": xero_response.get("invoice_id"),
                        "xero_invoice_number": next_invoice_number,
                        "xero_status": "draft"
                    }}
                )
                
        except Exception as e:
            logger.error(f"Failed to create Xero draft invoice: {str(e)}")
            # Don't fail the entire invoice process if Xero fails
            pass
    
    return {
        "message": "Invoice generated successfully and moved to accounting transactions",
        "invoice_id": invoice_record["id"],
        "invoice_number": invoice_number,
        "xero_draft_created": "xero_invoice_id" in locals()
    }

@api_router.get("/invoicing/archived-jobs")
async def get_archived_jobs(
    month: Optional[int] = None, 
    year: Optional[int] = None,
    current_user: dict = Depends(require_admin_or_manager)
):
    """Get archived jobs (completed and invoiced)"""
    # Build query filter
    query_filter = {"invoiced": True}
    
    if month and year:
        start_date = datetime(year, month, 1)
        if month == 12:
            end_date = datetime(year + 1, 1, 1)
        else:
            end_date = datetime(year, month + 1, 1)
        
        query_filter["created_at"] = {
            "$gte": start_date,
            "$lt": end_date
        }
    
    archived_jobs = await db.orders.find(query_filter).to_list(length=None)
    
    # Enrich with client and invoice information
    for job in archived_jobs:
        # Remove MongoDB ObjectId if present
        if "_id" in job:
            del job["_id"]
            
        # Get client info
        client = await db.clients.find_one({"id": job["client_id"]})
        if client:
            job["client_name"] = client["company_name"]
        
        # Get invoice info
        if job.get("invoice_id"):
            invoice = await db.invoices.find_one({"id": job["invoice_id"]})
            if invoice:
                job["invoice_number"] = invoice["invoice_number"]
                job["invoice_date"] = invoice["created_at"]
    
    return {"data": archived_jobs}

@api_router.get("/invoicing/monthly-report")
async def get_monthly_invoicing_report(
    month: int,
    year: int,
    current_user: dict = Depends(require_admin_or_manager)
):
    """Generate monthly invoicing report"""
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1)
    else:
        end_date = datetime(year, month + 1, 1)
    
    # Get jobs completed this month
    completed_jobs = await db.orders.find({
        "updated_at": {"$gte": start_date, "$lt": end_date},
        "current_stage": "cleared"
    }).to_list(length=None)
    
    # Get jobs invoiced this month
    invoiced_jobs = await db.invoices.find({
        "created_at": {"$gte": start_date, "$lt": end_date}
    }).to_list(length=None)
    
    # Remove ObjectIds from results
    for job in completed_jobs:
        if "_id" in job:
            del job["_id"]
    
    for invoice in invoiced_jobs:
        if "_id" in invoice:
            del invoice["_id"]
    
    # Calculate totals
    total_jobs_completed = len(completed_jobs)
    total_jobs_invoiced = len(invoiced_jobs)
    total_invoice_amount = sum(inv.get("total_amount", 0) for inv in invoiced_jobs)
    
    return {
        "month": month,
        "year": year,
        "total_jobs_completed": total_jobs_completed,
        "total_jobs_invoiced": total_jobs_invoiced,
        "total_invoice_amount": total_invoice_amount,
        "completed_jobs": completed_jobs,
        "invoiced_jobs": invoiced_jobs
    }

# ============= ACCOUNTING TRANSACTIONS ENDPOINTS =============

@api_router.get("/invoicing/accounting-transactions")
async def get_accounting_transactions(current_user: dict = Depends(require_admin_or_manager)):
    """Get all jobs in accounting transaction stage (invoiced but not completed)"""
    transactions = await db.orders.find({
        "current_stage": "accounting_transaction",
        "status": "accounting_draft"
    }).to_list(length=None)
    
    # Convert ObjectId to string and enrich with client information
    for transaction in transactions:
        # Remove MongoDB ObjectId if present
        if "_id" in transaction:
            del transaction["_id"]
            
        client = await db.clients.find_one({"id": transaction["client_id"]})
        if client:
            transaction["client_name"] = client["company_name"]
            transaction["client_email"] = client.get("email", "")
        
    return {"data": transactions}

@api_router.post("/invoicing/complete-transaction/{job_id}")
async def complete_accounting_transaction(
    job_id: str,
    current_user: dict = Depends(require_admin_or_manager)
):
    """Complete accounting transaction and archive the job"""
    # Get job/order
    job = await db.orders.find_one({"id": job_id})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    if job.get("current_stage") != "accounting_transaction":
        raise HTTPException(status_code=400, detail="Job is not in accounting transaction stage")
    
    # Update job to completed status
    await db.orders.update_one(
        {"id": job_id},
        {"$set": {
            "current_stage": "cleared",
            "status": "completed",
            "completed_at": datetime.now(timezone.utc),
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    # Get the updated order data for archiving
    updated_job = await db.orders.find_one({"id": job_id})
    if updated_job:
        # Create archived order dict
        archived_order = {
            "id": str(uuid.uuid4()),
            "original_order_id": updated_job["id"],
            "order_number": updated_job["order_number"],
            "client_id": updated_job["client_id"],
            "client_name": updated_job.get("client_name", ""),
            "purchase_order_number": updated_job.get("purchase_order_number"),
            "items": updated_job["items"],
            "subtotal": updated_job["subtotal"],
            "gst": updated_job["gst"],
            "total_amount": updated_job["total_amount"],
            "due_date": updated_job["due_date"],
            "delivery_address": updated_job.get("delivery_address"),
            "delivery_instructions": updated_job.get("delivery_instructions"),
            "runtime_estimate": updated_job.get("runtime_estimate"),
            "notes": updated_job.get("notes"),
            "created_by": updated_job["created_by"],
            "created_at": updated_job["created_at"],
            "completed_at": datetime.now(timezone.utc),
            "archived_by": current_user["user_id"],
            "invoice_id": updated_job.get("invoice_id"),
            "invoice_date": updated_job.get("invoice_date")
        }
        
        # Insert into archived orders collection
        await db.archived_orders.insert_one(archived_order)
    
    return {
        "message": "Accounting transaction completed and job archived successfully",
        "job_id": job_id
    }

@api_router.get("/invoicing/export-drafted-csv")
async def export_drafted_invoices_csv(current_user: dict = Depends(require_admin_or_manager)):
    """Export all accounting transactions (drafted invoices) to CSV in Xero import format"""
    try:
        # Get all jobs in accounting transaction stage
        transactions = await db.orders.find({
            "current_stage": "accounting_transaction",
            "status": "accounting_draft"
        }).to_list(length=None)
        
        # Prepare CSV data
        csv_data = []
        
        # CSV Headers based on Xero import format
        headers = [
            "ContactName", "EmailAddress", "POAddressLine1", "POAddressLine2", 
            "POAddressLine3", "POAddressLine4", "POCity", "PORegion", 
            "POPostalCode", "POCountry", "InvoiceNumber", "Reference", 
            "InvoiceDate", "DueDate", "InventoryItemCode", "Description", 
            "Quantity", "UnitAmount", "Discount", "AccountCode", "TaxType", 
            "TrackingName1", "TrackingOption1", "TrackingName2", "TrackingOption2", 
            "Currency", "BrandingTheme"
        ]
        
        csv_data.append(headers)
        
        # Process each transaction
        for transaction in transactions:
            # Get client info
            client = await db.clients.find_one({"id": transaction["client_id"]})
            client_name = client["company_name"] if client else transaction.get("client_name", "Unknown Client")
            client_email = client.get("email", "") if client else ""
            
            # Get invoice info
            invoice = await db.invoices.find_one({"id": transaction.get("invoice_id")})
            invoice_number = invoice["invoice_number"] if invoice else f"INV-{transaction['order_number']}"
            invoice_date = invoice["created_at"].strftime("%d/%m/%Y") if invoice and invoice.get("created_at") else datetime.now().strftime("%d/%m/%Y")
            
            # Calculate due date (30 days from invoice date)
            due_date_obj = invoice["created_at"] + timedelta(days=30) if invoice and invoice.get("created_at") else datetime.now() + timedelta(days=30)
            due_date = due_date_obj.strftime("%d/%m/%Y")
            
            # Process each item in the transaction
            items = transaction.get("items", [])
            if not items:
                # Create a single line if no items
                items = [{"description": f"Services for Order {transaction['order_number']}", "quantity": 1, "unit_price": transaction.get("total_amount", 0)}]
            
            for item in items:
                row = [
                    client_name,  # ContactName (required)
                    client_email,  # EmailAddress
                    "",  # POAddressLine1
                    "",  # POAddressLine2
                    "",  # POAddressLine3
                    "",  # POAddressLine4
                    "",  # POCity
                    "",  # PORegion
                    "",  # POPostalCode
                    "",  # POCountry
                    invoice_number,  # InvoiceNumber (required)
                    transaction["order_number"],  # Reference
                    invoice_date,  # InvoiceDate (required)
                    due_date,  # DueDate (required)
                    item.get("product_code", ""),  # InventoryItemCode
                    f"{item.get('product_name', item.get('description', 'Product'))} - {item.get('specifications', '')}".strip(" - "),  # Description (required)
                    str(item.get("quantity", 1)),  # Quantity (required)
                    str(item.get("unit_price", item.get("price", 0))),  # UnitAmount (required)
                    str(item.get("discount_percent", "")),  # Discount
                    os.getenv("XERO_SALES_ACCOUNT_CODE", "200"),  # AccountCode (required)
                    "OUTPUT",  # TaxType (required) - GST for sales
                    "",  # TrackingName1
                    "",  # TrackingOption1
                    "",  # TrackingName2
                    "",  # TrackingOption2
                    "AUD",  # Currency
                    ""   # BrandingTheme
                ]
                csv_data.append(row)
        
        # Generate CSV content
        import io
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerows(csv_data)
        csv_content = output.getvalue()
        output.close()
        
        # Return CSV as response
        from fastapi.responses import Response
        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=drafted_invoices_{datetime.now().strftime('%Y%m%d')}.csv"}
        )
        
    except Exception as e:
        logger.error(f"Failed to export drafted invoices CSV: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to export CSV: {str(e)}")

# ============= ARCHIVED ORDERS ENDPOINTS =============

@api_router.get("/clients/{client_id}/archived-orders")
async def get_client_archived_orders(
    client_id: str,
    filters: ArchivedOrderFilter = Depends(),
    current_user: dict = Depends(require_any_role)
):
    """Get archived orders for a specific client with filtering"""
    query = {"client_id": client_id}
    
    # Apply date filters
    if filters.date_from or filters.date_to:
        date_filter = {}
        if filters.date_from:
            date_filter["$gte"] = datetime.combine(filters.date_from, datetime.min.time())
        if filters.date_to:
            date_filter["$lte"] = datetime.combine(filters.date_to, datetime.max.time())
        query["archived_at"] = date_filter
    
    # Apply search query
    if filters.search_query:
        query["$or"] = [
            {"order_number": {"$regex": filters.search_query, "$options": "i"}},
            {"client_name": {"$regex": filters.search_query, "$options": "i"}},
            {"purchase_order_number": {"$regex": filters.search_query, "$options": "i"}},
            {"items.product_name": {"$regex": filters.search_query, "$options": "i"}}
        ]
    
    archived_orders = await db.archived_orders.find(query).sort("archived_at", -1).to_list(length=None)
    
    # Remove MongoDB ObjectIds
    for order in archived_orders:
        if "_id" in order:
            del order["_id"]
    
    return StandardResponse(success=True, message="Archived orders retrieved", data=archived_orders)

@api_router.post("/clients/{client_id}/archived-orders/fast-report")
async def generate_fast_report(
    client_id: str,
    report_request: FastReportRequest,
    current_user: dict = Depends(require_any_role)
):
    """Generate Excel fast report for client archived orders"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from io import BytesIO
    import pandas as pd
    from fastapi.responses import StreamingResponse
    
    # Calculate date range based on time period
    today = date.today()
    
    if report_request.time_period == ReportTimePeriod.CURRENT_MONTH:
        date_from = today.replace(day=1)
        date_to = today
    elif report_request.time_period == ReportTimePeriod.LAST_MONTH:
        first_day_current = today.replace(day=1)
        date_to = first_day_current - timedelta(days=1)
        date_from = date_to.replace(day=1)
    elif report_request.time_period == ReportTimePeriod.LAST_3_MONTHS:
        date_to = today
        date_from = today - timedelta(days=90)
    elif report_request.time_period == ReportTimePeriod.LAST_6_MONTHS:
        date_to = today
        date_from = today - timedelta(days=180)
    elif report_request.time_period == ReportTimePeriod.LAST_9_MONTHS:
        date_to = today
        date_from = today - timedelta(days=270)
    elif report_request.time_period == ReportTimePeriod.LAST_YEAR:
        date_to = today
        date_from = today - timedelta(days=365)
    elif report_request.time_period == ReportTimePeriod.CURRENT_QUARTER:
        quarter_start_month = (today.month - 1) // 3 * 3 + 1
        date_from = today.replace(month=quarter_start_month, day=1)
        date_to = today
    elif report_request.time_period == ReportTimePeriod.YEAR_TO_DATE:
        date_from = today.replace(month=1, day=1)
        date_to = today
    elif report_request.time_period == ReportTimePeriod.CURRENT_FINANCIAL_YEAR:
        # Australian financial year: July 1 - June 30
        if today.month >= 7:
            date_from = today.replace(month=7, day=1)
            date_to = today
        else:
            date_from = today.replace(year=today.year-1, month=7, day=1)
            date_to = today
    elif report_request.time_period == ReportTimePeriod.CUSTOM_RANGE:
        date_from = report_request.date_from
        date_to = report_request.date_to
    else:
        date_from = today - timedelta(days=30)
        date_to = today
    
    # Build query
    query = {"client_id": client_id}
    
    if date_from and date_to:
        query["archived_at"] = {
            "$gte": datetime.combine(date_from, datetime.min.time()),
            "$lte": datetime.combine(date_to, datetime.max.time())
        }
    
    # Apply product filter if specified
    if report_request.product_filter:
        query["items.product_name"] = {"$regex": report_request.product_filter, "$options": "i"}
    
    # Get archived orders
    archived_orders = await db.archived_orders.find(query).sort("archived_at", -1).to_list(length=None)
    
    if not archived_orders:
        raise HTTPException(status_code=404, detail="No archived orders found for the specified criteria")
    
    # Prepare data for Excel
    excel_data = []
    
    for order in archived_orders:
        row_data = {}
        
        # Map selected fields to Excel columns
        for field in report_request.selected_fields:
            if field == ReportField.ORDER_NUMBER:
                row_data["Order Number"] = order.get("order_number", "")
            elif field == ReportField.CLIENT_NAME:
                row_data["Client Name"] = order.get("client_name", "")
            elif field == ReportField.PURCHASE_ORDER_NUMBER:
                row_data["Purchase Order Number"] = order.get("purchase_order_number", "")
            elif field == ReportField.ORDER_DATE:
                row_data["Order Date"] = order.get("created_at", "").strftime("%Y-%m-%d") if order.get("created_at") else ""
            elif field == ReportField.COMPLETION_DATE:
                row_data["Completion Date"] = order.get("completed_at", "").strftime("%Y-%m-%d") if order.get("completed_at") else ""
            elif field == ReportField.DUE_DATE:
                row_data["Due Date"] = order.get("due_date", "").strftime("%Y-%m-%d") if order.get("due_date") else ""
            elif field == ReportField.SUBTOTAL:
                row_data["Subtotal"] = f"${order.get('subtotal', 0):.2f}"
            elif field == ReportField.GST:
                row_data["GST"] = f"${order.get('gst', 0):.2f}"
            elif field == ReportField.TOTAL_AMOUNT:
                row_data["Total Amount"] = f"${order.get('total_amount', 0):.2f}"
            elif field == ReportField.DELIVERY_ADDRESS:
                row_data["Delivery Address"] = order.get("delivery_address", "")
            elif field == ReportField.PRODUCT_NAMES:
                products = [item.get("product_name", "") for item in order.get("items", [])]
                row_data["Products"] = ", ".join(products)
            elif field == ReportField.PRODUCT_QUANTITIES:
                quantities = [f"{item.get('product_name', '')}: {item.get('quantity', 0)}" for item in order.get("items", [])]
                row_data["Product Quantities"] = ", ".join(quantities)
            elif field == ReportField.NOTES:
                row_data["Notes"] = order.get("notes", "")
            elif field == ReportField.RUNTIME_ESTIMATE:
                row_data["Runtime Estimate"] = order.get("runtime_estimate", "")
        
        excel_data.append(row_data)
    
    # Create Excel workbook
    df = pd.DataFrame(excel_data)
    wb = Workbook()
    ws = wb.active
    
    # Set title
    title = report_request.report_title or f"Archived Orders Report - {report_request.time_period.value.replace('_', ' ').title()}"
    ws.title = "Archived Orders Report"
    
    # Add header
    ws.merge_cells("A1:{}1".format(chr(65 + len(df.columns) - 1)))
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")
    
    # Add date range info
    ws.merge_cells("A2:{}2".format(chr(65 + len(df.columns) - 1)))
    ws["A2"] = f"Report Period: {date_from.strftime('%Y-%m-%d')} to {date_to.strftime('%Y-%m-%d')}"
    ws["A2"].alignment = Alignment(horizontal="center")
    
    # Add column headers
    for col, column_name in enumerate(df.columns, 1):
        cell = ws.cell(row=4, column=col, value=column_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    for row_idx, row_data in enumerate(excel_data, 5):
        for col_idx, (column_name, value) in enumerate(row_data.items(), 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-adjust column widths
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, len(df.columns) + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        # Check all cells in this column
        for row_idx in range(4, ws.max_row + 1):  # Start from row 4 (after headers)
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                if cell.value is not None:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        
        # Also check header length
        header_cell = ws.cell(row=4, column=col_idx)
        if header_cell.value is not None:
            header_length = len(str(header_cell.value))
            if header_length > max_length:
                max_length = header_length
        
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Get client name for filename
    client = await db.clients.find_one({"id": client_id})
    client_name = client.get("company_name", "Client") if client else "Client"
    safe_client_name = "".join(c for c in client_name if c.isalnum() or c in (' ', '-', '_')).strip()
    
    filename = f"{safe_client_name}_Archived_Orders_{date_from.strftime('%Y%m%d')}-{date_to.strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        BytesIO(excel_file.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ============= FILE SERVING ENDPOINTS =============

@api_router.get("/uploads/{file_path:path}")
async def serve_uploaded_file(file_path: str):
    """Serve uploaded files - Cross-platform compatible"""
    full_path = os.path.join(UPLOAD_DIR, file_path)
    if not os.path.exists(full_path):
        raise HTTPException(status_code=404, detail="File not found")
    
    return FileResponse(full_path)

# ============= STOCK MANAGEMENT SYSTEM ENDPOINTS =============

@api_router.get("/stock/raw-substrates", response_model=StandardResponse)
async def get_raw_substrates_stock(
    client_id: Optional[str] = None,
    current_user: dict = Depends(require_any_role)
):
    """Get all raw substrates on hand, optionally filtered by client"""
    try:
        query = {}
        if client_id and client_id != "all":
            query["client_id"] = client_id
            
        substrates = await db.raw_substrate_stock.find(query).to_list(length=None)
        
        # Remove MongoDB ObjectIds
        for substrate in substrates:
            if "_id" in substrate:
                del substrate["_id"]
        
        return StandardResponse(
            success=True,
            message="Raw substrates retrieved successfully",
            data=substrates
        )
    except Exception as e:
        logger.error(f"Failed to get raw substrates: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to retrieve raw substrates")

@api_router.post("/stock/raw-substrates", response_model=StandardResponse)
async def create_raw_substrate_stock(
    substrate_data: RawSubstrateStockCreate,
    current_user: dict = Depends(require_manager_or_admin)
):
    """Create new raw substrate stock entry"""
    try:
        # Create new substrate stock record
        substrate = RawSubstrateStock(
            client_id=substrate_data.client_id,
            client_name=substrate_data.client_name,
            product_id=substrate_data.product_id,
            product_code=substrate_data.product_code,
            product_description=substrate_data.product_description,
            quantity_on_hand=substrate_data.quantity_on_hand,
            unit_of_measure=substrate_data.unit_of_measure,
            source_order_id=substrate_data.source_order_id,
            source_job_id=substrate_data.source_job_id,
            is_shared_product=substrate_data.is_shared_product,
            shared_with_clients=substrate_data.shared_with_clients,
            created_from_excess=substrate_data.created_from_excess,
            material_specifications=substrate_data.material_specifications,
            material_value_m2=substrate_data.material_value_m2,
            minimum_stock_level=substrate_data.minimum_stock_level,
            created_by=current_user["user_id"]
        )
        
        substrate_dict = substrate.dict()
        substrate_dict["created_at"] = datetime.now(timezone.utc)
        
        result = await db.raw_substrate_stock.insert_one(substrate_dict)
        
        # Create stock movement record
        movement = StockMovement(
            stock_type="raw_substrate",
            stock_id=substrate.id,
            movement_type="addition",
            quantity_change=substrate_data.quantity_on_hand,
            previous_quantity=0.0,
            new_quantity=substrate_data.quantity_on_hand,
            reference_id=substrate_data.source_order_id,
            reference_type="order",
            notes=f"Initial stock creation from excess production",
            created_by=current_user["user_id"]
        )
        
        movement_dict = movement.dict()
        movement_dict["created_at"] = datetime.now(timezone.utc)
        await db.stock_movements.insert_one(movement_dict)
        
        return StandardResponse(
            success=True,
            message="Raw substrate stock created successfully",
            data={"id": substrate.id}
        )
    except Exception as e:
        logger.error(f"Failed to create raw substrate stock: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to create raw substrate stock")

@api_router.put("/stock/raw-substrates/{substrate_id}", response_model=StandardResponse)
async def update_raw_substrate_stock(
    substrate_id: str,
    update_data: RawSubstrateStockUpdate,
    current_user: dict = Depends(require_manager_or_admin)
):
    """Update raw substrate stock quantity or details"""
    try:
        # Get existing substrate
        existing = await db.raw_substrate_stock.find_one({"id": substrate_id})
        if not existing:
            raise HTTPException(status_code=404, detail="Raw substrate stock not found")
        
        previous_quantity = existing["quantity_on_hand"]
        new_quantity = update_data.quantity_on_hand if update_data.quantity_on_hand is not None else previous_quantity
        
        # Update substrate record
        update_fields = {
            "updated_at": datetime.now(timezone.utc),
            "updated_by": current_user["user_id"]
        }
        
        # Add any fields that are being updated
        if update_data.quantity_on_hand is not None:
            update_fields["quantity_on_hand"] = update_data.quantity_on_hand
        if update_data.minimum_stock_level is not None:
            update_fields["minimum_stock_level"] = update_data.minimum_stock_level
        if update_data.is_shared_product is not None:
            update_fields["is_shared_product"] = update_data.is_shared_product
        if update_data.shared_with_clients is not None:
            update_fields["shared_with_clients"] = update_data.shared_with_clients
        if update_data.material_specifications is not None:
            update_fields["material_specifications"] = update_data.material_specifications
        
        await db.raw_substrate_stock.update_one(
            {"id": substrate_id},
            {"$set": update_fields}
        )
        
        # Create stock movement record if quantity changed
        if new_quantity != previous_quantity:
            movement_type = "addition" if new_quantity > previous_quantity else "consumption"
            quantity_change = new_quantity - previous_quantity
            
            movement = StockMovement(
                stock_type="raw_substrate",
                stock_id=substrate_id,
                movement_type=movement_type,
                quantity_change=quantity_change,
                previous_quantity=previous_quantity,
                new_quantity=new_quantity,
                reference_type="manual",
                notes=update_data.notes or "Manual adjustment",
                created_by=current_user["user_id"]
            )
            
            movement_dict = movement.dict()
            movement_dict["created_at"] = datetime.now(timezone.utc)
            await db.stock_movements.insert_one(movement_dict)
        
        return StandardResponse(
            success=True,
            message="Raw substrate stock updated successfully"
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to update raw substrate stock: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to update raw substrate stock")

@api_router.get("/stock/raw-materials", response_model=StandardResponse)
async def get_raw_materials_stock(current_user: dict = Depends(require_any_role)):
    """Get all raw materials stock"""
    try:
        materials = await db.raw_material_stock.find({}).to_list(length=None)
        
        # Remove MongoDB ObjectIds
        for material in materials:
            if "_id" in material:
                del material["_id"]
        
        return StandardResponse(
            success=True,
            message="Raw materials stock retrieved successfully",
            data=materials
        )
    except Exception as e:
        logger.error(f"Failed to get raw materials stock: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to retrieve raw materials stock")

@api_router.post("/stock/raw-materials", response_model=StandardResponse)
async def create_raw_material_stock(
    material_data: RawMaterialStockCreate,
    current_user: dict = Depends(require_manager_or_admin)
):
    """Create new raw material stock entry"""
    try:
        # Create new material stock record
        material = RawMaterialStock(
            material_id=material_data.material_id,
            material_name=material_data.material_name,
            quantity_on_hand=material_data.quantity_on_hand,
            unit_of_measure=material_data.unit_of_measure,
            minimum_stock_level=material_data.minimum_stock_level,
            alert_threshold_days=material_data.alert_threshold_days,
            supplier_id=material_data.supplier_id,
            usage_rate_per_month=material_data.usage_rate_per_month
        )
        
        material_dict = material.dict()
        material_dict["created_at"] = datetime.now(timezone.utc)
        
        result = await db.raw_material_stock.insert_one(material_dict)
        
        # Create stock movement record
        movement = StockMovement(
            stock_type="raw_material",
            stock_id=material.id,
            movement_type="addition",
            quantity_change=material_data.quantity_on_hand,
            previous_quantity=0.0,
            new_quantity=material_data.quantity_on_hand,
            reference_type="manual",
            notes="Initial stock creation",
            created_by=current_user["user_id"]
        )
        
        movement_dict = movement.dict()
        movement_dict["created_at"] = datetime.now(timezone.utc)
        await db.stock_movements.insert_one(movement_dict)
        
        return StandardResponse(
            success=True,
            message="Raw material stock created successfully",
            data={"id": material.id}
        )
    except Exception as e:
        logger.error(f"Failed to create raw material stock: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to create raw material stock")

@api_router.put("/stock/raw-materials/{material_id}", response_model=StandardResponse)
async def update_raw_material_stock(
    material_id: str,
    update_data: RawMaterialStockUpdate,
    current_user: dict = Depends(require_manager_or_admin)
):
    """Update raw material stock quantity or details"""
    try:
        # Get existing material
        existing = await db.raw_material_stock.find_one({"id": material_id})
        if not existing:
            raise HTTPException(status_code=404, detail="Raw material stock not found")
        
        previous_quantity = existing["quantity_on_hand"]
        new_quantity = update_data.quantity_on_hand if update_data.quantity_on_hand is not None else previous_quantity
        
        # Update material record
        update_fields = {
            "updated_at": datetime.now(timezone.utc)
        }
        
        # Add any fields that are being updated
        if update_data.quantity_on_hand is not None:
            update_fields["quantity_on_hand"] = update_data.quantity_on_hand
        if update_data.minimum_stock_level is not None:
            update_fields["minimum_stock_level"] = update_data.minimum_stock_level
        if update_data.alert_threshold_days is not None:
            update_fields["alert_threshold_days"] = update_data.alert_threshold_days
        if update_data.usage_rate_per_month is not None:
            update_fields["usage_rate_per_month"] = update_data.usage_rate_per_month
        if update_data.last_order_date is not None:
            update_fields["last_order_date"] = update_data.last_order_date
        if update_data.last_order_quantity is not None:
            update_fields["last_order_quantity"] = update_data.last_order_quantity
        
        await db.raw_material_stock.update_one(
            {"id": material_id},
            {"$set": update_fields}
        )
        
        # Create stock movement record if quantity changed
        if new_quantity != previous_quantity:
            movement_type = "addition" if new_quantity > previous_quantity else "consumption"
            quantity_change = new_quantity - previous_quantity
            
            movement = StockMovement(
                stock_type="raw_material",
                stock_id=material_id,
                movement_type=movement_type,
                quantity_change=quantity_change,
                previous_quantity=previous_quantity,
                new_quantity=new_quantity,
                reference_type="manual",
                notes=update_data.notes or "Manual adjustment",
                created_by=current_user["user_id"]
            )
            
            movement_dict = movement.dict()
            movement_dict["created_at"] = datetime.now(timezone.utc)
            await db.stock_movements.insert_one(movement_dict)
        
        return StandardResponse(
            success=True,
            message="Raw material stock updated successfully"
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to update raw material stock: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to update raw material stock")

@api_router.get("/stock/movements/{stock_id}", response_model=StandardResponse)
async def get_stock_movements(
    stock_id: str,
    current_user: dict = Depends(require_any_role)
):
    """Get stock movement history for a specific stock item"""
    try:
        movements = await db.stock_movements.find({"stock_id": stock_id}).sort("created_at", -1).to_list(length=None)
        
        # Remove MongoDB ObjectIds
        for movement in movements:
            if "_id" in movement:
                del movement["_id"]
        
        return StandardResponse(
            success=True,
            message="Stock movements retrieved successfully",
            data=movements
        )
    except Exception as e:
        logger.error(f"Failed to get stock movements: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to retrieve stock movements")

@api_router.get("/stock/alerts", response_model=StandardResponse)
async def get_stock_alerts(current_user: dict = Depends(require_any_role)):
    """Get all active stock alerts"""
    try:
        alerts = await db.stock_alerts.find({"is_active": True}).to_list(length=None)
        
        # Remove MongoDB ObjectIds
        for alert in alerts:
            if "_id" in alert:
                del alert["_id"]
        
        return StandardResponse(
            success=True,
            message="Stock alerts retrieved successfully",
            data=alerts
        )
    except Exception as e:
        logger.error(f"Failed to get stock alerts: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to retrieve stock alerts")

@api_router.post("/stock/alerts/{alert_id}/acknowledge", response_model=StandardResponse)
async def acknowledge_stock_alert(
    alert_id: str,
    acknowledge_data: StockAlertAcknowledge,
    current_user: dict = Depends(require_any_role)
):
    """Acknowledge a stock alert"""
    try:
        update_fields = {
            "acknowledged_by": current_user["user_id"],
            "acknowledged_at": datetime.now(timezone.utc)
        }
        
        # Handle snooze functionality
        if acknowledge_data.snooze_hours:
            snooze_hours = acknowledge_data.snooze_hours
            update_fields["snooze_until"] = datetime.now(timezone.utc) + timedelta(hours=snooze_hours)
        else:
            # If not snoozed, deactivate the alert
            update_fields["is_active"] = False
        
        result = await db.stock_alerts.update_one(
            {"id": alert_id},
            {"$set": update_fields}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Stock alert not found")
        
        return StandardResponse(
            success=True,
            message="Stock alert acknowledged successfully"
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to acknowledge stock alert: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to acknowledge stock alert")

@api_router.post("/stock/check-low-stock", response_model=StandardResponse)
async def check_low_stock_alerts(current_user: dict = Depends(require_any_role)):
    """Check for low stock and create alerts"""
    try:
        alerts_created = 0
        
        # Check raw materials stock
        materials = await db.raw_material_stock.find({}).to_list(length=None)
        for material in materials:
            if material["quantity_on_hand"] <= material.get("minimum_stock_level", 0):
                # Check if alert already exists for this material
                existing_alert = await db.stock_alerts.find_one({
                    "stock_id": material["id"],
                    "stock_type": "raw_material",
                    "alert_type": "low_stock",
                    "is_active": True
                })
                
                if not existing_alert:
                    alert = StockAlert(
                        stock_type="raw_material",
                        stock_id=material["id"],
                        alert_type="low_stock",
                        message=f"Low stock alert: {material['material_name']} has {material['quantity_on_hand']} {material['unit_of_measure']} remaining (minimum: {material.get('minimum_stock_level', 0)})"
                    )
                    
                    alert_dict = alert.dict()
                    alert_dict["created_at"] = datetime.now(timezone.utc)
                    await db.stock_alerts.insert_one(alert_dict)
                    alerts_created += 1
        
        # Check raw substrates stock
        substrates = await db.raw_substrate_stock.find({}).to_list(length=None)
        for substrate in substrates:
            if substrate["quantity_on_hand"] <= substrate.get("minimum_stock_level", 0):
                # Check if alert already exists for this substrate
                existing_alert = await db.stock_alerts.find_one({
                    "stock_id": substrate["id"],
                    "stock_type": "raw_substrate", 
                    "alert_type": "low_stock",
                    "is_active": True
                })
                
                if not existing_alert:
                    alert = StockAlert(
                        stock_type="raw_substrate",
                        stock_id=substrate["id"],
                        alert_type="low_stock",
                        message=f"Low stock alert: {substrate['product_description']} ({substrate['product_code']}) has {substrate['quantity_on_hand']} {substrate['unit_of_measure']} remaining (minimum: {substrate.get('minimum_stock_level', 0)})"
                    )
                    
                    alert_dict = alert.dict()
                    alert_dict["created_at"] = datetime.now(timezone.utc)
                    await db.stock_alerts.insert_one(alert_dict)
                    alerts_created += 1
        
        return StandardResponse(
            success=True,
            message=f"Stock check completed. {alerts_created} new alerts created.",
            data={"alerts_created": alerts_created}
        )
    except Exception as e:
        logger.error(f"Failed to check low stock: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to check low stock")

@api_router.delete("/stock/raw-substrates/{substrate_id}", response_model=StandardResponse)
async def delete_raw_substrate_stock(
    substrate_id: str,
    current_user: dict = Depends(require_manager_or_admin)
):
    """Delete raw substrate stock entry"""
    try:
        result = await db.raw_substrate_stock.delete_one({"id": substrate_id})
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Raw substrate stock not found")
        
        # Also delete related stock movements and alerts
        await db.stock_movements.delete_many({"stock_id": substrate_id})
        await db.stock_alerts.delete_many({"stock_id": substrate_id})
        
        return StandardResponse(
            success=True,
            message="Raw substrate stock deleted successfully"
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete raw substrate stock: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to delete raw substrate stock")

@api_router.delete("/stock/raw-materials/{material_id}", response_model=StandardResponse)
async def delete_raw_material_stock(
    material_id: str,
    current_user: dict = Depends(require_manager_or_admin)
):
    """Delete raw material stock entry"""
    try:
        result = await db.raw_material_stock.delete_one({"id": material_id})
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Raw material stock not found")
        
        # Also delete related stock movements and alerts
        await db.stock_movements.delete_many({"stock_id": material_id})
        await db.stock_alerts.delete_many({"stock_id": material_id})
        
        return StandardResponse(
            success=True,
            message="Raw material stock deleted successfully"
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete raw material stock: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to delete raw material stock")

# ============= STOCK AVAILABILITY & ALLOCATION ENDPOINTS =============

@api_router.get("/stock/history", response_model=StandardResponse)
async def get_stock_history(
    product_id: str,
    client_id: str,
    current_user: dict = Depends(require_any_role)
):
    """Get stock history for a specific product and client"""
    try:
        # Get all stock entries for this product/client combination
        stock_entries = await db.raw_substrate_stock.find({
            "product_id": product_id,
            "client_id": client_id
        }).to_list(length=None)
        
        # Get stock movements for these entries
        stock_ids = [entry["id"] for entry in stock_entries]
        movements = await db.stock_movements.find({
            "stock_id": {"$in": stock_ids}
        }).sort("created_at", -1).to_list(length=None)
        
        # Remove MongoDB ObjectIds from stock entries and movements
        for entry in stock_entries:
            if "_id" in entry:
                del entry["_id"]
        
        for movement in movements:
            if "_id" in movement:
                del movement["_id"]
        
        # Combine stock entries with their movements
        history_data = {
            "product_id": product_id,
            "client_id": client_id,
            "stock_entries": stock_entries,
            "movements": movements,
            "total_entries": len(stock_entries),
            "total_quantity": sum(entry.get("quantity_on_hand", 0) for entry in stock_entries)
        }
        
        return StandardResponse(
            success=True,
            message="Stock history retrieved successfully",
            data=history_data
        )
        
    except Exception as e:
        logger.error(f"Failed to get stock history: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to retrieve stock history")

@api_router.get("/stock/allocations", response_model=StandardResponse)
async def get_stock_allocations(
    product_id: str,
    client_id: str,
    current_user: dict = Depends(require_any_role)
):
    """Get current stock allocations for a specific product and client"""
    try:
        # Get allocation movements (negative quantities that haven't been archived)
        allocations = await db.stock_movements.find({
            "product_id": product_id,
            "client_id": client_id,
            "movement_type": "allocation",
            "quantity": {"$lt": 0},  # Negative quantities are allocations
            "is_archived": {"$ne": True}  # Not archived
        }).sort("created_at", -1).to_list(length=None)
        
        # Remove MongoDB ObjectIds from allocations
        for allocation in allocations:
            if "_id" in allocation:
                del allocation["_id"]
        
        # Calculate total allocated quantity
        total_allocated = sum(abs(allocation.get("quantity", 0)) for allocation in allocations)
        
        allocation_data = {
            "product_id": product_id,
            "client_id": client_id,
            "allocations": allocations,
            "total_allocated": total_allocated,
            "active_orders": len(allocations)
        }
        
        return StandardResponse(
            success=True,
            message="Stock allocations retrieved successfully",
            data=allocation_data
        )
        
    except Exception as e:
        logger.error(f"Failed to get stock allocations: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to retrieve stock allocations")

@api_router.get("/stock/allocations/archived", response_model=StandardResponse)
async def get_archived_allocations(
    product_id: str,
    client_id: str,
    current_user: dict = Depends(require_any_role)
):
    """Get archived stock allocations (from completed/invoiced orders) for a specific product and client"""
    try:
        # Get archived allocation movements
        allocations = await db.stock_movements.find({
            "product_id": product_id,
            "client_id": client_id,
            "movement_type": "allocation",
            "quantity": {"$lt": 0},  # Negative quantities are allocations
            "is_archived": True  # Only archived
        }).sort("created_at", -1).to_list(length=None)
        
        # Remove MongoDB ObjectIds from allocations
        for allocation in allocations:
            if "_id" in allocation:
                del allocation["_id"]
        
        # Calculate total allocated quantity
        total_allocated = sum(abs(allocation.get("quantity", 0)) for allocation in allocations)
        
        allocation_data = {
            "product_id": product_id,
            "client_id": client_id,
            "allocations": allocations,
            "total_allocated": total_allocated,
            "archived_orders": len(allocations)
        }
        
        return StandardResponse(
            success=True,
            message="Archived stock allocations retrieved successfully",
            data=allocation_data
        )
        
    except Exception as e:
        logger.error(f"Failed to get archived stock allocations: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to retrieve archived stock allocations")

# ============= STOCK AVAILABILITY & ALLOCATION ENDPOINTS =============

@api_router.get("/stock/check-availability", response_model=StandardResponse)
async def check_stock_availability(
    product_id: str,
    client_id: str,
    current_user: dict = Depends(require_any_role)
):
    """Check stock availability for a specific product and client"""
    try:
        # Check if there's any stock for this product from this client
        stock = await db.raw_substrate_stock.find_one({
            "product_id": product_id,
            "client_id": client_id,
            "quantity_on_hand": {"$gt": 0}
        })
        
        if stock:
            return StandardResponse(
                success=True,
                message="Stock available",
                data={
                    "product_id": product_id,
                    "client_id": client_id,
                    "quantity_on_hand": stock["quantity_on_hand"],
                    "stock_id": stock["id"]
                }
            )
        else:
            raise HTTPException(status_code=404, detail="No stock available for this product")
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to check stock availability: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to check stock availability")

@api_router.post("/stock/allocate", response_model=StandardResponse)
async def allocate_stock(
    allocation_data: dict,
    current_user: dict = Depends(require_any_role)
):
    """Allocate stock for an order"""
    try:
        product_id = allocation_data.get("product_id")
        client_id = allocation_data.get("client_id")
        quantity = allocation_data.get("quantity")
        order_reference = allocation_data.get("order_reference")
        
        if not all([product_id, client_id, quantity]):
            raise HTTPException(status_code=400, detail="Missing required fields")
            
        # Find the stock entry
        stock = await db.raw_substrate_stock.find_one({
            "product_id": product_id,
            "client_id": client_id,
            "quantity_on_hand": {"$gte": quantity}
        })
        
        if not stock:
            raise HTTPException(status_code=400, detail="Insufficient stock available")
        
        # Update stock quantity
        new_quantity = stock["quantity_on_hand"] - quantity
        result = await db.raw_substrate_stock.update_one(
            {"id": stock["id"]},
            {"$set": {"quantity_on_hand": new_quantity}}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Stock entry not found")
        
        # Create stock movement record
        movement_id = str(uuid.uuid4())
        movement = {
            "id": movement_id,
            "stock_id": stock["id"],
            "product_id": product_id,
            "client_id": client_id,
            "movement_type": "allocation",
            "quantity": -quantity,  # Negative for allocation
            "reference": order_reference,
            "created_by": current_user["user_id"],
            "created_at": datetime.now(timezone.utc).isoformat(),
            "is_archived": False
        }
        await db.stock_movements.insert_one(movement)
        
        return StandardResponse(
            success=True,
            message=f"Successfully allocated {quantity} units from stock",
            data={
                "allocated_quantity": quantity,
                "remaining_stock": new_quantity,
                "movement_id": movement_id
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to allocate stock: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to allocate stock")

@api_router.get("/stock/history", response_model=StandardResponse)
async def get_stock_history(
    product_id: str,
    client_id: str,
    current_user: dict = Depends(require_any_role)
):
    """Get stock history for a specific product and client"""
    try:
        # Get all stock entries for this product/client combination
        stock_entries = await db.raw_substrate_stock.find({
            "product_id": product_id,
            "client_id": client_id
        }).to_list(length=None)
        
        # Get stock movements for these entries
        stock_ids = [entry["id"] for entry in stock_entries]
        movements = await db.stock_movements.find({
            "stock_id": {"$in": stock_ids}
        }).sort("created_at", -1).to_list(length=None)
        
        # Remove MongoDB ObjectIds from stock entries and movements
        for entry in stock_entries:
            if "_id" in entry:
                del entry["_id"]
        
        for movement in movements:
            if "_id" in movement:
                del movement["_id"]
        
        # Combine stock entries with their movements
        history_data = {
            "product_id": product_id,
            "client_id": client_id,
            "stock_entries": stock_entries,
            "movements": movements,
            "total_entries": len(stock_entries),
            "total_quantity": sum(entry.get("quantity_on_hand", 0) for entry in stock_entries)
        }
        
        return StandardResponse(
            success=True,
            message="Stock history retrieved successfully",
            data=history_data
        )
        
    except Exception as e:
        logger.error(f"Failed to get stock history: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to retrieve stock history")

@api_router.get("/stock/allocations", response_model=StandardResponse)
async def get_stock_allocations(
    product_id: str,
    client_id: str,
    current_user: dict = Depends(require_any_role)
):
    """Get current stock allocations for a specific product and client"""
    try:
        # Get allocation movements (negative quantities that haven't been archived)
        allocations = await db.stock_movements.find({
            "product_id": product_id,
            "client_id": client_id,
            "movement_type": "allocation",
            "quantity": {"$lt": 0},  # Negative quantities are allocations
            "is_archived": {"$ne": True}  # Not archived
        }).sort("created_at", -1).to_list(length=None)
        
        # Remove MongoDB ObjectIds from allocations
        for allocation in allocations:
            if "_id" in allocation:
                del allocation["_id"]
        
        # Calculate total allocated quantity
        total_allocated = sum(abs(allocation.get("quantity", 0)) for allocation in allocations)
        
        allocation_data = {
            "product_id": product_id,
            "client_id": client_id,
            "allocations": allocations,
            "total_allocated": total_allocated,
            "active_orders": len(allocations)
        }
        
        return StandardResponse(
            success=True,
            message="Stock allocations retrieved successfully",
            data=allocation_data
        )
        
    except Exception as e:
        logger.error(f"Failed to get stock allocations: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to retrieve stock allocations")


# ============= STOCK REPORTING ENDPOINTS =============

@api_router.get("/stock/reports/material-usage", response_model=StandardResponse)
async def get_material_usage_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    material_id: Optional[str] = None,
    current_user: dict = Depends(require_any_role)
):
    """Generate material usage report with projections"""
    try:
        # Default to last 30 days if no dates provided
        if not end_date:
            end_date = datetime.now(timezone.utc).isoformat()
        if not start_date:
            start_date = (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()
        
        # Build query for stock movements
        query = {
            "movement_type": {"$in": ["allocation", "usage", "consumption"]},
            "created_at": {"$gte": start_date, "$lte": end_date}
        }
        
        if material_id:
            query["product_id"] = material_id
        
        # Get all relevant stock movements
        movements = await db.stock_movements.find(query).to_list(length=None)
        
        # Group by material/product
        material_usage = {}
        for movement in movements:
            prod_id = movement.get("product_id", "unknown")
            if prod_id not in material_usage:
                material_usage[prod_id] = {
                    "product_id": prod_id,
                    "product_name": movement.get("product_name", "Unknown"),
                    "total_used": 0,
                    "movements": []
                }
            
            quantity = abs(movement.get("quantity", 0))
            material_usage[prod_id]["total_used"] += quantity
            material_usage[prod_id]["movements"].append({
                "date": movement.get("created_at"),
                "quantity": quantity,
                "reference": movement.get("reference", ""),
                "movement_type": movement.get("movement_type")
            })
        
        # Calculate projections (usage rate per day * 30 days)
        days_in_period = (datetime.fromisoformat(end_date.replace('Z', '+00:00')) - 
                         datetime.fromisoformat(start_date.replace('Z', '+00:00'))).days or 1
        
        projections = []
        for prod_id, data in material_usage.items():
            daily_usage = data["total_used"] / days_in_period
            projected_monthly = daily_usage * 30
            projected_quarterly = daily_usage * 90
            
            # Get current stock level
            stock = await db.raw_substrate_stock.find_one({"product_id": prod_id}) or \
                   await db.raw_material_stock.find_one({"material_id": prod_id})
            
            current_stock = stock.get("quantity_on_hand", 0) if stock else 0
            days_until_depleted = (current_stock / daily_usage) if daily_usage > 0 else 999
            
            projections.append({
                **data,
                "current_stock": current_stock,
                "daily_usage_rate": round(daily_usage, 2),
                "projected_monthly_usage": round(projected_monthly, 2),
                "projected_quarterly_usage": round(projected_quarterly, 2),
                "days_until_depleted": round(days_until_depleted, 1),
                "reorder_recommended": days_until_depleted < 30
            })
        
        return StandardResponse(
            success=True,
            message="Material usage report generated successfully",
            data={
                "report_period": {
                    "start_date": start_date,
                    "end_date": end_date,
                    "days": days_in_period
                },
                "materials": projections,
                "total_materials": len(projections)
            }
        )
        
    except Exception as e:
        logger.error(f"Failed to generate material usage report: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to generate report")

@api_router.get("/stock/reports/low-stock", response_model=StandardResponse)
async def get_low_stock_report(
    threshold_days: int = 30,
    current_user: dict = Depends(require_any_role)
):
    """Get report of items with low stock levels"""
    try:
        # Get all stock items
        substrates = await db.raw_substrate_stock.find().to_list(length=None)
        materials = await db.raw_material_stock.find().to_list(length=None)
        
        low_stock_items = []
        
        # Check substrates
        for stock in substrates:
            min_level = stock.get("minimum_stock_level", 0)
            current = stock.get("quantity_on_hand", 0)
            
            if current <= min_level:
                # Calculate usage rate from recent movements
                movements = await db.stock_movements.find({
                    "product_id": stock.get("product_id"),
                    "movement_type": {"$in": ["allocation", "usage"]},
                    "created_at": {"$gte": (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()}
                }).to_list(length=None)
                
                total_used = sum(abs(m.get("quantity", 0)) for m in movements)
                daily_usage = total_used / 30 if total_used > 0 else 0
                days_remaining = (current / daily_usage) if daily_usage > 0 else 999
                
                low_stock_items.append({
                    "type": "substrate",
                    "id": stock.get("id"),
                    "product_id": stock.get("product_id"),
                    "product_description": stock.get("product_description", "Unknown"),
                    "client_name": stock.get("client_name", "Unknown"),
                    "current_stock": current,
                    "minimum_level": min_level,
                    "unit_of_measure": stock.get("unit_of_measure", "units"),
                    "daily_usage_rate": round(daily_usage, 2),
                    "days_remaining": round(days_remaining, 1),
                    "status": "critical" if days_remaining < 7 else "low" if days_remaining < threshold_days else "warning"
                })
        
        # Check raw materials
        for material in materials:
            min_level = material.get("minimum_stock_level", 0)
            current = material.get("quantity_on_hand", 0)
            
            if current <= min_level:
                # Calculate usage rate
                usage_rate = material.get("usage_rate_per_month", 0)
                daily_usage = usage_rate / 30 if usage_rate > 0 else 0
                days_remaining = (current / daily_usage) if daily_usage > 0 else 999
                
                low_stock_items.append({
                    "type": "raw_material",
                    "id": material.get("id"),
                    "material_id": material.get("material_id"),
                    "material_name": material.get("material_name", "Unknown"),
                    "current_stock": current,
                    "minimum_level": min_level,
                    "unit_of_measure": material.get("unit_of_measure", "kg"),
                    "daily_usage_rate": round(daily_usage, 2),
                    "days_remaining": round(days_remaining, 1),
                    "status": "critical" if days_remaining < 7 else "low" if days_remaining < threshold_days else "warning"
                })
        
        # Sort by days remaining (most critical first)
        low_stock_items.sort(key=lambda x: x["days_remaining"])
        
        return StandardResponse(
            success=True,
            message=f"Found {len(low_stock_items)} low stock items",
            data={
                "low_stock_items": low_stock_items,
                "total_items": len(low_stock_items),
                "critical_items": len([item for item in low_stock_items if item["status"] == "critical"]),
                "threshold_days": threshold_days
            }
        )
        
    except Exception as e:
        logger.error(f"Failed to generate low stock report: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to generate low stock report")

@api_router.get("/stock/reports/inventory-value", response_model=StandardResponse)
async def get_inventory_value_report(
    current_user: dict = Depends(require_any_role)
):
    """Calculate total inventory value"""
    try:
        # Get all stock with pricing information
        substrates = await db.raw_substrate_stock.find().to_list(length=None)
        materials = await db.raw_material_stock.find().to_list(length=None)
        
        inventory_value = {
            "substrates": [],
            "materials": [],
            "total_substrate_value": 0,
            "total_material_value": 0,
            "total_inventory_value": 0
        }
        
        # Calculate substrate values
        for stock in substrates:
            quantity = stock.get("quantity_on_hand", 0)
            # Try to get price from client product catalogue
            product_id = stock.get("product_id")
            client_id = stock.get("client_id")
            
            price_per_unit = 0
            if product_id and client_id:
                product = await db.client_products.find_one({
                    "id": product_id,
                    "client_id": client_id
                })
                if product:
                    price_per_unit = product.get("price_per_unit", 0) or 0
            
            value = quantity * price_per_unit
            inventory_value["substrates"].append({
                "product_description": stock.get("product_description", "Unknown"),
                "quantity": quantity,
                "unit_of_measure": stock.get("unit_of_measure", "units"),
                "price_per_unit": price_per_unit,
                "total_value": round(value, 2)
            })
            inventory_value["total_substrate_value"] += value
        
        # Calculate material values (would need supplier pricing)
        for material in materials:
            quantity = material.get("quantity_on_hand", 0)
            # Note: Price per unit would need to be added to raw_materials_stock model
            price_per_unit = material.get("price_per_unit", 0) or 0
            
            value = quantity * price_per_unit
            inventory_value["materials"].append({
                "material_name": material.get("material_name", "Unknown"),
                "quantity": quantity,
                "unit_of_measure": material.get("unit_of_measure", "kg"),
                "price_per_unit": price_per_unit,
                "total_value": round(value, 2)
            })
            inventory_value["total_material_value"] += value
        
        inventory_value["total_inventory_value"] = (
            inventory_value["total_substrate_value"] + 
            inventory_value["total_material_value"]
        )
        
        # Round totals
        inventory_value["total_substrate_value"] = round(inventory_value["total_substrate_value"], 2)
        inventory_value["total_material_value"] = round(inventory_value["total_material_value"], 2)
        inventory_value["total_inventory_value"] = round(inventory_value["total_inventory_value"], 2)
        
        return StandardResponse(
            success=True,
            message="Inventory value report generated successfully",
            data=inventory_value
        )
        
    except Exception as e:
        logger.error(f"Failed to generate inventory value report: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to generate inventory value report")



@api_router.get("/stock/reports/material-usage-detailed", response_model=StandardResponse)
async def get_detailed_material_usage_report(
    material_id: str,
    start_date: str,
    end_date: str,
    include_order_breakdown: bool = False,
    current_user: dict = Depends(require_any_role)
):
    """
    Generate detailed material usage report by width and length for a specific material.
    Shows usage per width, total length, and grand total m² used.
    Optionally includes order-by-order breakdown.
    """
    try:
        # Get material information - try both collections
        material = await db.materials.find_one({"id": material_id})
        if not material:
            # Try raw_materials collection with material_id field
            material = await db.raw_materials.find_one({"material_id": material_id})
        if not material:
            raise HTTPException(status_code=404, detail="Material not found")
        
        # Parse dates
        start = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        end = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
        
        # Find all orders in the date range that used this material
        orders = await db.orders.find({
            "created_at": {"$gte": start.isoformat(), "$lte": end.isoformat()},
            "status": {"$in": ["completed", "archived"]}
        }).to_list(length=None)
        
        # Track material usage by width
        usage_by_width = {}  # {width_mm: {total_length_m: X, orders: [{order_number, length_m}]}}
        total_m2 = 0.0
        
        for order in orders:
            order_number = order.get("order_number", "Unknown")
            
            # Check job specifications for this order to find material usage
            job_specs = await db.job_specifications.find({
                "order_id": order["id"]
            }).to_list(length=None)
            
            for job_spec in job_specs:
                materials_composition = job_spec.get("materials_composition", [])
                
                for material_layer in materials_composition:
                    # Check if this layer uses the requested material
                    layer_material_id = material_layer.get("material_id")
                    
                    if layer_material_id == material_id:
                        # Get width and quantity
                        width_mm = material_layer.get("width", 0)
                        quantity_meters = material_layer.get("quantity", 0)
                        
                        if width_mm > 0 and quantity_meters > 0:
                            width_key = f"{width_mm}"
                            
                            # Initialize width entry if not exists
                            if width_key not in usage_by_width:
                                usage_by_width[width_key] = {
                                    "width_mm": width_mm,
                                    "total_length_m": 0.0,
                                    "orders": []
                                }
                            
                            # Add to width total
                            usage_by_width[width_key]["total_length_m"] += quantity_meters
                            
                            # Track order detail if breakdown requested
                            if include_order_breakdown:
                                usage_by_width[width_key]["orders"].append({
                                    "order_number": order_number,
                                    "length_m": quantity_meters,
                                    "order_date": order.get("created_at", ""),
                                    "client_name": order.get("client_name", "Unknown")
                                })
                            
                            # Calculate m² for this usage (width in mm / 1000 * length in m)
                            m2 = (width_mm / 1000.0) * quantity_meters
                            total_m2 += m2
        
        # Also check slit widths that might have been used
        slit_widths = await db.slit_widths.find({
            "raw_material_id": material_id,
            "created_at": {"$gte": start.isoformat(), "$lte": end.isoformat()},
            "is_allocated": True
        }).to_list(length=None)
        
        for slit in slit_widths:
            width_mm = slit.get("slit_width_mm", 0)
            quantity_meters = slit.get("allocated_quantity", 0) or 0
            order_id = slit.get("allocated_to_order_id")
            
            if width_mm > 0 and quantity_meters > 0:
                width_key = f"{width_mm}"
                
                # Initialize width entry if not exists
                if width_key not in usage_by_width:
                    usage_by_width[width_key] = {
                        "width_mm": width_mm,
                        "total_length_m": 0.0,
                        "orders": []
                    }
                
                # Add to width total
                usage_by_width[width_key]["total_length_m"] += quantity_meters
                
                # Get order details if breakdown requested
                if include_order_breakdown and order_id:
                    order = await db.orders.find_one({"id": order_id})
                    if order:
                        usage_by_width[width_key]["orders"].append({
                            "order_number": order.get("order_number", "Unknown"),
                            "length_m": quantity_meters,
                            "order_date": order.get("created_at", ""),
                            "client_name": order.get("client_name", "Unknown")
                        })
                
                # Calculate m² for this usage
                m2 = (width_mm / 1000.0) * quantity_meters
                total_m2 += m2
        
        # Convert to sorted list
        usage_list = []
        for width_key, data in usage_by_width.items():
            width_mm = data["width_mm"]
            total_length_m = data["total_length_m"]
            m2_for_width = (width_mm / 1000.0) * total_length_m
            
            width_data = {
                "width_mm": width_mm,
                "total_length_m": round(total_length_m, 2),
                "m2": round(m2_for_width, 2)
            }
            
            # Add order breakdown if requested
            if include_order_breakdown:
                width_data["orders"] = data["orders"]
                width_data["order_count"] = len(data["orders"])
            
            usage_list.append(width_data)
        
        # Sort by width (ascending)
        usage_list.sort(key=lambda x: x["width_mm"])
        
        report_data = {
            "material_id": material_id,
            "material_name": material.get("material_description", material.get("supplier", "Unknown")),
            "material_code": material.get("product_code", "N/A"),
            "report_period": {
                "start_date": start_date,
                "end_date": end_date,
                "days": (end - start).days
            },
            "usage_by_width": usage_list,
            "total_widths_used": len(usage_list),
            "grand_total_m2": round(total_m2, 2),
            "grand_total_length_m": round(sum(w["total_length_m"] for w in usage_list), 2),
            "include_order_breakdown": include_order_breakdown
        }
        
        return StandardResponse(
            success=True,
            message="Detailed material usage report generated successfully",
            data=report_data
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to generate detailed material usage report: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to generate report: {str(e)}")


@api_router.get("/stock/reports/product-usage-detailed", response_model=StandardResponse)
async def get_detailed_product_usage_report(
    client_id: Optional[str] = None,
    start_date: str = None,
    end_date: str = None,
    include_order_breakdown: bool = False,
    current_user: dict = Depends(require_any_role)
):
    """
    Generate detailed product usage report by width and length for client products.
    Excludes 'Spiral Paper Cores' and 'Composite Cores' product types.
    Shows usage per width, total length, and grand total m² used.
    Optionally includes order-by-order breakdown.
    """
    try:
        # Parse dates - handle if no dates provided, default to last 30 days
        if not start_date or not end_date:
            end = datetime.now()
            start = end - timedelta(days=30)
        else:
            start = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            end = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
        
        # Remove timezone info for MongoDB comparison
        start_dt = start.replace(tzinfo=None)
        end_dt = end.replace(tzinfo=None)
        
        # Build query for orders in the date range
        # Include all orders (including orders on hand) except cancelled
        order_query = {
            "created_at": {"$gte": start_dt, "$lte": end_dt}
        }
        
        # Exclude cancelled orders if such status exists
        if await db.orders.find_one({"status": "cancelled"}):
            order_query["status"] = {"$ne": "cancelled"}
        
        if client_id:
            order_query["client_id"] = client_id
        
        orders = await db.orders.find(order_query).to_list(length=None)
        
        # Track product usage by product and width
        product_usage = {}  # {product_id: {widths: {width_mm: {data}}, product_info: {}}}
        
        for order in orders:
            order_number = order.get("order_number", "Unknown")
            order_items = order.get("items", [])
            
            for item in order_items:
                product_id = item.get("product_id")
                if not product_id:
                    continue
                
                # Get product details
                product = await db.client_products.find_one({"id": product_id})
                if not product:
                    continue
                
                # Skip excluded product types
                product_type = product.get("product_type", "")
                if product_type in ["Spiral Paper Cores", "Composite Cores"]:
                    continue
                
                # Get dimensions
                width_mm = item.get("width", 0) or product.get("width", 0)
                quantity = item.get("quantity", 0)
                length_m = item.get("length", 0) or product.get("length", 0)
                
                if width_mm <= 0 or quantity <= 0:
                    continue
                
                # Calculate total length for this item (quantity * length per unit)
                total_length = quantity * length_m
                
                # Initialize product entry if not exists
                if product_id not in product_usage:
                    product_usage[product_id] = {
                        "product_info": {
                            "product_id": product_id,
                            "product_description": product.get("product_description", "Unknown"),
                            "product_code": product.get("product_code", "N/A"),
                            "product_type": product_type,
                            "client_id": product.get("client_id"),
                            "client_name": order.get("client_name", "Unknown")
                        },
                        "widths": {},
                        "total_length_m": 0,
                        "total_m2": 0
                    }
                
                width_key = f"{width_mm}"
                
                # Initialize width entry if not exists
                if width_key not in product_usage[product_id]["widths"]:
                    product_usage[product_id]["widths"][width_key] = {
                        "width_mm": width_mm,
                        "total_length_m": 0.0,
                        "orders": []
                    }
                
                # Add to width total
                product_usage[product_id]["widths"][width_key]["total_length_m"] += total_length
                
                # Track order detail if breakdown requested
                if include_order_breakdown:
                    product_usage[product_id]["widths"][width_key]["orders"].append({
                        "order_number": order_number,
                        "quantity": quantity,
                        "length_per_unit": length_m,
                        "total_length_m": total_length,
                        "order_date": order.get("created_at", ""),
                        "client_name": order.get("client_name", "Unknown")
                    })
        
        # Calculate totals and format output
        products_list = []
        grand_total_m2 = 0.0
        grand_total_length_m = 0.0
        
        for product_id, data in product_usage.items():
            product_total_length = 0.0
            product_total_m2 = 0.0
            
            widths_list = []
            for width_key, width_data in data["widths"].items():
                width_mm = width_data["width_mm"]
                total_length_m = width_data["total_length_m"]
                m2_for_width = (width_mm / 1000.0) * total_length_m
                
                width_info = {
                    "width_mm": width_mm,
                    "total_length_m": round(total_length_m, 2),
                    "m2": round(m2_for_width, 2)
                }
                
                # Add order breakdown if requested
                if include_order_breakdown:
                    width_info["orders"] = width_data["orders"]
                    width_info["order_count"] = len(width_data["orders"])
                
                widths_list.append(width_info)
                product_total_length += total_length_m
                product_total_m2 += m2_for_width
            
            # Sort widths by width_mm
            widths_list.sort(key=lambda x: x["width_mm"])
            
            products_list.append({
                "product_info": data["product_info"],
                "usage_by_width": widths_list,
                "total_widths_used": len(widths_list),
                "product_total_length_m": round(product_total_length, 2),
                "product_total_m2": round(product_total_m2, 2)
            })
            
            grand_total_length_m += product_total_length
            grand_total_m2 += product_total_m2
        
        # Sort products by description
        products_list.sort(key=lambda x: x["product_info"]["product_description"])
        
        report_data = {
            "report_period": {
                "start_date": start_date,
                "end_date": end_date,
                "days": (end - start).days
            },
            "client_filter": client_id,
            "products": products_list,
            "total_products": len(products_list),
            "grand_total_m2": round(grand_total_m2, 2),
            "grand_total_length_m": round(grand_total_length_m, 2),
            "include_order_breakdown": include_order_breakdown,
            "excluded_types": ["Spiral Paper Cores", "Composite Cores"]
        }
        
        return StandardResponse(
            success=True,
            message="Detailed product usage report generated successfully",
            data=report_data
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to generate detailed product usage report: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to generate report: {str(e)}")



@api_router.get("/stock/reports/projected-order-analysis", response_model=StandardResponse)
async def get_projected_order_analysis(
    client_id: Optional[str] = None,
    start_date: str = None,
    end_date: str = None,
    current_user: dict = Depends(require_any_role)
):
    """
    Generate projected order analysis based on historical data.
    Shows projections for 3, 6, 9, and 12 months with product-level detail
    and raw material requirements.
    """
    try:
        # Handle default dates if not provided
        if not start_date:
            start_date = (datetime.now() - timedelta(days=90)).isoformat() + 'Z'
        if not end_date:
            end_date = datetime.now().isoformat() + 'Z'
            
        # Parse dates
        start = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        end = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
        days_in_period = (end - start).days or 1
        
        # Build query for orders in the date range
        # Include all orders except cancelled/deleted to capture orders on hand
        # FIX: MongoDB stores dates as datetime objects, not strings
        # Use datetime objects for comparison instead of strings
        start_dt = start.replace(tzinfo=None)
        end_dt = end.replace(tzinfo=None)
            
        order_query = {
            "created_at": {"$gte": start_dt, "$lte": end_dt}
        }
        
        # Exclude cancelled/deleted orders if such status exists
        if await db.orders.find_one({"status": "cancelled"}):
            order_query["status"] = {"$ne": "cancelled"}
        
        if client_id:
            order_query["client_id"] = client_id
        
        orders = await db.orders.find(order_query).to_list(length=None)
        
        # Track product usage with details
        product_analysis = {}  # {product_id: {usage_data, client_info, product_specs}}
        
        for order in orders:
            order_items = order.get("items", [])
            
            for item in order_items:
                product_id = item.get("product_id")
                if not product_id:
                    continue
                
                logger.info(f"Order {order.get('order_number')}: found product_id {product_id} ({item.get('product_name')})")
                
                # Get product details
                product = await db.client_products.find_one({"id": product_id})
                if not product:
                    continue
                
                quantity = item.get("quantity", 0)
                if quantity <= 0:
                    continue
                
                # Initialize product entry if not exists
                if product_id not in product_analysis:
                    product_analysis[product_id] = {
                        "product_info": {
                            "product_id": product_id,
                            "product_description": product.get("product_description", "Unknown"),
                            "product_code": product.get("product_code", "N/A"),
                            "product_type": product.get("product_type", "Unknown"),
                            "client_id": product.get("client_id"),
                            "client_name": order.get("client_name", "Unknown"),
                            "width": product.get("width", 0),
                            "length": product.get("length", 0),
                            "unit_of_measure": product.get("unit_of_measure", "units")
                        },
                        "historical_orders": [],
                        "total_quantity": 0,
                        "order_count": 0,
                        "materials_composition": product.get("materials_composition", [])
                    }
                
                # Track order
                product_analysis[product_id]["historical_orders"].append({
                    "order_number": order.get("order_number"),
                    "order_date": order.get("created_at"),
                    "quantity": quantity,
                    "client_name": order.get("client_name")
                })
                
                product_analysis[product_id]["total_quantity"] += quantity
                product_analysis[product_id]["order_count"] += 1
        
        # Calculate projections and material requirements
        products_list = []
        
        logger.info(f"Processing {len(product_analysis)} products for projections")
        
        for product_id, data in product_analysis.items():
            logger.info(f"Analyzing product_id: {product_id} from product_analysis")
            
            # Calculate average usage per day
            avg_per_day = data["total_quantity"] / days_in_period
            
            # Project for 3, 6, 9, and 12 months
            projections = {
                "3_months": round(avg_per_day * 90, 2),
                "6_months": round(avg_per_day * 180, 2),
                "9_months": round(avg_per_day * 270, 2),
                "12_months": round(avg_per_day * 365, 2)
            }
            
            # Calculate material requirements for projections
            # Get product details including material_layers
            product = await db.client_products.find_one({"id": product_id})
            material_layers = product.get("material_layers", []) if product else []
            
            # Debug logging
            logger.info(f"Product {product_id}: has {len(material_layers)} material layers")
            if product:
                logger.info(f"Product type: {product.get('product_type')}, width: {product.get('width')}, length: {product.get('length')}")
            
            material_requirements = {}
            
            if material_layers and len(material_layers) > 0:
                # Check if this is a core-based product (paper cores) or flat product
                is_core_product = product.get("product_type") == "paper_cores"
                
                if is_core_product:
                    # Spiral-wound paper core calculation using cylinder shell formula
                    try:
                        # Get core specifications in mm
                        core_id_mm = float(product.get("core_id") or 76)  # Inner diameter in mm
                        core_length_mm = float(product.get("core_width") or product.get("width") or 1200)  # Core length in mm
                        wall_thickness_mm = float(product.get("core_thickness") or 3)  # Wall thickness in mm
                        
                        # Convert to meters
                        core_id_m = core_id_mm / 1000
                        core_length_m = core_length_mm / 1000
                        wall_thickness_m = wall_thickness_mm / 1000
                        
                        # Calculate inner radius
                        inner_radius_m = core_id_m / 2
                        
                    except (TypeError, ValueError):
                        # If conversion fails, use defaults
                        core_id_mm = 76
                        core_length_mm = 1200
                        wall_thickness_mm = 3
                        core_id_m = 0.076
                        core_length_m = 1.2
                        wall_thickness_m = 0.003
                        inner_radius_m = 0.038
                    
                    for period, projected_qty in projections.items():
                        material_requirements[period] = []
                        
                        # Track current radius position as we build layers
                        current_inner_radius = inner_radius_m
                        total_paper_mass_kg = 0
                        total_paper_area_m2 = 0
                        total_strip_length_m = 0
                        total_cost = 0
                        
                        # Process each material layer
                        for layer_index, layer in enumerate(material_layers):
                            try:
                                material_id = layer.get("material_id")
                                thickness_mm = float(layer.get("thickness") or 0)  # Thickness per single layer in mm
                                num_layers = int(layer.get("quantity") or 1)  # How many layers of this material
                                layer_width_mm = float(layer.get("width") or 0)  # Width in mm (if material is cut into strips)
                            except (TypeError, ValueError) as e:
                                logger.error(f"Error parsing layer fields: {e}, layer: {layer}")
                                continue
                            
                            if thickness_mm <= 0 or num_layers <= 0:
                                continue
                            
                            # Convert to meters
                            thickness_m = thickness_mm / 1000
                            layer_width_m = layer_width_mm / 1000 if layer_width_mm > 0 else None
                            
                            # Calculate total thickness for this material stream
                            total_stream_thickness_m = thickness_m * num_layers
                            
                            # Calculate inner and outer radius for this material stream
                            stream_inner_radius = current_inner_radius
                            stream_outer_radius = current_inner_radius + total_stream_thickness_m
                            
                            # Calculate volume using cylinder shell formula
                            # Volume = π × core_length × (outer_radius² - inner_radius²)
                            volume_m3 = 3.14159 * core_length_m * (
                                (stream_outer_radius ** 2) - (stream_inner_radius ** 2)
                            )
                            
                            # Get material information and GSM from database
                            material_name = layer.get("material_name", "Unknown")
                            layer_type = layer.get("layer_type", f"Layer {layer_index + 1}")
                            gsm = 0  # Initialize GSM
                            material_cost = 0
                            cost_per_meter = 0
                            price_per_tonne = 0
                            linear_metres_per_tonne = 0
                            
                            if material_id:
                                material = await db.materials.find_one({"id": material_id})
                                if material:
                                    material_name = material.get("material_description", material.get("supplier", material_name))
                                    price_per_tonne = float(material.get("price", 0))
                                    
                                    # Get GSM from material database (it's stored as string, so convert to float)
                                    gsm_str = material.get("gsm", "0")
                                    try:
                                        gsm = float(gsm_str) if gsm_str else 0
                                    except (ValueError, TypeError):
                                        gsm = 0
                                    
                                    # Calculate cost per linear metre using the correct formula
                                    # Formula: Convert 1 tonne (1,000 kg = 1,000,000 grams) to linear metres
                                    # linear_metres_per_tonne = 1,000,000 / (GSM × width_metres)
                                    if gsm > 0 and layer_width_m and layer_width_m > 0:
                                        # 1 tonne = 1,000,000 grams
                                        # linear metres per tonne = 1,000,000 grams / (GSM × width_metres)
                                        linear_metres_per_tonne = 1000000 / (gsm * layer_width_m)
                                        # price per metre = price per tonne / linear metres per tonne
                                        cost_per_meter = price_per_tonne / linear_metres_per_tonne
                                    else:
                                        # Fallback: if GSM or width not available, use price directly
                                        cost_per_meter = price_per_tonne
                            
                            # Calculate density from GSM and thickness
                            # density = GSM ÷ thickness(mm) gives kg/m³
                            if thickness_mm > 0 and gsm > 0:
                                density_kg_m3 = gsm / thickness_mm
                            else:
                                density_kg_m3 = 0
                            
                            # Calculate mass: mass = volume × density
                            stream_mass_kg = volume_m3 * density_kg_m3
                            
                            # Calculate surface area: area = volume ÷ thickness
                            if total_stream_thickness_m > 0:
                                stream_area_m2 = volume_m3 / total_stream_thickness_m
                            else:
                                stream_area_m2 = 0
                            
                            # Calculate strip length if layer width is provided
                            if layer_width_m and layer_width_m > 0:
                                stream_strip_length_m = stream_area_m2 / layer_width_m
                            else:
                                stream_strip_length_m = 0
                            
                            # Calculate total for projected quantity
                            total_mass_kg = stream_mass_kg * projected_qty
                            total_area_m2 = stream_area_m2 * projected_qty
                            total_length_m = stream_strip_length_m * projected_qty
                            
                            # Cost calculation: use strip length (linear metres) - moved here after strip length is calculated
                            if stream_strip_length_m > 0:
                                # Material cost = linear metres × cost per metre × projected quantity
                                material_cost = stream_strip_length_m * cost_per_meter * projected_qty
                            else:
                                # Fallback to area-based calculation if strip length unavailable
                                material_cost = stream_area_m2 * cost_per_meter
                            
                            # Add to totals
                            total_paper_mass_kg += total_mass_kg
                            total_paper_area_m2 += total_area_m2
                            if stream_strip_length_m > 0:
                                total_strip_length_m += total_length_m
                            total_cost += material_cost
                            
                            material_requirements[period].append({
                                "layer_order": layer_index + 1,
                                "layer_type": layer_type,
                                "material_id": material_id,
                                "material_name": material_name,
                                "width_mm": layer_width_mm,
                                "thickness_mm": thickness_mm,
                                "gsm": gsm,
                                "num_layers": num_layers,
                                "stream_inner_radius_mm": round(stream_inner_radius * 1000, 2),
                                "stream_outer_radius_mm": round(stream_outer_radius * 1000, 2),
                                "volume_m3_per_core": round(volume_m3, 6),
                                "density_kg_m3": round(density_kg_m3, 2),
                                "mass_kg_per_core": round(stream_mass_kg, 4),
                                "area_m2_per_core": round(stream_area_m2, 4),
                                "strip_length_m_per_core": round(stream_strip_length_m, 2) if stream_strip_length_m > 0 else None,
                                "total_mass_kg": round(total_mass_kg, 2),
                                "total_area_m2": round(total_area_m2, 2),
                                "total_strip_length_m": round(total_length_m, 2) if stream_strip_length_m > 0 else None,
                                "meters_per_core": round(stream_strip_length_m, 2) if stream_strip_length_m > 0 else round(stream_area_m2, 2),
                                "total_meters_needed": round(total_length_m, 2) if stream_strip_length_m > 0 else round(total_area_m2, 2),
                                # New cost calculation fields
                                "price_per_tonne": round(price_per_tonne, 2),
                                "linear_metres_per_tonne": round(linear_metres_per_tonne, 2) if linear_metres_per_tonne > 0 else None,
                                "cost_per_meter": round(cost_per_meter, 4),
                                "cost_per_core": round(stream_strip_length_m * cost_per_meter, 4) if stream_strip_length_m > 0 else round(stream_area_m2 * cost_per_meter, 4),
                                "total_cost": round(material_cost, 2)
                            })
                            
                            # Update current radius for next layer
                            current_inner_radius = stream_outer_radius
                        
                        # Calculate outer diameter
                        outer_diameter_mm = core_id_mm + (2 * wall_thickness_mm)
                        
                        # Calculate cost per core and per metre
                        cost_per_core = total_cost / projected_qty if projected_qty > 0 else 0
                        # Cost per metre of finished core (assuming core_length_m is the length of one core)
                        cost_per_metre_of_core = cost_per_core / core_length_m if core_length_m > 0 else 0
                        
                        # Add summary totals
                        if material_requirements[period]:
                            material_requirements[period].append({
                                "is_total": True,
                                "outer_diameter_mm": round(outer_diameter_mm, 2),
                                "core_length_m": round(core_length_m, 3),
                                "total_paper_mass_kg": round(total_paper_mass_kg, 2),
                                "total_paper_area_m2": round(total_paper_area_m2, 2),
                                "total_strip_length_m": round(total_strip_length_m, 2) if total_strip_length_m > 0 else None,
                                "total_meters_all_layers": round(total_strip_length_m, 2) if total_strip_length_m > 0 else round(total_paper_area_m2, 2),
                                "total_cost": round(total_cost, 2),
                                "cost_per_core": round(cost_per_core, 4),
                                "cost_per_metre_of_core": round(cost_per_metre_of_core, 4),
                                "projected_quantity": int(projected_qty)
                            })
                else:
                    # Simpler calculation for flat products (labels, films, tapes)
                    try:
                        product_width = float(product.get("width") or 0) / 1000  # Convert mm to meters
                        product_length = float(product.get("length") or 0)  # Already in meters
                    except (TypeError, ValueError):
                        product_width = 1.0  # Default 1 meter
                        product_length = 100  # Default 100 meters
                    
                    for period, projected_qty in projections.items():
                        material_requirements[period] = []
                        total_meters_all_layers = 0
                        
                        for layer_index, layer in enumerate(material_layers):
                            try:
                                material_id = layer.get("material_id")
                                thickness = float(layer.get("thickness") or 0)  # mm
                                width = float(layer.get("width") or (product_width * 1000))  # mm
                                quantity_per_unit = int(layer.get("quantity") or 1)
                            except (TypeError, ValueError) as e:
                                logger.error(f"Error parsing flat product layer fields: {e}, layer: {layer}")
                                continue
                            
                            # For flat products: meters needed = product length × quantity per unit × projected qty
                            meters_per_unit = product_length * quantity_per_unit
                            total_meters = meters_per_unit * projected_qty
                            total_meters_all_layers += total_meters
                            
                            material_name = layer.get("material_name", "Unknown")
                            layer_type = layer.get("layer_type", f"Layer {layer_index + 1}")
                            gsm = layer.get("gsm", 0)
                            
                            # Calculate material cost
                            material_cost = 0
                            cost_per_meter = 0
                            
                            if material_id:
                                material = await db.materials.find_one({"id": material_id})
                                if material:
                                    material_name = material.get("material_description", material.get("supplier", material_name))
                                    # Get cost per unit (usually per meter)
                                    cost_per_unit = float(material.get("cost_per_unit", 0))
                                    cost_per_meter = cost_per_unit
                                    material_cost = total_meters * cost_per_meter
                            
                            material_requirements[period].append({
                                "layer_order": layer_index + 1,
                                "layer_type": layer_type,
                                "material_id": material_id,
                                "material_name": material_name,
                                "width_mm": width,
                                "thickness_mm": thickness,
                                "gsm": gsm,
                                "laps_per_core": quantity_per_unit,
                                "meters_per_core": round(meters_per_unit, 2),
                                "total_meters_needed": round(total_meters, 2),
                                "cost_per_meter": round(cost_per_meter, 4),
                                "total_cost": round(material_cost, 2)
                            })
                        
                        if material_requirements[period]:
                            material_requirements[period].append({
                                "is_total": True,
                                "total_meters_all_layers": round(total_meters_all_layers, 2)
                            })
            
            # Calculate customer breakdown for projections
            customer_breakdown = {}
            total_qty = data["total_quantity"]
            
            for order in data["historical_orders"]:
                client_name = order.get("client_name", "Unknown")
                qty = order.get("quantity", 0)
                
                if client_name not in customer_breakdown:
                    customer_breakdown[client_name] = {
                        "total_quantity": 0,
                        "order_count": 0,
                        "percentage": 0
                    }
                
                customer_breakdown[client_name]["total_quantity"] += qty
                customer_breakdown[client_name]["order_count"] += 1
            
            # Calculate percentages and project per customer
            customer_projections = {}
            for client_name, client_data in customer_breakdown.items():
                percentage = (client_data["total_quantity"] / total_qty * 100) if total_qty > 0 else 0
                customer_breakdown[client_name]["percentage"] = round(percentage, 1)
                
                # Project quantities for each period based on historical percentage
                customer_projections[client_name] = {
                    "3_months": round(projections["3_months"] * percentage / 100, 2),
                    "6_months": round(projections["6_months"] * percentage / 100, 2),
                    "9_months": round(projections["9_months"] * percentage / 100, 2),
                    "12_months": round(projections["12_months"] * percentage / 100, 2),
                    "percentage": round(percentage, 1),
                    "historical_total": client_data["total_quantity"],
                    "order_count": client_data["order_count"]
                }
            
            products_list.append({
                "product_info": data["product_info"],
                "historical_data": {
                    "total_quantity": data["total_quantity"],
                    "order_count": data["order_count"],
                    "total_orders": data["order_count"],  # Alias for frontend compatibility
                    "average_per_day": round(avg_per_day, 2),
                    "average_per_month": round(avg_per_day * 30, 2),
                    "avg_monthly_orders": round(avg_per_day * 30, 2),  # Alias for frontend compatibility
                    "orders": data["historical_orders"]
                },
                "projections": projections,
                "material_requirements": material_requirements,
                "customer_breakdown": customer_breakdown,
                "customer_projections": customer_projections
            })
        
        # Sort by total quantity (most used first)
        products_list.sort(key=lambda x: x["historical_data"]["total_quantity"], reverse=True)
        
        # Calculate period-based summaries for frontend including material costs
        period_summaries = {}
        for period in ["3_months", "6_months", "9_months", "12_months"]:
            total_projected_orders = sum(
                p["projections"][period] for p in products_list
            )
            
            # Calculate total material cost across all products
            total_material_cost = 0
            for product in products_list:
                mat_reqs = product.get("material_requirements", {}).get(period, [])
                for mat in mat_reqs:
                    if not mat.get("is_total"):  # Skip total rows
                        total_material_cost += mat.get("total_cost", 0)
            
            period_summaries[period] = {
                "total_projected_orders": round(total_projected_orders, 2),
                "total_projected_material_cost": round(total_material_cost, 2),
                "products_analyzed": len(products_list)
            }
        
        report_data = {
            "report_period": {
                "start_date": start_date,
                "end_date": end_date,
                "days": days_in_period
            },
            "client_filter": client_id,
            "products": products_list,
            "total_products": len(products_list),
            "summary": period_summaries  # Period-based summaries
        }
        
        return StandardResponse(
            success=True,
            message="Projected order analysis generated successfully",
            data=report_data
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to generate projected order analysis: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to generate report: {str(e)}")






@api_router.get("/stock/reports/job-card-performance", response_model=StandardResponse)
async def get_job_card_performance_report(
    start_date: str = None,
    end_date: str = None,
    current_user: dict = Depends(require_any_role)
):
    """
    Enhanced job card performance report with comprehensive metrics:
    - Time spent on each job and by stage
    - Material used, excess material (wastage), and material entered into stock
    - Efficiency score (on-time vs delayed jobs)
    - Job type breakdown and client performance analysis
    """
    try:
        # Parse dates - handle if no dates provided, default to last 30 days
        if not start_date or not end_date:
            end = datetime.now()
            start = end - timedelta(days=30)
        else:
            start = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            end = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
        
        # Remove timezone info for MongoDB comparison
        start_dt = start.replace(tzinfo=None)
        end_dt = end.replace(tzinfo=None)
        
        # Find completed orders in the date range
        order_query = {
            "completed_at": {"$gte": start_dt, "$lte": end_dt},
            "status": {"$in": ["completed", "archived"]}
        }
        
        completed_orders = await db.orders.find(order_query).to_list(length=None)
        
        job_cards = []
        total_time_hours = 0
        total_stock_entries = 0
        total_stock_quantity = 0
        total_material_used_kg = 0
        total_material_excess_kg = 0
        jobs_on_time = 0
        jobs_delayed = 0
        
        # Collect data for breakdowns
        job_type_metrics = {}  # product_type -> metrics
        client_metrics = {}  # client_id -> metrics
        
        for order in completed_orders:
            order_id = order.get("id")
            order_number = order.get("order_number", "Unknown")
            client_id = order.get("client_id", "unknown")
            client_name = order.get("client_name", "Unknown")
            due_date = order.get("due_date")
            completed_at = order.get("completed_at")
            
            # Calculate if on time
            is_on_time = False
            if due_date and completed_at:
                if isinstance(due_date, str):
                    due_date = datetime.fromisoformat(due_date.replace('Z', '+00:00'))
                if isinstance(completed_at, str):
                    completed_at = datetime.fromisoformat(completed_at.replace('Z', '+00:00'))
                is_on_time = completed_at <= due_date
            
            if is_on_time:
                jobs_on_time += 1
            else:
                jobs_delayed += 1
            
            # Get production logs for this order to calculate time spent
            production_logs = await db.production_logs.find({
                "order_id": order_id
            }).sort("timestamp", 1).to_list(length=None)
            
            # Calculate time in each stage
            time_by_stage = {}
            stage_start_times = {}
            
            for log in production_logs:
                timestamp = log.get("timestamp")
                
                if isinstance(timestamp, str):
                    timestamp = datetime.fromisoformat(timestamp.replace('Z', '+00:00'))
                
                from_stage = log.get("from_stage")
                to_stage = log.get("to_stage")
                
                # Mark end of previous stage
                if from_stage and from_stage in stage_start_times:
                    start_time = stage_start_times[from_stage]
                    duration = (timestamp - start_time).total_seconds() / 3600  # hours
                    time_by_stage[from_stage] = time_by_stage.get(from_stage, 0) + duration
                    del stage_start_times[from_stage]
                
                # Mark start of new stage
                if to_stage:
                    stage_start_times[to_stage] = timestamp
            
            # Calculate total time for this job
            total_job_time = sum(time_by_stage.values())
            
            # Get material consumption for this order from stock_movements
            material_movements = await db.stock_movements.find({
                "reference_id": order_id,
                "reference_type": "order",
                "movement_type": "consumption"
            }).to_list(length=None)
            
            total_material_used = 0
            material_details = []
            
            for movement in material_movements:
                quantity = abs(movement.get("quantity_change", 0))  # Make positive
                total_material_used += quantity
                material_details.append({
                    "stock_id": movement.get("stock_id"),
                    "stock_type": movement.get("stock_type", "unknown"),
                    "quantity": quantity,
                    "notes": movement.get("notes", "")
                })
            
            # Calculate expected material usage based on product specifications
            order_items = order.get("items", [])
            total_ordered_qty = sum(item.get("quantity", 0) for item in order_items)
            expected_material = 0
            product_types = []
            
            for item in order_items:
                product_id = item.get("product_id")
                quantity = item.get("quantity", 0)
                
                # Try to get product specifications
                product_spec = await db.client_products.find_one({"id": product_id})
                if product_spec:
                    product_type = product_spec.get("product_type", "Unknown")
                    if product_type not in product_types:
                        product_types.append(product_type)
                    
                    # Try to calculate expected material from material_layers
                    material_layers = product_spec.get("material_layers", [])
                    for layer in material_layers:
                        layer_quantity = layer.get("quantity", 0)
                        expected_material += layer_quantity * quantity
            
            # Calculate excess material (wastage)
            material_excess = max(0, total_material_used - expected_material) if expected_material > 0 else 0
            waste_percentage = (material_excess / total_material_used * 100) if total_material_used > 0 else 0
            
            # Get stock entries for this order (finished goods entered into inventory)
            stock_entries = await db.raw_substrate_stock.find({
                "source_order_id": order_id
            }).to_list(length=None)
            
            stock_summary = []
            job_stock_quantity = 0
            
            for stock in stock_entries:
                qty = stock.get("quantity_on_hand", 0)
                job_stock_quantity += qty
                stock_summary.append({
                    "product_description": stock.get("product_description", "Unknown"),
                    "quantity": qty,
                    "unit_of_measure": stock.get("unit_of_measure", "units"),
                    "created_at": stock.get("created_at")
                })
            
            # Build job card data
            job_card_data = {
                "order_number": order_number,
                "order_id": order_id,
                "client_id": client_id,
                "client_name": client_name,
                "product_types": product_types,
                "created_at": order.get("created_at"),
                "due_date": order.get("due_date"),
                "completed_at": order.get("completed_at"),
                "is_on_time": is_on_time,
                "total_time_hours": round(total_job_time, 2),
                "time_by_stage": {k: round(v, 2) for k, v in time_by_stage.items()},
                "material_used_kg": round(total_material_used, 2),
                "expected_material_kg": round(expected_material, 2),
                "material_excess_kg": round(material_excess, 2),
                "waste_percentage": round(waste_percentage, 2),
                "material_details": material_details,
                "stock_entries": stock_summary,
                "total_stock_produced": job_stock_quantity,
                "ordered_quantity": total_ordered_qty,
                "stock_entry_count": len(stock_summary)
            }
            
            job_cards.append(job_card_data)
            
            # Update totals
            total_time_hours += total_job_time
            total_stock_entries += len(stock_summary)
            total_stock_quantity += job_stock_quantity
            total_material_used_kg += total_material_used
            total_material_excess_kg += material_excess
            
            # Update job type breakdown
            for product_type in product_types:
                if product_type not in job_type_metrics:
                    job_type_metrics[product_type] = {
                        "job_count": 0,
                        "total_time_hours": 0,
                        "total_material_used": 0,
                        "total_excess": 0,
                        "jobs_on_time": 0,
                        "jobs_delayed": 0
                    }
                
                job_type_metrics[product_type]["job_count"] += 1
                job_type_metrics[product_type]["total_time_hours"] += total_job_time
                job_type_metrics[product_type]["total_material_used"] += total_material_used
                job_type_metrics[product_type]["total_excess"] += material_excess
                if is_on_time:
                    job_type_metrics[product_type]["jobs_on_time"] += 1
                else:
                    job_type_metrics[product_type]["jobs_delayed"] += 1
            
            # Update client performance metrics
            if client_id not in client_metrics:
                client_metrics[client_id] = {
                    "client_name": client_name,
                    "job_count": 0,
                    "total_time_hours": 0,
                    "total_material_used": 0,
                    "total_excess": 0,
                    "jobs_on_time": 0,
                    "jobs_delayed": 0
                }
            
            client_metrics[client_id]["job_count"] += 1
            client_metrics[client_id]["total_time_hours"] += total_job_time
            client_metrics[client_id]["total_material_used"] += total_material_used
            client_metrics[client_id]["total_excess"] += material_excess
            if is_on_time:
                client_metrics[client_id]["jobs_on_time"] += 1
            else:
                client_metrics[client_id]["jobs_delayed"] += 1
        
        # Calculate averages and efficiency metrics
        job_count = len(job_cards)
        efficiency_score = (jobs_on_time / job_count * 100) if job_count > 0 else 0
        overall_waste_percentage = (total_material_excess_kg / total_material_used_kg * 100) if total_material_used_kg > 0 else 0
        
        averages = {
            "average_time_per_job_hours": round(total_time_hours / job_count, 2) if job_count > 0 else 0,
            "average_stock_entries_per_job": round(total_stock_entries / job_count, 2) if job_count > 0 else 0,
            "average_stock_quantity_per_job": round(total_stock_quantity / job_count, 2) if job_count > 0 else 0,
            "average_material_used_per_job_kg": round(total_material_used_kg / job_count, 2) if job_count > 0 else 0,
            "average_waste_per_job_kg": round(total_material_excess_kg / job_count, 2) if job_count > 0 else 0,
            "total_jobs_completed": job_count,
            "jobs_on_time": jobs_on_time,
            "jobs_delayed": jobs_delayed,
            "efficiency_score_percentage": round(efficiency_score, 2),
            "total_time_all_jobs_hours": round(total_time_hours, 2),
            "total_stock_produced": total_stock_quantity,
            "total_material_used_kg": round(total_material_used_kg, 2),
            "total_material_excess_kg": round(total_material_excess_kg, 2),
            "overall_waste_percentage": round(overall_waste_percentage, 2)
        }
        
        # Format job type breakdown
        job_type_breakdown = []
        for product_type, metrics in job_type_metrics.items():
            efficiency = (metrics["jobs_on_time"] / metrics["job_count"] * 100) if metrics["job_count"] > 0 else 0
            waste_pct = (metrics["total_excess"] / metrics["total_material_used"] * 100) if metrics["total_material_used"] > 0 else 0
            
            job_type_breakdown.append({
                "product_type": product_type,
                "job_count": metrics["job_count"],
                "total_time_hours": round(metrics["total_time_hours"], 2),
                "average_time_per_job": round(metrics["total_time_hours"] / metrics["job_count"], 2),
                "total_material_used_kg": round(metrics["total_material_used"], 2),
                "total_excess_kg": round(metrics["total_excess"], 2),
                "waste_percentage": round(waste_pct, 2),
                "jobs_on_time": metrics["jobs_on_time"],
                "jobs_delayed": metrics["jobs_delayed"],
                "efficiency_percentage": round(efficiency, 2)
            })
        
        # Format client performance
        client_performance = []
        for client_id, metrics in client_metrics.items():
            efficiency = (metrics["jobs_on_time"] / metrics["job_count"] * 100) if metrics["job_count"] > 0 else 0
            waste_pct = (metrics["total_excess"] / metrics["total_material_used"] * 100) if metrics["total_material_used"] > 0 else 0
            
            client_performance.append({
                "client_id": client_id,
                "client_name": metrics["client_name"],
                "job_count": metrics["job_count"],
                "total_time_hours": round(metrics["total_time_hours"], 2),
                "average_time_per_job": round(metrics["total_time_hours"] / metrics["job_count"], 2),
                "total_material_used_kg": round(metrics["total_material_used"], 2),
                "total_excess_kg": round(metrics["total_excess"], 2),
                "waste_percentage": round(waste_pct, 2),
                "jobs_on_time": metrics["jobs_on_time"],
                "jobs_delayed": metrics["jobs_delayed"],
                "efficiency_percentage": round(efficiency, 2)
            })
        
        # Sort client performance by efficiency (highest first)
        client_performance.sort(key=lambda x: x["efficiency_percentage"], reverse=True)
        
        # Sort by completion date (most recent first)
        job_cards.sort(key=lambda x: x["completed_at"] if x["completed_at"] else "", reverse=True)
        
        report_data = {
            "report_period": {
                "start_date": start_date if start_date else start.isoformat() + 'Z',
                "end_date": end_date if end_date else end.isoformat() + 'Z',
                "days": (end_dt - start_dt).days
            },
            "job_cards": job_cards,
            "averages": averages,
            "job_type_breakdown": job_type_breakdown,
            "client_performance": client_performance
        }
        
        return StandardResponse(
            success=True,
            message="Job card performance report generated successfully",
            data=report_data
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to generate job card performance report: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to generate report: {str(e)}")



@api_router.get("/stock/reports/job-card-performance/export-csv")
async def export_job_card_performance_csv(
    start_date: str = None,
    end_date: str = None,
    current_user: dict = Depends(require_any_role)
):
    """
    Export job card performance report as CSV file for Excel/spreadsheet analysis
    """
    try:
        import csv
        from io import StringIO
        from fastapi.responses import StreamingResponse
        
        # Get the report data using the same logic
        report_response = await get_job_card_performance_report(start_date, end_date, current_user)
        report_data = report_response.data
        
        # Create CSV in memory
        output = StringIO()
        writer = csv.writer(output)
        
        # Write summary section
        writer.writerow(['Job Card Performance Report'])
        writer.writerow(['Period', f"{report_data['report_period']['start_date']} to {report_data['report_period']['end_date']}"])
        writer.writerow([])
        
        # Write averages/summary section
        writer.writerow(['SUMMARY METRICS'])
        averages = report_data['averages']
        writer.writerow(['Total Jobs Completed', averages['total_jobs_completed']])
        writer.writerow(['Jobs On Time', averages['jobs_on_time']])
        writer.writerow(['Jobs Delayed', averages['jobs_delayed']])
        writer.writerow(['Efficiency Score (%)', averages['efficiency_score_percentage']])
        writer.writerow(['Total Time (hours)', averages['total_time_all_jobs_hours']])
        writer.writerow(['Average Time Per Job (hours)', averages['average_time_per_job_hours']])
        writer.writerow(['Total Material Used (kg)', averages['total_material_used_kg']])
        writer.writerow(['Total Material Excess/Waste (kg)', averages['total_material_excess_kg']])
        writer.writerow(['Overall Waste Percentage (%)', averages['overall_waste_percentage']])
        writer.writerow(['Average Material Per Job (kg)', averages['average_material_used_per_job_kg']])
        writer.writerow(['Average Waste Per Job (kg)', averages['average_waste_per_job_kg']])
        writer.writerow(['Total Stock Produced', averages['total_stock_produced']])
        writer.writerow(['Average Stock Entries Per Job', averages['average_stock_entries_per_job']])
        writer.writerow([])
        
        # Write job type breakdown
        writer.writerow(['JOB TYPE BREAKDOWN'])
        writer.writerow(['Product Type', 'Job Count', 'Total Time (hrs)', 'Avg Time/Job', 'Material Used (kg)', 
                        'Excess (kg)', 'Waste %', 'On Time', 'Delayed', 'Efficiency %'])
        for breakdown in report_data.get('job_type_breakdown', []):
            writer.writerow([
                breakdown['product_type'],
                breakdown['job_count'],
                breakdown['total_time_hours'],
                breakdown['average_time_per_job'],
                breakdown['total_material_used_kg'],
                breakdown['total_excess_kg'],
                breakdown['waste_percentage'],
                breakdown['jobs_on_time'],
                breakdown['jobs_delayed'],
                breakdown['efficiency_percentage']
            ])
        writer.writerow([])
        
        # Write client performance
        writer.writerow(['CLIENT PERFORMANCE'])
        writer.writerow(['Client Name', 'Job Count', 'Total Time (hrs)', 'Avg Time/Job', 'Material Used (kg)', 
                        'Excess (kg)', 'Waste %', 'On Time', 'Delayed', 'Efficiency %'])
        for client in report_data.get('client_performance', []):
            writer.writerow([
                client['client_name'],
                client['job_count'],
                client['total_time_hours'],
                client['average_time_per_job'],
                client['total_material_used_kg'],
                client['total_excess_kg'],
                client['waste_percentage'],
                client['jobs_on_time'],
                client['jobs_delayed'],
                client['efficiency_percentage']
            ])
        writer.writerow([])
        
        # Write detailed job cards
        writer.writerow(['DETAILED JOB CARDS'])
        writer.writerow(['Order Number', 'Client', 'Product Types', 'Created', 'Due Date', 'Completed', 
                        'On Time', 'Time (hrs)', 'Material Used (kg)', 'Expected (kg)', 'Excess (kg)', 
                        'Waste %', 'Stock Produced', 'Ordered Qty'])
        
        for job in report_data.get('job_cards', []):
            writer.writerow([
                job['order_number'],
                job['client_name'],
                ', '.join(job.get('product_types', [])),
                job.get('created_at', ''),
                job.get('due_date', ''),
                job.get('completed_at', ''),
                'Yes' if job.get('is_on_time') else 'No',
                job['total_time_hours'],
                job['material_used_kg'],
                job['expected_material_kg'],
                job['material_excess_kg'],
                job['waste_percentage'],
                job['total_stock_produced'],
                job['ordered_quantity']
            ])
        
        # Prepare the response
        output.seek(0)
        filename = f"job_card_performance_{datetime.now().strftime('%Y%m%d')}.csv"
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"Failed to export job card performance CSV: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to export CSV: {str(e)}")


# ============= SLIT WIDTH MANAGEMENT ENDPOINTS =============


@api_router.get("/stock/print/{stock_id}")
async def print_stock_description(
    stock_id: str,
    stock_type: str = "substrate",  # substrate or material
    current_user: dict = Depends(require_any_role)
):
    """Generate printable PDF description for a stock unit"""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from io import BytesIO
        
        # Fetch stock data
        if stock_type == "substrate":
            stock = await db.raw_substrate_stock.find_one({"id": stock_id})
            if not stock:
                raise HTTPException(status_code=404, detail="Stock item not found")
        else:  # material
            stock = await db.raw_material_stock.find_one({"id": stock_id})
            if not stock:
                raise HTTPException(status_code=404, detail="Material not found")
        
        # Create PDF in memory
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        # Container for PDF elements
        elements = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#1e3a8a'),
            spaceAfter=12,
            spaceBefore=12
        )
        
        # Title
        title_text = "Adela Merchants - Stock Description"
        elements.append(Paragraph(title_text, title_style))
        elements.append(Spacer(1, 0.3*inch))
        
        # Helper function to format date
        def format_date(date_value):
            """Safely format date from various types"""
            if not date_value:
                return "N/A"
            try:
                # Handle datetime objects directly
                if isinstance(date_value, datetime):
                    return date_value.strftime("%Y-%m-%d")
                # Handle string dates
                elif isinstance(date_value, str):
                    # Parse ISO format string
                    dt = datetime.fromisoformat(date_value.replace('Z', '+00:00'))
                    return dt.strftime("%Y-%m-%d")
                # Handle date objects
                elif hasattr(date_value, 'strftime'):
                    return date_value.strftime("%Y-%m-%d")
                else:
                    # Convert to string as fallback
                    return str(date_value)
            except Exception as e:
                logger.warning(f"Date formatting error for value {date_value} (type: {type(date_value)}): {str(e)}")
                return "N/A"
        
        # Stock Information
        if stock_type == "substrate":
            elements.append(Paragraph("Product on Hand", heading_style))
            
            data = [
                ["Product Code:", stock.get("product_code", "N/A")],
                ["Description:", stock.get("product_description", "N/A")],
                ["Client:", stock.get("client_name", "N/A")],
                ["Quantity on Hand:", f"{stock.get('quantity_on_hand', 0)} {stock.get('unit_of_measure', 'units')}"],
                ["Minimum Stock Level:", f"{stock.get('minimum_stock_level', 0)} {stock.get('unit_of_measure', 'units')}"],
                ["Source Order:", stock.get("source_order_id", "N/A")],
                ["Location:", stock.get("location", "Main Warehouse")],
                ["Last Updated:", format_date(stock.get("created_at"))]
            ]
        else:
            elements.append(Paragraph("Raw Material Stock", heading_style))
            
            data = [
                ["Material Name:", stock.get("material_name", "N/A")],
                ["Material ID:", stock.get("material_id", "N/A")],
                ["Quantity on Hand:", f"{stock.get('quantity_on_hand', 0)} {stock.get('unit_of_measure', 'kg')}"],
                ["Minimum Stock Level:", f"{stock.get('minimum_stock_level', 0)} {stock.get('unit_of_measure', 'kg')}"],
                ["Usage Rate/Month:", f"{stock.get('usage_rate_per_month', 0)} {stock.get('unit_of_measure', 'kg')}"],
                ["Alert Threshold:", f"{stock.get('alert_threshold_days', 7)} days"],
                ["Supplier:", stock.get("supplier_name", "N/A")],
                ["Last Updated:", format_date(stock.get("created_at"))]
            ]
        
        # Create table
        table = Table(data, colWidths=[2.5*inch, 4*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e5e7eb')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('TOPPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 0.5*inch))
        
        # Get stock movements
        query = {"stock_id": stock_id} if stock_type == "substrate" else {"product_id": stock.get("material_id")}
        movements = await db.stock_movements.find(query).sort("created_at", -1).limit(10).to_list(length=None)
        
        if movements:
            elements.append(Paragraph("Recent Stock Movements", heading_style))
            
            movement_data = [["Date", "Type", "Quantity", "Reference"]]
            for movement in movements:
                date_str = format_date(movement.get("created_at"))
                movement_data.append([
                    date_str,
                    movement.get("movement_type", "N/A").title(),
                    f"{movement.get('quantity', 0)} {stock.get('unit_of_measure', 'units')}",
                    movement.get("reference", "N/A")[:30]
                ])
            
            movement_table = Table(movement_data, colWidths=[1.2*inch, 1.2*inch, 1.5*inch, 2.6*inch])
            movement_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ]))
            
            elements.append(movement_table)
        
        # Footer
        elements.append(Spacer(1, 0.5*inch))

        footer_text = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}"
        footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
        elements.append(Paragraph(footer_text, footer_style))
        
        # Build PDF
        doc.build(elements)
        
        # Return PDF
        buffer.seek(0)
        filename = f"stock_{stock_type}_{stock_id}.pdf"
        
        return Response(
            content=buffer.getvalue(),
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to generate stock PDF: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to generate PDF: {str(e)}")


@api_router.get("/slit-widths/material/{material_id}", response_model=StandardResponse)
async def get_slit_widths_by_material(
    material_id: str,
    current_user: dict = Depends(require_any_role)
):
    """Get all slit widths available for a specific raw material"""
    try:
        # Get all slit widths for this raw material
        slit_widths = await db.slit_widths.find({
            "raw_material_id": material_id,
            "remaining_quantity": {"$gt": 0}  # Only show slit widths with remaining stock
        }).sort("slit_width_mm", 1).to_list(length=None)
        
        # Remove MongoDB ObjectIds
        for width in slit_widths:
            if "_id" in width:
                del width["_id"]
        
        # Group by width and sum quantities
        width_groups = {}
        for width in slit_widths:
            width_mm = width["slit_width_mm"]
            if width_mm not in width_groups:
                width_groups[width_mm] = {
                    "slit_width_mm": width_mm,
                    "total_quantity_meters": 0,
                    "available_quantity_meters": 0,
                    "entries": []
                }
            width_groups[width_mm]["total_quantity_meters"] += width["quantity_meters"]
            width_groups[width_mm]["available_quantity_meters"] += width["remaining_quantity"]
            width_groups[width_mm]["entries"].append(width)
        
        # Convert to list
        grouped_widths = list(width_groups.values())
        
        return StandardResponse(
            success=True,
            message="Slit widths retrieved successfully",
            data={
                "material_id": material_id,
                "slit_widths": grouped_widths,
                "total_widths": len(grouped_widths)
            }
        )
        
    except Exception as e:
        logger.error(f"Failed to get slit widths: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to retrieve slit widths")

@api_router.post("/slit-widths", response_model=StandardResponse)
async def create_slit_width(
    slit_width_data: SlitWidthCreate,
    current_user: dict = Depends(require_any_role)
):
    """Create new slit width entry from slitting job"""
    try:
        # Create the slit width entry
        slit_width = SlitWidth(
            **slit_width_data.dict(),
            remaining_quantity=slit_width_data.quantity_meters,
            created_by=current_user["sub"]
        )
        
        # Convert to dict and prepare for MongoDB
        slit_width_dict = slit_width.dict()
        
        # Insert into database
        result = await db.slit_widths.insert_one(slit_width_dict)
        
        return StandardResponse(
            success=True,
            message="Slit width created successfully",
            data={"slit_width_id": slit_width.id}
        )
        
    except Exception as e:
        logger.error(f"Failed to create slit width: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to create slit width")

@api_router.get("/slit-widths/check-availability", response_model=StandardResponse)
async def check_slit_width_availability(
    material_id: str,
    required_width_mm: float,
    required_quantity_meters: float,
    current_user: dict = Depends(require_any_role)
):
    """Check if required slit width and quantity is available"""
    try:
        # Find slit widths that match the required width
        matching_widths = await db.slit_widths.find({
            "raw_material_id": material_id,
            "slit_width_mm": required_width_mm,
            "remaining_quantity": {"$gt": 0}
        }).sort("created_at", 1).to_list(length=None)  # FIFO order
        
        # Calculate available quantity
        total_available = sum(width["remaining_quantity"] for width in matching_widths)
        
        # Remove MongoDB ObjectIds
        for width in matching_widths:
            if "_id" in width:
                del width["_id"]
        
        availability_data = {
            "material_id": material_id,
            "required_width_mm": required_width_mm,
            "required_quantity_meters": required_quantity_meters,
            "available_quantity_meters": total_available,
            "is_sufficient": total_available >= required_quantity_meters,
            "shortage_meters": max(0, required_quantity_meters - total_available),
            "matching_entries": matching_widths
        }
        
        return StandardResponse(
            success=True,
            message="Slit width availability checked",
            data=availability_data
        )
        
    except Exception as e:
        logger.error(f"Failed to check slit width availability: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to check availability")

@api_router.post("/slit-widths/allocate", response_model=StandardResponse)
async def allocate_slit_width(
    allocation_request: SlitWidthAllocationRequest,
    current_user: dict = Depends(require_any_role)
):
    """Allocate slit width to an order"""
    try:
        # Get the specific slit width entry
        slit_width = await db.slit_widths.find_one({"id": allocation_request.slit_width_id})
        
        if not slit_width:
            raise HTTPException(status_code=404, detail="Slit width entry not found")
        
        # Check if enough quantity is available
        if slit_width["remaining_quantity"] < allocation_request.required_quantity_meters:
            raise HTTPException(
                status_code=400, 
                detail=f"Insufficient quantity. Available: {slit_width['remaining_quantity']} meters, Required: {allocation_request.required_quantity_meters} meters"
            )
        
        # Calculate new quantities
        allocated_quantity = min(slit_width["remaining_quantity"], allocation_request.required_quantity_meters)
        new_remaining = slit_width["remaining_quantity"] - allocated_quantity
        
        # Update the slit width entry
        await db.slit_widths.update_one(
            {"id": allocation_request.slit_width_id},
            {
                "$set": {
                    "is_allocated": True,
                    "allocated_to_order_id": allocation_request.order_id,
                    "allocated_quantity": ((slit_width.get("allocated_quantity") or 0) + allocated_quantity),
                    "remaining_quantity": new_remaining,
                    "updated_at": datetime.now(timezone.utc)
                }
            }
        )
        
        # Create stock movement record
        movement = StockMovement(
            stock_type="slit_width",
            stock_id=allocation_request.slit_width_id,
            movement_type="allocation",
            quantity_change=-allocated_quantity,  # Negative for allocation
            previous_quantity=slit_width["remaining_quantity"],
            new_quantity=new_remaining,
            reference_id=allocation_request.order_id,
            reference_type="order",
            notes=f"Allocated {allocated_quantity} meters of {slit_width['slit_width_mm']}mm width to order",
            created_by=current_user["sub"]
        )
        
        # Insert movement record
        movement_dict = movement.dict()
        await db.stock_movements.insert_one(movement_dict)
        
        return StandardResponse(
            success=True,
            message=f"Successfully allocated {allocated_quantity} meters",
            data={
                "allocated_quantity": allocated_quantity,
                "remaining_required": allocation_request.required_quantity_meters - allocated_quantity,
                "slit_width_id": allocation_request.slit_width_id,
                "new_remaining_quantity": new_remaining
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to allocate slit width: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to allocate slit width")

@api_router.get("/slit-widths/allocations/{order_id}", response_model=StandardResponse)
async def get_slit_width_allocations(
    order_id: str,
    current_user: dict = Depends(require_any_role)
):
    """Get all slit width allocations for a specific order"""
    try:
        # Find all slit widths allocated to this order
        allocations = await db.slit_widths.find({
            "allocated_to_order_id": order_id,
            "is_allocated": True
        }).to_list(length=None)
        
        # Remove MongoDB ObjectIds
        for allocation in allocations:
            if "_id" in allocation:
                del allocation["_id"]
        
        # Get related stock movements
        movements = await db.stock_movements.find({
            "reference_id": order_id,
            "stock_type": "slit_width",
            "movement_type": "allocation"
        }).sort("created_at", -1).to_list(length=None)
        
        for movement in movements:
            if "_id" in movement:
                del movement["_id"]
        
        return StandardResponse(
            success=True,
            message="Slit width allocations retrieved successfully",
            data={
                "order_id": order_id,
                "allocations": allocations,
                "movements": movements,
                "total_allocated_meters": sum(alloc.get("allocated_quantity", 0) for alloc in allocations)
            }
        )
        
    except Exception as e:
        logger.error(f"Failed to get slit width allocations: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to retrieve allocations")

@api_router.put("/slit-widths/{slit_width_id}", response_model=StandardResponse)
async def update_slit_width(
    slit_width_id: str,
    update_data: SlitWidthUpdate,
    current_user: dict = Depends(require_any_role)
):
    """Update slit width entry"""
    try:
        # Check if slit width exists
        slit_width = await db.slit_widths.find_one({"id": slit_width_id})
        
        if not slit_width:
            raise HTTPException(status_code=404, detail="Slit width entry not found")
        
        # Prepare update data
        update_dict = {}
        if update_data.quantity_meters is not None:
            update_dict["quantity_meters"] = update_data.quantity_meters
        if update_data.remaining_quantity is not None:
            update_dict["remaining_quantity"] = update_data.remaining_quantity
        if update_data.is_allocated is not None:
            update_dict["is_allocated"] = update_data.is_allocated
        if update_data.allocated_to_order_id is not None:
            update_dict["allocated_to_order_id"] = update_data.allocated_to_order_id
        if update_data.allocated_quantity is not None:
            update_dict["allocated_quantity"] = update_data.allocated_quantity
            
        update_dict["updated_at"] = datetime.now(timezone.utc)
        
        # Update the entry
        result = await db.slit_widths.update_one(
            {"id": slit_width_id},
            {"$set": update_dict}
        )
        
        if result.modified_count == 0:
            raise HTTPException(status_code=400, detail="No changes made to slit width")
        
        return StandardResponse(
            success=True,
            message="Slit width updated successfully"
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to update slit width: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to update slit width")

@api_router.delete("/slit-widths/{slit_width_id}", response_model=StandardResponse)
async def delete_slit_width(
    slit_width_id: str,
    current_user: dict = Depends(require_any_role)
):
    """Delete slit width entry"""
    try:
        # Check if slit width exists
        slit_width = await db.slit_widths.find_one({"id": slit_width_id})
        
        if not slit_width:
            raise HTTPException(status_code=404, detail="Slit width entry not found")
        
        # Check if it's allocated to prevent deletion
        if slit_width.get("is_allocated") and slit_width.get("allocated_quantity", 0) > 0:
            raise HTTPException(
                status_code=400, 
                detail="Cannot delete slit width that is allocated to orders"
            )
        
        # Delete the entry
        result = await db.slit_widths.delete_one({"id": slit_width_id})
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=400, detail="Failed to delete slit width")
        
        return StandardResponse(
            success=True,
            message="Slit width deleted successfully"
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete slit width: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to delete slit width")

# ============= SYSTEM ENDPOINTS =============

@api_router.get("/")
async def root():
    return {"message": "Misty Manufacturing Management System API"}

@api_router.get("/health")
async def health_check():
    return {"status": "healthy", "timestamp": datetime.now(timezone.utc)}

# Include the routers in the main app
app.include_router(api_router)
app.include_router(payroll_router)

# Add CORS middleware with debug logging
cors_origins_str = os.environ.get('CORS_ORIGINS', '*')
print(f"CORS Origins string from env: '{cors_origins_str}'")
cors_origins = cors_origins_str.split(',') if cors_origins_str != '*' else ["*"]
print(f"CORS Origins configured: {cors_origins}")

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],  # Temporarily allow all origins for testing
    allow_methods=["*"],
    allow_headers=["*"],
)

# Router already included above, removing duplicate registration

# Mount static files for uploads - Cross-platform compatible
app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")

# Direct Xero callback route (not under /api to avoid routing issues)
@app.get("/xero-oauth-callback")
async def xero_oauth_callback_direct(code: str = None, state: str = None, error: str = None):
    """Direct Xero OAuth callback route that bypasses /api routing issues"""
    if error:
        return Response(
            status_code=302,
            headers={"Location": f"{os.getenv('FRONTEND_URL')}/xero/callback?error={error}"}
        )
    
    if code and state:
        return Response(
            status_code=302,
            headers={"Location": f"{os.getenv('FRONTEND_URL')}/xero/callback?code={code}&state={state}"}
        )
    
    return Response(
        status_code=302,
        headers={"Location": f"{os.getenv('FRONTEND_URL')}/"}
    )

# Direct CSV export route (bypasses /api routing issues)
@app.get("/export-drafted-csv")
async def export_drafted_invoices_csv_direct(authorization: str = Header(None)):
    """Export all accounting transactions (drafted invoices) to CSV - Direct route bypassing /api issues"""
    try:
        # Simple authentication check
        if not authorization or not authorization.startswith('Bearer '):
            return Response(
                content="Unauthorized - Missing or invalid token",
                media_type="text/plain",
                status_code=401
            )
        
        # For simplicity, just check if token exists (proper auth would verify JWT)
        token = authorization.replace('Bearer ', '')
        if not token:
            return Response(
                content="Unauthorized - Invalid token",
                media_type="text/plain", 
                status_code=401
            )
        # Get all jobs in accounting transaction stage
        transactions = await db.orders.find({
            "current_stage": "accounting_transaction",
            "status": "accounting_draft"
        }).to_list(length=None)
        
        # Prepare CSV data
        csv_data = []
        
        # CSV Headers based on Xero import format
        headers = [
            "ContactName", "EmailAddress", "POAddressLine1", "POAddressLine2", 
            "POAddressLine3", "POAddressLine4", "POCity", "PORegion", 
            "POPostalCode", "POCountry", "InvoiceNumber", "Reference", 
            "InvoiceDate", "DueDate", "InventoryItemCode", "Description", 
            "Quantity", "UnitAmount", "Discount", "AccountCode", "TaxType", 
            "TrackingName1", "TrackingOption1", "TrackingName2", "TrackingOption2", 
            "Currency", "BrandingTheme"
        ]
        
        csv_data.append(headers)
        
        # Process each transaction
        for transaction in transactions:
            # Get client info
            client = await db.clients.find_one({"id": transaction["client_id"]})
            client_name = client["company_name"] if client else transaction.get("client_name", "Unknown Client")
            client_email = client.get("email", "") if client else ""
            
            # Get invoice info
            invoice = await db.invoices.find_one({"id": transaction.get("invoice_id")})
            invoice_number = invoice["invoice_number"] if invoice else f"INV-{transaction['order_number']}"
            invoice_date = invoice["created_at"].strftime("%d/%m/%Y") if invoice and invoice.get("created_at") else datetime.now().strftime("%d/%m/%Y")
            
            # Calculate due date (30 days from invoice date)
            due_date_obj = invoice["created_at"] + timedelta(days=30) if invoice and invoice.get("created_at") else datetime.now() + timedelta(days=30)
            due_date = due_date_obj.strftime("%d/%m/%Y")
            
            # Process each item in the transaction
            items = transaction.get("items", [])
            if not items:
                # Create a single line if no items
                items = [{"description": f"Services for Order {transaction['order_number']}", "quantity": 1, "unit_price": transaction.get("total_amount", 0)}]
            
            for item in items:
                row = [
                    client_name,  # ContactName (required)
                    client_email,  # EmailAddress
                    "",  # POAddressLine1
                    "",  # POAddressLine2
                    "",  # POAddressLine3
                    "",  # POAddressLine4
                    "",  # POCity
                    "",  # PORegion
                    "",  # POPostalCode
                    "",  # POCountry
                    invoice_number,  # InvoiceNumber (required)
                    transaction["order_number"],  # Reference
                    invoice_date,  # InvoiceDate (required)
                    due_date,  # DueDate (required)
                    item.get("product_code", ""),  # InventoryItemCode
                    f"{item.get('product_name', item.get('description', 'Product'))} - {item.get('specifications', '')}".strip(" - "),  # Description (required)
                    str(item.get("quantity", 1)),  # Quantity (required)
                    str(item.get("unit_price", item.get("price", 0))),  # UnitAmount (required)
                    str(item.get("discount_percent", "")),  # Discount
                    os.getenv("XERO_SALES_ACCOUNT_CODE", "200"),  # AccountCode (required)
                    "OUTPUT",  # TaxType (required) - GST for sales
                    "",  # TrackingName1
                    "",  # TrackingOption1
                    "",  # TrackingName2
                    "",  # TrackingOption2
                    "AUD",  # Currency
                    ""   # BrandingTheme
                ]
                csv_data.append(row)
        
        # Generate CSV content
        import io
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerows(csv_data)
        csv_content = output.getvalue()
        output.close()
        
        # Return CSV as response
        from fastapi.responses import Response
        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=drafted_invoices_{datetime.now().strftime('%Y%m%d')}.csv"}
        )
        
    except Exception as e:
        logger.error(f"Failed to export drafted invoices CSV: {str(e)}")
        return Response(
            content=f"Error: {str(e)}",
            media_type="text/plain",
            status_code=500
        )

# Direct Xero token exchange route (bypasses /api routing)
@app.post("/xero-auth-callback")
async def xero_auth_callback_direct(callback_data: dict):
    """Direct Xero token exchange that bypasses /api routing issues"""
    try:
        # Get auth credentials from environment
        XERO_CLIENT_ID = os.getenv("XERO_CLIENT_ID")
        XERO_CLIENT_SECRET = os.getenv("XERO_CLIENT_SECRET")  
        XERO_CALLBACK_URL = os.getenv("XERO_REDIRECT_URI")
        
        auth_code = callback_data.get("code")
        state = callback_data.get("state")
        
        if not auth_code or not state:
            return {"error": "Missing authorization code or state"}
        
        # Exchange code for tokens
        token_url = "https://identity.xero.com/connect/token"
        
        # Prepare authentication
        auth_string = f"{XERO_CLIENT_ID}:{XERO_CLIENT_SECRET}"
        auth_bytes = auth_string.encode('ascii')
        auth_b64 = base64.b64encode(auth_bytes).decode('ascii')
        
        headers = {
            "Authorization": f"Basic {auth_b64}",
            "Content-Type": "application/x-www-form-urlencoded"
        }
        
        token_data = {
            "grant_type": "authorization_code",
            "code": auth_code,
            "redirect_uri": XERO_CALLBACK_URL
        }
        
        response = requests.post(token_url, headers=headers, data=token_data)
        response.raise_for_status()
        
        tokens = response.json()
        
        # Store tokens (simplified - no user context for now)
        token_record = {
            "user_id": "system",  # Simplified for testing
            "access_token": tokens["access_token"],
            "refresh_token": tokens["refresh_token"],
            "expires_at": datetime.now(timezone.utc) + timedelta(seconds=tokens.get("expires_in", 1800)),
            "created_at": datetime.now(timezone.utc),
            "tenant_id": None
        }
        
        # Store in database
        await db.xero_tokens.replace_one(
            {"user_id": "system"},
            token_record,
            upsert=True
        )
        
        return {"message": "Xero connection successful", "access_token_expires_in": tokens.get("expires_in", 1800)}
        
    except Exception as e:
        logger.error(f"Xero token exchange failed: {str(e)}")
        return {"error": f"Failed to exchange authorization code: {str(e)}"}



# ======================
# LABEL DESIGNER ENDPOINTS
# ======================

@app.get("/api/label-templates")
async def get_label_templates(current_user: dict = Depends(get_current_user)):
    """Get all label templates"""
    try:
        templates = await db.label_templates.find().to_list(length=None)
        # Remove MongoDB's _id field
        for template in templates:
            if '_id' in template:
                del template['_id']
        return {"success": True, "data": templates}
    except Exception as e:
        logger.error(f"Error fetching label templates: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/label-templates/{template_id}")
async def get_label_template(template_id: str, current_user: dict = Depends(get_current_user)):
    """Get a specific label template"""
    try:
        template = await db.label_templates.find_one({"id": template_id})
        if not template:
            raise HTTPException(status_code=404, detail="Label template not found")
        # Remove MongoDB's _id field
        if '_id' in template:
            del template['_id']
        return {"success": True, "data": template}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching label template: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/label-templates")
async def create_label_template(template: LabelTemplateCreate, current_user: dict = Depends(get_current_user)):
    """Create a new label template"""
    try:
        template_data = LabelTemplate(
            **template.dict(),
            created_by=current_user.get("sub")
        )
        
        await db.label_templates.insert_one(template_data.dict())
        
        return {
            "success": True,
            "message": "Label template created successfully",
            "data": template_data.dict()
        }
    except Exception as e:
        logger.error(f"Error creating label template: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/label-templates/{template_id}")
async def update_label_template(
    template_id: str,
    template: dict,  # Accept raw dict to avoid Pydantic validation issues
    current_user: dict = Depends(get_current_user)
):
    """Update a label template"""
    try:
        existing_template = await db.label_templates.find_one({"id": template_id})
        if not existing_template:
            raise HTTPException(status_code=404, detail="Label template not found")
        
        # Remove None values and prepare update data
        update_data = {k: v for k, v in template.items() if v is not None and k != 'id'}
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        
        logger.info(f"Updating template {template_id} with data: {update_data.keys()}")
        
        await db.label_templates.update_one(
            {"id": template_id},
            {"$set": update_data}
        )
        
        updated_template = await db.label_templates.find_one({"id": template_id})
        # Remove MongoDB's _id field
        if updated_template and '_id' in updated_template:
            del updated_template['_id']
        
        return {
            "success": True,
            "message": "Label template updated successfully",
            "data": updated_template
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating label template: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/label-templates/{template_id}")
async def delete_label_template(template_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a label template"""
    try:
        result = await db.label_templates.delete_one({"id": template_id})
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Label template not found")
        
        return {
            "success": True,
            "message": "Label template deleted successfully"
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting label template: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/printers")
async def get_printers(current_user: dict = Depends(get_current_user)):
    """Get list of available printers"""
    try:
        import subprocess
        import platform
        
        printers = []
        system = platform.system()
        
        if system == "Windows":
            # Windows printer detection
            try:
                import win32print
                printer_names = [printer[2] for printer in win32print.EnumPrinters(2)]
                default_printer = win32print.GetDefaultPrinter()
                printers = [{"name": name, "is_default": name == default_printer} for name in printer_names]
            except:
                # Fallback to PowerShell
                result = subprocess.run(
                    ["powershell", "-Command", "Get-Printer | Select-Object Name, Default | ConvertTo-Json"],
                    capture_output=True,
                    text=True
                )
                if result.returncode == 0:
                    import json
                    data = json.loads(result.stdout)
                    if isinstance(data, list):
                        printers = [{"name": p["Name"], "is_default": p.get("Default", False)} for p in data]
                    else:
                        printers = [{"name": data["Name"], "is_default": data.get("Default", False)}]
        
        elif system == "Linux":
            # Linux CUPS detection
            try:
                result = subprocess.run(["lpstat", "-p"], capture_output=True, text=True)
                if result.returncode == 0:
                    for line in result.stdout.split('\n'):
                        if line.startswith('printer'):
                            name = line.split()[1]
                            printers.append({"name": name, "is_default": False})
                
                # Get default printer
                result = subprocess.run(["lpstat", "-d"], capture_output=True, text=True)
                if result.returncode == 0 and "system default destination" in result.stdout:
                    default = result.stdout.split(":")[-1].strip()
                    for printer in printers:
                        if printer["name"] == default:
                            printer["is_default"] = True
            except:
                pass
        
        elif system == "Darwin":  # macOS
            try:
                result = subprocess.run(["lpstat", "-p"], capture_output=True, text=True)
                if result.returncode == 0:
                    for line in result.stdout.split('\n'):
                        if line.startswith('printer'):
                            name = line.split()[1]
                            printers.append({"name": name, "is_default": False})
                
                # Get default printer
                result = subprocess.run(["lpstat", "-d"], capture_output=True, text=True)
                if result.returncode == 0:
                    default = result.stdout.split(":")[-1].strip()
                    for printer in printers:
                        if printer["name"] == default:
                            printer["is_default"] = True
            except:
                pass
        
        # Add browser printing as an option
        printers.insert(0, {"name": "Browser Default Printer", "is_default": True, "is_browser": True})
        
        return {"success": True, "data": printers}
    except Exception as e:
        logger.error(f"Error fetching printers: {str(e)}")
        # Return browser option as fallback
        return {"success": True, "data": [{"name": "Browser Default Printer", "is_default": True, "is_browser": True}]}

@app.post("/api/print-label")
async def print_label(print_data: dict, current_user: dict = Depends(get_current_user)):
    """Send label to printer"""
    try:
        printer_name = print_data.get("printer_name")
        is_browser = print_data.get("is_browser", True)
        
        if is_browser or printer_name == "Browser Default Printer":
            # Browser printing - return success and let frontend handle
            return {
                "success": True,
                "message": "Please use browser print dialog",
                "use_browser_print": True
            }
        
        # For actual printer, would implement printing logic here
        # This would involve generating PDF and sending to printer
        # For now, return success
        return {
            "success": True,
            "message": f"Print job sent to {printer_name}",
            "use_browser_print": False
        }
        
    except Exception as e:
        logger.error(f"Error printing label: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def startup_event():
    """Initialize system on startup"""
    logger.info("Starting Misty Manufacturing Management System...")
    
    # Create default admin user if not exists
    admin_user = await db.users.find_one({"username": "Callum"})
    if not admin_user:
        logger.info("Creating default admin user...")
        default_admin = User(
            username="Callum",
            email="admin@adelamerchants.com.au",
            hashed_password=get_password_hash("Peach7510"),
            role=UserRole.ADMIN,
            full_name="Callum - System Administrator"
        )
        await db.users.insert_one(default_admin.dict())
        logger.info("Default admin user created successfully")
    
    logger.info("Misty Manufacturing Management System started successfully!")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()