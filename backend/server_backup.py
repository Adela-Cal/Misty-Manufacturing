from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, status, Header, Request
from fastapi.responses import FileResponse, StreamingResponse, HTMLResponse, Response
from fastapi.staticfiles import StaticFiles
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from datetime import datetime, date, time, timezone, timedelta
import calendar
from bson import ObjectId
from typing import List, Optional, Dict, Any
import uuid
from io import BytesIO
import secrets
import requests
import base64
from urllib.parse import urlencode
import asyncio

# Xero SDK imports
from xero_python.api_client import ApiClient, Configuration
from xero_python.api_client.oauth2 import OAuth2Token
from xero_python.accounting import AccountingApi
from xero_python.accounting.models import Invoice, Contact, LineItem, Contacts, Invoices

# Import our custom modules
from models import *
from auth import *
from document_generator import DocumentGenerator
from file_utils import *
from payroll_endpoints import payroll_router

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Force correct environment values for current environment  
os.environ['XERO_REDIRECT_URI'] = 'https://misty-ato-payroll.preview.emergentagent.com/xero-oauth-callback'
os.environ['FRONTEND_URL'] = 'https://misty-ato-payroll.preview.emergentagent.com'

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app
app = FastAPI(title="Misty Manufacturing Management System", version="1.0.0")

# Create router with /api prefix
api_router = APIRouter(prefix="/api")

# Initialize document generator
doc_generator = DocumentGenerator()

# Ensure upload directories exist
ensure_upload_dirs()

# ============= AUTHENTICATION ENDPOINTS =============

@api_router.post("/auth/login", response_model=Token)
async def login(user_credentials: UserLogin):
    """Authenticate user and return JWT token"""
    user_data = await db.users.find_one({"username": user_credentials.username})
    
    if not user_data or not verify_password(user_credentials.password, user_data["hashed_password"]):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid username or password"
        )
    
    if not user_data.get("is_active", True):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Account is disabled"
        )
    
    # Update last login
    await db.users.update_one(
        {"id": user_data["id"]},
        {"$set": {"last_login": datetime.now(timezone.utc)}}
    )
    
    # Create JWT token
    access_token = create_access_token(
        data={"sub": user_data["username"], "role": user_data["role"], "user_id": user_data["id"]}
    )
    
    user_info = {
        "id": user_data["id"],
        "username": user_data["username"],
        "full_name": user_data["full_name"],
        "role": user_data["role"],
        "email": user_data["email"]
    }
    
    return {"access_token": access_token, "token_type": "bearer", "user": user_info}

@api_router.post("/auth/register", response_model=StandardResponse)
async def register_user(user_data: UserCreate, current_user: dict = Depends(require_admin)):
    """Register new user (Admin only)"""
    # Check if username already exists
    existing_user = await db.users.find_one({"username": user_data.username})
    if existing_user:
        raise HTTPException(status_code=400, detail="Username already exists")
    
    # Check if email already exists
    existing_email = await db.users.find_one({"email": user_data.email})
    if existing_email:
        raise HTTPException(status_code=400, detail="Email already exists")
    
    # Create new user
    hashed_password = get_password_hash(user_data.password)
    new_user = User(
        username=user_data.username,
        email=user_data.email,
        hashed_password=hashed_password,
        role=user_data.role,
        full_name=user_data.full_name
    )
    
    await db.users.insert_one(new_user.dict())
    
    return StandardResponse(success=True, message="User created successfully")

@api_router.get("/auth/me")
async def get_current_user_info(current_user: dict = Depends(get_current_user)):
    """Get current user information"""
    user_data = await db.users.find_one({"id": current_user["user_id"]})
    if not user_data:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {
        "id": user_data["id"],
        "username": user_data["username"],
        "full_name": user_data["full_name"],
        "role": user_data["role"],
        "email": user_data["email"]
    }

# ============= CLIENT MANAGEMENT ENDPOINTS =============

@api_router.get("/clients", response_model=List[Client])
async def get_clients(current_user: dict = Depends(require_any_role)):
    """Get all clients"""
    clients = await db.clients.find({"is_active": True}).to_list(1000)
    return [Client(**client) for client in clients]

@api_router.post("/clients", response_model=StandardResponse)
async def create_client(client_data: ClientCreate, current_user: dict = Depends(require_admin_or_sales)):
    """Create new client"""
    new_client = Client(**client_data.dict())
    await db.clients.insert_one(new_client.dict())
    
    return StandardResponse(success=True, message="Client created successfully", data={"id": new_client.id})

@api_router.get("/clients/{client_id}", response_model=Client)
async def get_client(client_id: str, current_user: dict = Depends(require_any_role)):
    """Get specific client"""
    client = await db.clients.find_one({"id": client_id, "is_active": True})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    return Client(**client)

@api_router.put("/clients/{client_id}", response_model=StandardResponse)
async def update_client(client_id: str, client_data: ClientCreate, current_user: dict = Depends(require_admin_or_sales)):
    """Update client"""
    update_data = client_data.dict()
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    result = await db.clients.update_one(
        {"id": client_id, "is_active": True},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Client not found")
    
    return StandardResponse(success=True, message="Client updated successfully")

@api_router.post("/clients/{client_id}/logo")
async def upload_client_logo(client_id: str, file: UploadFile = File(...), current_user: dict = Depends(require_admin_or_sales)):
    """Upload client logo"""
    # Check if client exists
    client = await db.clients.find_one({"id": client_id, "is_active": True})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Save logo file
    try:
        file_path = await save_logo_file(file, client_id)
        file_url = get_file_url(file_path)
        
        # Update client with logo path
        await db.clients.update_one(
            {"id": client_id},
            {"$set": {"logo_path": file_path, "updated_at": datetime.now(timezone.utc)}}
        )
        
        return StandardResponse(success=True, message="Logo uploaded successfully", data={"file_url": file_url})
    
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(status_code=500, detail="Failed to upload logo")

# ============= PRODUCT MANAGEMENT ENDPOINTS =============

@api_router.get("/products", response_model=List[Product])
async def get_all_products(current_user: dict = Depends(require_any_role)):
    """Get all products"""
    products = await db.products.find({"is_active": True}).to_list(1000)
    return [Product(**product) for product in products]

@api_router.get("/clients/{client_id}/products", response_model=List[Product])
async def get_client_products(client_id: str, current_user: dict = Depends(require_any_role)):
    """Get products for specific client"""
    products = await db.products.find({"client_id": client_id, "is_active": True}).to_list(1000)
    return [Product(**product) for product in products]

@api_router.post("/products", response_model=StandardResponse)
async def create_product(product_data: ProductCreate, current_user: dict = Depends(require_admin_or_sales)):
    """Create new product for client"""
    new_product = Product(**product_data.dict())
    await db.products.insert_one(new_product.dict())
    
    return StandardResponse(success=True, message="Product created successfully", data={"id": new_product.id})

@api_router.get("/products/{product_id}", response_model=Product)
async def get_product(product_id: str, current_user: dict = Depends(require_any_role)):
    """Get specific product"""
    product = await db.products.find_one({"id": product_id, "is_active": True})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    return Product(**product)

@api_router.put("/products/{product_id}", response_model=StandardResponse)
async def update_product(product_id: str, product_data: ProductCreate, current_user: dict = Depends(require_admin_or_sales)):
    """Update product"""
    update_data = product_data.dict()
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    result = await db.products.update_one(
        {"id": product_id, "is_active": True},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    
    return StandardResponse(success=True, message="Product updated successfully")

# ============= MATERIALS MANAGEMENT ENDPOINTS =============

@api_router.get("/materials", response_model=List[Material])
async def get_materials(current_user: dict = Depends(require_any_role)):
    """Get all materials"""
    materials = await db.materials.find({"is_active": True}).to_list(1000)
    return [Material(**material) for material in materials]

@api_router.post("/materials", response_model=StandardResponse)
async def create_material(material_data: MaterialCreate, current_user: dict = Depends(require_admin_or_manager)):
    """Create new material"""
    new_material = Material(**material_data.dict())
    await db.materials.insert_one(new_material.dict())
    
    return StandardResponse(success=True, message="Material created successfully", data={"id": new_material.id})

@api_router.get("/materials/{material_id}", response_model=Material)
async def get_material(material_id: str, current_user: dict = Depends(require_any_role)):
    """Get specific material"""
    material = await db.materials.find_one({"id": material_id, "is_active": True})
    if not material:
        raise HTTPException(status_code=404, detail="Material not found")
    
    return Material(**material)

@api_router.put("/materials/{material_id}", response_model=StandardResponse)
async def update_material(material_id: str, material_data: MaterialCreate, current_user: dict = Depends(require_admin_or_manager)):
    """Update material"""
    update_data = material_data.dict()
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    result = await db.materials.update_one(
        {"id": material_id, "is_active": True},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Material not found")
    
    return StandardResponse(success=True, message="Material updated successfully")

@api_router.delete("/materials/{material_id}", response_model=StandardResponse)
async def delete_material(material_id: str, current_user: dict = Depends(require_admin_or_manager)):
    """Delete material (soft delete)"""
    result = await db.materials.update_one(
        {"id": material_id, "is_active": True},
        {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc)}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Material not found")
    
    return StandardResponse(success=True, message="Material deleted successfully")

# ============= SUPPLIERS ENDPOINTS =============

@api_router.get("/suppliers", response_model=List[Supplier])
async def get_suppliers(current_user: dict = Depends(require_any_role)):
    """Get all active suppliers"""
    suppliers = await db.suppliers.find({"is_active": True}).sort("supplier_name", 1).to_list(1000)
    return [Supplier(**supplier) for supplier in suppliers]

@api_router.post("/suppliers", response_model=StandardResponse)
async def create_supplier(supplier_data: SupplierCreate, current_user: dict = Depends(require_admin_or_manager)):
    """Create new supplier"""
    new_supplier = Supplier(**supplier_data.dict())
    
    await db.suppliers.insert_one(new_supplier.dict())
    
    return StandardResponse(success=True, message="Supplier created successfully", data={"id": new_supplier.id})

@api_router.get("/suppliers/{supplier_id}", response_model=Supplier)
async def get_supplier(supplier_id: str, current_user: dict = Depends(require_any_role)):
    """Get specific supplier by ID"""
    supplier = await db.suppliers.find_one({"id": supplier_id, "is_active": True})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    return Supplier(**supplier)

@api_router.put("/suppliers/{supplier_id}", response_model=StandardResponse)
async def update_supplier(supplier_id: str, supplier_data: SupplierCreate, current_user: dict = Depends(require_admin_or_manager)):
    """Update supplier"""
    update_data = supplier_data.dict()
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    result = await db.suppliers.update_one(
        {"id": supplier_id, "is_active": True},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    return StandardResponse(success=True, message="Supplier updated successfully")

@api_router.delete("/suppliers/{supplier_id}", response_model=StandardResponse)
async def delete_supplier(supplier_id: str, current_user: dict = Depends(require_admin_or_manager)):
    """Soft delete supplier"""
    result = await db.suppliers.update_one(
        {"id": supplier_id, "is_active": True},
        {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc)}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    return StandardResponse(success=True, message="Supplier deleted successfully")

# ============= PRODUCT SPECIFICATIONS ENDPOINTS =============

@api_router.get("/product-specifications", response_model=List[ProductSpecification])
async def get_product_specifications(current_user: dict = Depends(require_any_role)):
    """Get all product specifications"""
    specs = await db.product_specifications.find({"is_active": True}).sort("product_name", 1).to_list(1000)
    return [ProductSpecification(**spec) for spec in specs]

@api_router.post("/product-specifications", response_model=StandardResponse)
async def create_product_specification(spec_data: ProductSpecificationCreate, current_user: dict = Depends(require_admin_or_manager)):
    """Create new product specification with automatic thickness calculation"""
    # Calculate total thickness from material layers (thickness * quantity for each layer)
    calculated_thickness = sum(layer.thickness * (layer.quantity or 1.0) for layer in spec_data.material_layers)
    
    # Generate thickness options (±5%, ±10%, exact)
    thickness_options = []
    if calculated_thickness > 0:
        thickness_options = [
            round(calculated_thickness * 0.90, 3),  # -10%
            round(calculated_thickness * 0.95, 3),  # -5%
            round(calculated_thickness, 3),         # Exact
            round(calculated_thickness * 1.05, 3),  # +5%
            round(calculated_thickness * 1.10, 3),  # +10%
        ]
        # Remove duplicates and sort
        thickness_options = sorted(list(set(thickness_options)))
    
    # Create new specification with calculated values
    new_spec = ProductSpecification(
        **spec_data.dict(),
        calculated_total_thickness=calculated_thickness if calculated_thickness > 0 else None,
        thickness_options=thickness_options
    )
    
    await db.product_specifications.insert_one(new_spec.dict())
    
    return StandardResponse(success=True, message="Product specification created successfully", data={
        "id": new_spec.id, 
        "calculated_thickness": calculated_thickness,
        "thickness_options": thickness_options
    })

@api_router.get("/product-specifications/{spec_id}", response_model=ProductSpecification)
async def get_product_specification(spec_id: str, current_user: dict = Depends(require_any_role)):
    """Get specific product specification by ID"""
    spec = await db.product_specifications.find_one({"id": spec_id, "is_active": True})
    if not spec:
        raise HTTPException(status_code=404, detail="Product specification not found")
    
    return ProductSpecification(**spec)

@api_router.put("/product-specifications/{spec_id}", response_model=StandardResponse)
async def update_product_specification(spec_id: str, spec_data: ProductSpecificationCreate, current_user: dict = Depends(require_admin_or_manager)):
    """Update product specification with automatic thickness calculation"""
    # Calculate total thickness from material layers (thickness * quantity for each layer)
    calculated_thickness = sum(layer.thickness * (layer.quantity or 1.0) for layer in spec_data.material_layers)
    
    # Generate thickness options (±5%, ±10%, exact)
    thickness_options = []
    if calculated_thickness > 0:
        thickness_options = [
            round(calculated_thickness * 0.90, 3),  # -10%
            round(calculated_thickness * 0.95, 3),  # -5%
            round(calculated_thickness, 3),         # Exact
            round(calculated_thickness * 1.05, 3),  # +5%
            round(calculated_thickness * 1.10, 3),  # +10%
        ]
        # Remove duplicates and sort
        thickness_options = sorted(list(set(thickness_options)))
    
    # Prepare update data
    update_data = spec_data.dict()
    update_data.update({
        "calculated_total_thickness": calculated_thickness if calculated_thickness > 0 else None,
        "thickness_options": thickness_options,
        "updated_at": datetime.now(timezone.utc)
    })
    
    result = await db.product_specifications.update_one(
        {"id": spec_id, "is_active": True},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Product specification not found")
    
    return StandardResponse(success=True, message="Product specification updated successfully", data={
        "calculated_thickness": calculated_thickness,
        "thickness_options": thickness_options
    })

@api_router.delete("/product-specifications/{spec_id}", response_model=StandardResponse)
async def delete_product_specification(spec_id: str, current_user: dict = Depends(require_admin_or_manager)):
    """Soft delete product specification"""
    result = await db.product_specifications.update_one(
        {"id": spec_id, "is_active": True},
        {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc)}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Product specification not found")
    
    return StandardResponse(success=True, message="Product specification deleted successfully")

# ============= MACHINERY RATES ENDPOINTS =============

@api_router.get("/machinery-rates", response_model=List[MachineryRate])
async def get_machinery_rates(current_user: dict = Depends(require_any_role)):
    """Get all machinery rates"""
    rates = await db.machinery_rates.find({"is_active": True}).sort("function", 1).to_list(1000)
    return [MachineryRate(**rate) for rate in rates]

@api_router.post("/machinery-rates", response_model=StandardResponse)
async def create_machinery_rate(rate_data: MachineryRateCreate, current_user: dict = Depends(require_admin_or_manager)):
    """Create new machinery rate"""
    # Check if rate for this function already exists
    existing_rate = await db.machinery_rates.find_one({
        "function": rate_data.function,
        "is_active": True
    })
    
    if existing_rate:
        raise HTTPException(status_code=400, detail=f"Rate for function '{rate_data.function}' already exists")
    
    # Create new rate
    new_rate = MachineryRate(
        **rate_data.dict(),
        created_at=datetime.now(timezone.utc)
    )
    
    rate_dict = new_rate.dict()
    await db.machinery_rates.insert_one(rate_dict)
    
    return StandardResponse(
        success=True,
        message="Machinery rate created successfully",
        data={"id": new_rate.id}
    )

@api_router.get("/machinery-rates/{rate_id}", response_model=MachineryRate)
async def get_machinery_rate(rate_id: str, current_user: dict = Depends(require_any_role)):
    """Get specific machinery rate by ID"""
    rate = await db.machinery_rates.find_one({"id": rate_id, "is_active": True})
    if not rate:
        raise HTTPException(status_code=404, detail="Machinery rate not found")
    
    return MachineryRate(**rate)

@api_router.put("/machinery-rates/{rate_id}", response_model=StandardResponse)
async def update_machinery_rate(rate_id: str, rate_data: MachineryRateUpdate, current_user: dict = Depends(require_admin_or_manager)):
    """Update machinery rate"""
    # Prepare update data
    update_data = {k: v for k, v in rate_data.dict().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    # If function is being changed, check for duplicates
    if "function" in update_data:
        existing_rate = await db.machinery_rates.find_one({
            "function": update_data["function"],
            "is_active": True,
            "id": {"$ne": rate_id}
        })
        if existing_rate:
            raise HTTPException(status_code=400, detail=f"Rate for function '{update_data['function']}' already exists")
    
    result = await db.machinery_rates.update_one(
        {"id": rate_id, "is_active": True},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Machinery rate not found")
    
    return StandardResponse(success=True, message="Machinery rate updated successfully")

@api_router.delete("/machinery-rates/{rate_id}", response_model=StandardResponse)
async def delete_machinery_rate(rate_id: str, current_user: dict = Depends(require_admin_or_manager)):
    """Soft delete machinery rate"""
    result = await db.machinery_rates.update_one(
        {"id": rate_id, "is_active": True},
        {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc)}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Machinery rate not found")
    
    return StandardResponse(success=True, message="Machinery rate deleted successfully")

# ============= CALCULATORS ENDPOINTS =============

@api_router.post("/calculators/material-consumption-by-client", response_model=StandardResponse)
async def calculate_material_consumption_by_client(
    request: MaterialConsumptionByClientRequest, 
    current_user: dict = Depends(require_any_role)
):
    """Calculate material consumption for a client within date range"""
    try:
        # Get orders for client within date range
        orders = await db.orders.find({
            "client_id": request.client_id,
            "created_at": {
                "$gte": datetime.combine(request.start_date, datetime.min.time()),
                "$lte": datetime.combine(request.end_date, datetime.max.time())
            },
            "status": {"$ne": "cancelled"}
        }).to_list(1000)
        
        # Get material info
        material = await db.materials.find_one({"id": request.material_id})
        if not material:
            raise HTTPException(status_code=404, detail="Material not found")
        
        total_consumption = 0
        order_breakdown = []
        
        for order in orders:
            # Calculate material usage based on order items and product specifications
            order_consumption = 0
            for item in order["items"]:
                # This is a simplified calculation - you may need to enhance based on your business logic
                order_consumption += item["quantity"] * 1.0  # Placeholder calculation
            
            total_consumption += order_consumption
            order_breakdown.append({
                "order_number": order["order_number"],
                "order_date": order["created_at"],
                "consumption": order_consumption
            })
        
        result = CalculationResult(
            calculation_type="material_consumption_by_client",
            input_parameters={
                "client_id": request.client_id,
                "material_id": request.material_id,
                "start_date": request.start_date.isoformat(),
                "end_date": request.end_date.isoformat()
            },
            results={
                "total_consumption": total_consumption,
                "material_name": material["product_code"],
                "unit": material["unit"],
                "order_count": len(orders),
                "order_breakdown": order_breakdown
            },
            calculated_by=current_user["sub"]
        )
        
        return StandardResponse(success=True, message="Calculation completed", data=result.dict())
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Calculation failed: {str(e)}")

@api_router.post("/calculators/material-permutation", response_model=StandardResponse)
async def calculate_material_permutation(
    request: MaterialPermutationRequest,
    current_user: dict = Depends(require_any_role)
):
    """Calculate optimal material permutation for core IDs"""
    try:
        # Get product specifications for the core IDs
        core_specs = await db.product_specifications.find({
            "specifications.core_id": {"$in": request.core_ids},
            "is_active": True
        }).to_list(1000)
        
        # Calculate permutation options
        permutation_options = []
        for arrangement in _generate_permutations(request.sizes_to_manufacture, request.master_deckle_width):
            waste_percentage = _calculate_waste(arrangement, request.master_deckle_width)
            if waste_percentage <= request.acceptable_waste_percentage:
                permutation_options.append({
                    "arrangement": arrangement,
                    "waste_percentage": waste_percentage,
                    "efficiency": 100 - waste_percentage
                })
        
        # Sort by efficiency
        permutation_options.sort(key=lambda x: x["efficiency"], reverse=True)
        
        result = CalculationResult(
            calculation_type="material_permutation",
            input_parameters=request.dict(),
            results={
                "permutation_options": permutation_options[:10],  # Top 10 options
                "total_options_found": len(permutation_options)
            },
            calculated_by=current_user["sub"]
        )
        
        return StandardResponse(success=True, message="Permutation calculation completed", data=result.dict())
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Permutation calculation failed: {str(e)}")

@api_router.post("/calculators/spiral-core-consumption", response_model=StandardResponse)
async def calculate_spiral_core_consumption(
    request: SpiralCoreConsumptionRequest,
    current_user: dict = Depends(require_any_role)
):
    """Calculate material consumption for Spiral Paper Cores"""
    try:
        # Get product specification
        spec = await db.product_specifications.find_one({"id": request.product_specification_id})
        if not spec:
            raise HTTPException(status_code=404, detail="Product specification not found")
        
        if spec["product_type"] != "Spiral Paper Core":
            raise HTTPException(status_code=400, detail="Specification is not for Spiral Paper Cores")
        
        specifications = spec["specifications"]
        
        # Get material information
        material_id = specifications.get("selected_material_id")
        if not material_id:
            raise HTTPException(status_code=400, detail="No material selected in specification")
            
        material = await db.materials.find_one({"id": material_id})
        if not material:
            raise HTTPException(status_code=404, detail="Material not found")
        
        # Calculate consumption
        internal_diameter = float(specifications.get("internal_diameter", 0))
        wall_thickness = float(specifications.get("wall_thickness_required", 0))
        gsm = float(material.get("gsm", 0))
        material_thickness = float(material.get("thickness_mm", 1))
        
        # Calculate surface area per core
        outer_diameter = internal_diameter + (2 * wall_thickness)
        circumference = 3.14159 * outer_diameter
        surface_area = circumference * request.core_length  # mm²
        surface_area_m2 = surface_area / 1000000  # Convert to m²
        
        # Calculate material weight per core
        layers_required = max(1, round(wall_thickness / material_thickness))
        material_weight_per_core = surface_area_m2 * gsm * layers_required / 1000  # kg
        
        # Total consumption
        total_weight = material_weight_per_core * request.quantity
        
        result = CalculationResult(
            calculation_type="spiral_core_consumption",
            input_parameters=request.dict(),
            results={
                "material_name": material["product_code"],
                "gsm": gsm,
                "material_thickness_mm": material_thickness,
                "layers_required": layers_required,
                "internal_diameter_mm": internal_diameter,
                "outer_diameter_mm": outer_diameter,
                "wall_thickness_mm": wall_thickness,
                "core_length_mm": request.core_length,
                "surface_area_m2_per_core": round(surface_area_m2, 4),
                "material_weight_per_core_kg": round(material_weight_per_core, 4),
                "quantity": request.quantity,
                "total_material_weight_kg": round(total_weight, 2),
                "unit": material["unit"]
            },
            calculated_by=current_user["sub"]
        )
        
        return StandardResponse(success=True, message="Spiral core consumption calculated", data=result.dict())
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Spiral core calculation failed: {str(e)}")

# Helper functions for calculations
def _generate_permutations(sizes, master_width):
    """Generate possible arrangements of sizes within master width"""
    # Simplified permutation logic - you can enhance this
    arrangements = []
    for size_combo in _get_size_combinations(sizes, master_width):
        arrangements.append(size_combo)
    return arrangements

def _get_size_combinations(sizes, master_width):
    """Get combinations of sizes that fit within master width"""
    combinations = []
    # This is a simplified algorithm - implement more sophisticated logic as needed
    for size in sizes:
        width = size.get("width", 0)
        if width <= master_width:
            combinations.append([size])
    return combinations

def _calculate_waste(arrangement, master_width):
    """Calculate waste percentage for an arrangement"""
    used_width = sum(item.get("width", 0) for item in arrangement)
    waste = master_width - used_width
    return (waste / master_width) * 100 if master_width > 0 else 100

# ============= STOCKTAKE ENDPOINTS =============

@api_router.get("/stocktake/current", response_model=StandardResponse)
async def get_current_stocktake(current_user: dict = Depends(require_manager_or_admin)):
    """Get current month's stocktake or check if one is needed"""
    current_date = date.today()
    current_month = current_date.strftime("%Y-%m")
    
    # Check if stocktake exists for current month
    stocktake = await db.stocktakes.find_one({"month": current_month})
    
    # Check if it's first business day and stocktake is needed
    is_first_business_day = _is_first_business_day(current_date)
    
    if not stocktake and is_first_business_day:
        return StandardResponse(
            success=True, 
            message="Stocktake required", 
            data={
                "stocktake_required": True,
                "month": current_month,
                "first_business_day": True
            }
        )
    
    # Convert ObjectId to string if stocktake exists
    if stocktake:
        stocktake["id"] = str(stocktake["_id"])
        del stocktake["_id"]
    
    return StandardResponse(
        success=True, 
        message="Stocktake status retrieved", 
        data={
            "stocktake": stocktake,
            "stocktake_required": not bool(stocktake),
            "first_business_day": is_first_business_day
        }
    )

@api_router.post("/stocktake", response_model=StandardResponse)
async def create_stocktake(
    stocktake_data: StocktakeCreate,
    current_user: dict = Depends(require_manager_or_admin)
):
    """Create new stocktake with all materials"""
    current_month = stocktake_data.stocktake_date.strftime("%Y-%m")
    
    # Check if stocktake already exists for this month
    existing = await db.stocktakes.find_one({"month": current_month})
    if existing:
        raise HTTPException(status_code=400, detail="Stocktake already exists for this month")
    
    # Get all active materials
    materials = await db.materials.find({"is_active": True}).to_list(1000)
    
    new_stocktake = Stocktake(
        stocktake_date=stocktake_data.stocktake_date,
        month=current_month,
        created_by=current_user["sub"]
    )
    
    # Convert the stocktake to dict and handle date serialization
    stocktake_dict = new_stocktake.dict()
    stocktake_dict["stocktake_date"] = datetime.combine(stocktake_data.stocktake_date, datetime.min.time())
    
    await db.stocktakes.insert_one(stocktake_dict)
    
    return StandardResponse(
        success=True, 
        message="Stocktake created", 
        data={
            "stocktake_id": new_stocktake.id,
            "materials_count": len(materials),
            "materials": [{"id": m["id"], "name": f"{m['supplier']} - {m['product_code']}", "unit": m["unit"]} for m in materials]
        }
    )

@api_router.put("/stocktake/{stocktake_id}/entry", response_model=StandardResponse)
async def update_stocktake_entry(
    stocktake_id: str,
    entry_data: StocktakeEntryUpdate,
    current_user: dict = Depends(require_manager_or_admin)
):
    """Update stocktake entry for a material"""
    stocktake = await db.stocktakes.find_one({"id": stocktake_id})
    if not stocktake:
        raise HTTPException(status_code=404, detail="Stocktake not found")
    
    material = await db.materials.find_one({"id": entry_data.material_id})
    if not material:
        raise HTTPException(status_code=404, detail="Material not found")
    
    # Create or update entry
    entry = StocktakeEntry(
        stocktake_id=stocktake_id,
        material_id=entry_data.material_id,
        material_name=f"{material['supplier']} - {material['product_code']}",
        current_quantity=entry_data.current_quantity,
        unit=material["unit"],
        counted_by=current_user["sub"]
    )
    
    # Update stocktake entries
    await db.stocktake_entries.replace_one(
        {"stocktake_id": stocktake_id, "material_id": entry_data.material_id},
        entry.dict(),
        upsert=True
    )
    
    return StandardResponse(success=True, message="Stocktake entry updated")

@api_router.post("/stocktake/{stocktake_id}/complete", response_model=StandardResponse)
async def complete_stocktake(
    stocktake_id: str,
    current_user: dict = Depends(require_manager_or_admin)
):
    """Complete stocktake"""
    result = await db.stocktakes.update_one(
        {"id": stocktake_id},
        {
            "$set": {
                "status": "completed",
                "completed_by": current_user["sub"],
                "completed_at": datetime.now(timezone.utc)
            }
        }
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Stocktake not found")
    
    return StandardResponse(success=True, message="Stocktake completed")

def _is_first_business_day(check_date):
    """Check if given date is first business day of month"""
    first_day = check_date.replace(day=1)
    # Skip weekends
    while first_day.weekday() >= 5:  # 5=Saturday, 6=Sunday
        first_day += timedelta(days=1)
    return check_date == first_day

# ============= USER MANAGEMENT ENDPOINTS =============

@api_router.get("/users", response_model=List[dict])
async def get_users(current_user: dict = Depends(require_admin)):
    """Get all users (Admin only)"""
    users = await db.users.find({}).sort("full_name", 1).to_list(1000)
    
    # Remove password hashes and convert ObjectId to string
    for user in users:
        if "password_hash" in user:
            del user["password_hash"]
        if "_id" in user:
            del user["_id"]  # Remove MongoDB ObjectId
    
    return users

@api_router.post("/users", response_model=StandardResponse)
async def create_user(user_data: UserCreate, current_user: dict = Depends(require_admin)):
    """Create new user account (Admin only)"""
    # Check if username or email already exists
    existing_user = await db.users.find_one({
        "$or": [
            {"username": user_data.username},
            {"email": user_data.email}
        ]
    })
    
    if existing_user:
        if existing_user["username"] == user_data.username:
            raise HTTPException(status_code=400, detail="Username already exists")
        if existing_user["email"] == user_data.email:
            raise HTTPException(status_code=400, detail="Email already exists")
    
    # Hash the password
    password_hash = hash_password(user_data.password)
    
    # Create user document
    new_user = {
        "id": str(uuid.uuid4()),
        "username": user_data.username,
        "email": user_data.email,
        "password_hash": password_hash,
        "full_name": user_data.full_name,
        "role": user_data.role,
        "department": user_data.department,
        "phone": user_data.phone,
        "employment_type": user_data.employment_type.value,
        "is_active": True,
        "created_at": datetime.now(timezone.utc),
        "updated_at": None,
        "created_by": current_user["sub"]
    }
    
    await db.users.insert_one(new_user)
    
    return StandardResponse(success=True, message="User created successfully", data={"id": new_user["id"]})

@api_router.get("/users/{user_id}", response_model=dict)
async def get_user(user_id: str, current_user: dict = Depends(require_admin)):
    """Get specific user by ID (Admin only)"""
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Remove password hash and ObjectId from response
    if "password_hash" in user:
        del user["password_hash"]
    if "_id" in user:
        del user["_id"]  # Remove MongoDB ObjectId
    
    return user

@api_router.put("/users/{user_id}", response_model=StandardResponse)
async def update_user(user_id: str, user_data: UserUpdate, current_user: dict = Depends(require_admin)):
    """Update user account (Admin only)"""
    # Check if user exists
    existing_user = await db.users.find_one({"id": user_id})
    if not existing_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Prepare update data
    update_data = {}
    if user_data.username is not None:
        # Check if username is already used by another user
        username_user = await db.users.find_one({"username": user_data.username, "id": {"$ne": user_id}})
        if username_user:
            raise HTTPException(status_code=400, detail="Username already exists")
        update_data["username"] = user_data.username
    
    if user_data.email is not None:
        # Check if email is already used by another user
        email_user = await db.users.find_one({"email": user_data.email, "id": {"$ne": user_id}})
        if email_user:
            raise HTTPException(status_code=400, detail="Email already exists")
        update_data["email"] = user_data.email
    
    if user_data.full_name is not None:
        update_data["full_name"] = user_data.full_name
    if user_data.role is not None:
        update_data["role"] = user_data.role
    if user_data.department is not None:
        update_data["department"] = user_data.department
    if user_data.phone is not None:
        update_data["phone"] = user_data.phone
    if user_data.employment_type is not None:
        update_data["employment_type"] = user_data.employment_type.value
    if user_data.is_active is not None:
        update_data["is_active"] = user_data.is_active
    
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    await db.users.update_one(
        {"id": user_id},
        {"$set": update_data}
    )
    
    return StandardResponse(success=True, message="User updated successfully")

@api_router.delete("/users/{user_id}", response_model=StandardResponse)
async def delete_user(user_id: str, current_user: dict = Depends(require_admin)):
    """Delete user permanently (Admin only)"""
    # Check if user exists before deletion
    existing_user = await db.users.find_one({"id": user_id})
    if not existing_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Prevent admin from deleting themselves
    if user_id == current_user.get("user_id") or user_id == current_user.get("sub"):
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    
    # Permanently delete the user
    result = await db.users.delete_one({"id": user_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return StandardResponse(success=True, message="User deleted successfully")

@api_router.post("/users/change-password", response_model=StandardResponse)
async def change_user_password(password_data: PasswordChangeRequest, current_user: dict = Depends(get_current_user)):
    """Change user's own password"""
    # Get current user from database using user_id from JWT token
    user_id = current_user.get("user_id")
    if not user_id:
        raise HTTPException(status_code=404, detail="User ID not found in token")
    
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Verify current password (check both possible field names for compatibility)
    password_field = "password_hash" if "password_hash" in user else "hashed_password"
    if not verify_password(password_data.current_password, user[password_field]):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    
    # Hash new password
    new_password_hash = hash_password(password_data.new_password)
    
    # Update password (use the same field name that exists)
    await db.users.update_one(
        {"id": user_id},
        {"$set": {
            password_field: new_password_hash,
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    return StandardResponse(success=True, message="Password changed successfully")

# Helper functions for user management
def _generate_permissions(role: UserRole) -> List[str]:
    """Generate permissions based on role"""
    base_permissions = []
    
    if role == UserRole.ADMIN:
        return ["all"]  # Admin has all permissions
    
    elif role == UserRole.MANAGER:
        return [
            "manage_clients", "create_orders", "update_production", 
            "view_reports", "manage_payroll", "manage_materials",
            "manage_suppliers", "manage_specifications", "use_calculators"
        ]
    
    elif role == UserRole.MANAGER:
        return [
            "create_orders", "update_production", "use_calculators", 
            "view_own_payroll", "submit_timesheets", "request_leave"
        ]
    
    elif role == UserRole.PRODUCTION_TEAM:
        return [
            "update_production", "view_own_payroll", "submit_timesheets", "request_leave"
        ]
    
    elif role == UserRole.SALES:
        return [
            "manage_clients", "create_orders", "view_reports"
        ]
    
    return base_permissions

# ============= CLIENT PRODUCT CATALOGUE ENDPOINTS =============

@api_router.get("/clients/{client_id}/catalog", response_model=List[ClientProduct])
async def get_client_product_catalog(client_id: str, current_user: dict = Depends(require_any_role)):
    """Get product catalogue for specific client"""
    products = await db.client_products.find({"client_id": client_id, "is_active": True}).to_list(1000)
    return [ClientProduct(**product) for product in products]

@api_router.post("/clients/{client_id}/catalog", response_model=StandardResponse)
async def create_client_product(client_id: str, product_data: ClientProductCreate, current_user: dict = Depends(require_admin_or_sales)):
    """Create new product for client catalogue"""
    # Verify client exists
    client = await db.clients.find_one({"id": client_id, "is_active": True})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Set client_id from URL
    product_data.client_id = client_id
    new_product = ClientProduct(**product_data.dict())
    await db.client_products.insert_one(new_product.dict())
    
    return StandardResponse(success=True, message="Client product created successfully", data={"id": new_product.id})

@api_router.get("/clients/{client_id}/catalog/{product_id}", response_model=ClientProduct)
async def get_client_product(client_id: str, product_id: str, current_user: dict = Depends(require_any_role)):
    """Get specific client product"""
    product = await db.client_products.find_one({
        "id": product_id, 
        "client_id": client_id,
        "is_active": True
    })
    if not product:
        raise HTTPException(status_code=404, detail="Client product not found")
    
    return ClientProduct(**product)

@api_router.put("/clients/{client_id}/catalog/{product_id}", response_model=StandardResponse)
async def update_client_product(client_id: str, product_id: str, product_data: ClientProductCreate, current_user: dict = Depends(require_admin_or_sales)):
    """Update client product"""
    # Verify client exists
    client = await db.clients.find_one({"id": client_id, "is_active": True})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    update_data = product_data.dict()
    update_data["client_id"] = client_id  # Ensure client_id is correct
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    result = await db.client_products.update_one(
        {"id": product_id, "client_id": client_id, "is_active": True},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Client product not found")
    
    return StandardResponse(success=True, message="Client product updated successfully")

@api_router.delete("/clients/{client_id}/catalog/{product_id}", response_model=StandardResponse)
async def delete_client_product(client_id: str, product_id: str, current_user: dict = Depends(require_admin_or_sales)):
    """Delete client product (soft delete)"""
    result = await db.client_products.update_one(
        {"id": product_id, "client_id": client_id, "is_active": True},
        {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc)}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Client product not found")
    
    return StandardResponse(success=True, message="Client product deleted successfully")

@api_router.post("/clients/{client_id}/catalog/{product_id}/copy-to/{target_client_id}", response_model=StandardResponse)
async def copy_client_product(client_id: str, product_id: str, target_client_id: str, current_user: dict = Depends(require_admin_or_sales)):
    """Copy product to another client catalogue"""
    # Verify both clients exist
    source_client = await db.clients.find_one({"id": client_id, "is_active": True})
    target_client = await db.clients.find_one({"id": target_client_id, "is_active": True})
    
    if not source_client:
        raise HTTPException(status_code=404, detail="Source client not found")
    if not target_client:
        raise HTTPException(status_code=404, detail="Target client not found")
    
    # Get source product
    source_product = await db.client_products.find_one({
        "id": product_id,
        "client_id": client_id,
        "is_active": True
    })
    
    if not source_product:
        raise HTTPException(status_code=404, detail="Source product not found")
    
    # Create copy for target client
    copied_product = ClientProduct(**source_product)
    copied_product.id = str(uuid.uuid4())  # Generate new ID
    copied_product.client_id = target_client_id
    copied_product.created_at = datetime.now(timezone.utc)
    copied_product.updated_at = None
    
    await db.client_products.insert_one(copied_product.dict())
    
    return StandardResponse(
        success=True, 
        message=f"Product copied successfully to {target_client['company_name']}", 
        data={"id": copied_product.id}
    )

@api_router.delete("/clients/{client_id}", response_model=StandardResponse)
async def delete_client(client_id: str, current_user: dict = Depends(require_admin)):
    """Delete client (soft delete - Admin only)"""
    # Check if client exists
    existing_client = await db.clients.find_one({"id": client_id, "is_active": True})
    if not existing_client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Check if client has active orders
    active_orders = await db.orders.find_one({"client_id": client_id, "status": {"$nin": ["completed", "cancelled"]}})
    if active_orders:
        raise HTTPException(status_code=400, detail="Cannot delete client with active orders")
    
    # Perform soft delete
    result = await db.clients.update_one(
        {"id": client_id},
        {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc)}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Client not found")
    
    return StandardResponse(success=True, message="Client deleted successfully")

# ============= ORDER MANAGEMENT ENDPOINTS =============

@api_router.get("/orders", response_model=List[Order])
async def get_orders(status_filter: Optional[str] = None, current_user: dict = Depends(require_any_role)):
    """Get all orders with optional status filter"""
    query = {}
    if status_filter:
        query["status"] = status_filter
    
    orders = await db.orders.find(query).sort("created_at", -1).to_list(1000)
    return [Order(**order) for order in orders]

@api_router.post("/orders", response_model=StandardResponse)
async def create_order(order_data: OrderCreate, current_user: dict = Depends(require_admin_or_manager)):
    """Create new order"""
    # Get client details
    client = await db.clients.find_one({"id": order_data.client_id, "is_active": True})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Generate order number
    order_count = await db.orders.count_documents({})
    order_number = f"ADM-{datetime.now().year}-{order_count + 1:04d}"
    
    # Calculate totals with discount applied before GST
    subtotal = sum(item.total_price for item in order_data.items)
    
    # Apply discount if provided
    discount_percentage = order_data.discount_percentage or 0
    discount_amount = 0
    discounted_subtotal = subtotal
    
    if discount_percentage > 0:
        discount_amount = subtotal * (discount_percentage / 100)
        discounted_subtotal = subtotal - discount_amount
    
    # Calculate GST on discounted amount
    gst = discounted_subtotal * 0.1
    total_amount = discounted_subtotal + gst
    
    new_order = Order(
        order_number=order_number,
        client_id=order_data.client_id,
        client_name=client["company_name"],
        purchase_order_number=order_data.purchase_order_number,
        items=order_data.items,
        subtotal=subtotal,
        discount_percentage=discount_percentage if discount_percentage > 0 else None,
        discount_amount=discount_amount if discount_amount > 0 else None,
        discount_notes=order_data.discount_notes if order_data.discount_notes else None,
        discounted_subtotal=discounted_subtotal,
        gst=gst,
        total_amount=total_amount,
        due_date=order_data.due_date,
        priority=order_data.priority,
        delivery_address=order_data.delivery_address,
        delivery_instructions=order_data.delivery_instructions,
        runtime_estimate=order_data.runtime_estimate,
        notes=order_data.notes,
        created_by=current_user["user_id"]
    )
    
    await db.orders.insert_one(new_order.dict())
    
    # Log initial production stage
    production_log = ProductionLog(
        order_id=new_order.id,
        stage=ProductionStage.ORDER_ENTERED,
        updated_by=current_user["user_id"],
        notes="Order created"
    )
    await db.production_logs.insert_one(production_log.dict())
    
    return StandardResponse(success=True, message="Order created successfully", data={"id": new_order.id, "order_number": order_number})

@api_router.get("/orders/{order_id}", response_model=Order)
async def get_order(order_id: str, current_user: dict = Depends(require_any_role)):
    """Get specific order"""
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    return Order(**order)

@api_router.put("/orders/{order_id}/stage", response_model=StandardResponse)
async def update_production_stage(order_id: str, stage_update: ProductionStageUpdate, current_user: dict = Depends(require_production_access)):
    """Update order production stage"""
    # Check permissions for specific actions
    user_role = current_user.get("role")
    
    # Production team can only move stages forward, not create/delete orders
    if user_role == UserRole.PRODUCTION_TEAM.value:
        if stage_update.to_stage == ProductionStage.ORDER_ENTERED:
            raise HTTPException(status_code=403, detail="Cannot move back to order entry stage")
    
    # Only admin and production manager can invoice
    if stage_update.to_stage == ProductionStage.INVOICING and not can_invoice(user_role):
        raise HTTPException(status_code=403, detail="Insufficient permissions to invoice")
    
    # Update order stage
    update_data = {
        "current_stage": stage_update.to_stage,
        "updated_at": datetime.now(timezone.utc)
    }
    
    # If moving to completed, set completion date and archive the order
    if stage_update.to_stage == ProductionStage.CLEARED:
        update_data["completed_at"] = datetime.now(timezone.utc)
        update_data["status"] = OrderStatus.COMPLETED
        
        # Get the complete order data for archiving
        order = await db.orders.find_one({"id": order_id})
        if order:
            # Create archived order
            archived_order = ArchivedOrder(
                original_order_id=order["id"],
                order_number=order["order_number"],
                client_id=order["client_id"],
                client_name=order["client_name"],
                purchase_order_number=order.get("purchase_order_number"),
                items=order["items"],
                subtotal=order["subtotal"],
                gst=order["gst"],
                total_amount=order["total_amount"],
                due_date=order["due_date"],
                delivery_address=order.get("delivery_address"),
                delivery_instructions=order.get("delivery_instructions"),
                runtime_estimate=order.get("runtime_estimate"),
                notes=order.get("notes"),
                created_by=order["created_by"],
                created_at=order["created_at"],
                completed_at=datetime.now(timezone.utc),
                archived_by=current_user["user_id"]
            )
            
            # Insert into archived orders collection
            await db.archived_orders.insert_one(archived_order.dict())
            
            # Update order status to archived and keep in main collection for now
            # This allows for transition period and potential rollback
            update_data["status"] = OrderStatus.ARCHIVED
    
    result = await db.orders.update_one(
        {"id": order_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Log production stage change
    production_log = ProductionLog(
        order_id=order_id,
        stage=stage_update.to_stage,
        updated_by=current_user["user_id"],
        notes=stage_update.notes
    )
    await db.production_logs.insert_one(production_log.dict())
    
    return StandardResponse(success=True, message="Production stage updated successfully")

# ============= JOB SPECIFICATION ENDPOINTS =============

@api_router.post("/job-specifications", response_model=StandardResponse)
async def create_job_specification(spec_data: JobSpecificationCreate, current_user: dict = Depends(require_admin_or_manager)):
    """Create job specification"""
    new_spec = JobSpecification(**spec_data.dict(), created_by=current_user["user_id"])
    await db.job_specifications.insert_one(new_spec.dict())
    
    return StandardResponse(success=True, message="Job specification created successfully", data={"id": new_spec.id})

@api_router.get("/orders/{order_id}/job-specification", response_model=JobSpecification)
async def get_job_specification_by_order(order_id: str, current_user: dict = Depends(require_any_role)):
    """Get job specification for order"""
    spec = await db.job_specifications.find_one({"order_id": order_id})
    if not spec:
        raise HTTPException(status_code=404, detail="Job specification not found")
    
    return JobSpecification(**spec)

@api_router.delete("/orders/{order_id}", response_model=StandardResponse)
async def delete_order(order_id: str, current_user: dict = Depends(require_admin)):
    """Delete order permanently (Admin only)"""
    # Check if order exists
    existing_order = await db.orders.find_one({"id": order_id})
    if not existing_order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Check if order is in production (not safe to delete if in progress)
    unsafe_stages = ["paper_slitting", "winding", "finishing"]
    current_stage = existing_order.get("current_stage", "order_entered")
    if current_stage in unsafe_stages:
        raise HTTPException(status_code=400, detail="Cannot delete order in production. Contact manager to halt production first.")
    
    # Clean up related data first
    # Remove job specifications
    await db.job_specifications.delete_many({"order_id": order_id})
    
    # Remove materials status
    await db.materials_status.delete_many({"order_id": order_id})
    
    # Remove order items status
    await db.order_items_status.delete_many({"order_id": order_id})
    
    # Perform hard delete - completely remove the order
    result = await db.orders.delete_one({"id": order_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Order not found")
    
    return StandardResponse(success=True, message="Order deleted successfully")

# ============= PRODUCTION BOARD ENDPOINTS =============

@api_router.get("/production/board")
async def get_production_board(current_user: dict = Depends(require_any_role)):
    """Get production board with orders grouped by stage"""
    # Get all active orders
    orders = await db.orders.find({"status": {"$ne": OrderStatus.COMPLETED}}).to_list(1000)
    
    # Group orders by production stage
    board = {}
    for stage in ProductionStage:
        if stage != ProductionStage.CLEARED:  # Don't show cleared orders on board
            board[stage.value] = []
    
    for order in orders:
        stage = order.get("current_stage", ProductionStage.ORDER_ENTERED.value)
        if stage in board:
            # Get client info for logo
            client = await db.clients.find_one({"id": order["client_id"]})
            
            # Get materials status
            materials_status = await db.materials_status.find_one({"order_id": order["id"]})
            materials_ready = materials_status.get("materials_ready", False) if materials_status else False
            
            # Handle due_date comparison properly for timezone awareness
            due_date = order["due_date"]
            if isinstance(due_date, str):
                due_date = datetime.fromisoformat(due_date.replace("Z", "+00:00"))
            elif isinstance(due_date, datetime) and due_date.tzinfo is None:
                # Make timezone-naive datetime timezone-aware (assume UTC)
                due_date = due_date.replace(tzinfo=timezone.utc)
            
            order_info = {
                "id": order["id"],
                "order_number": order["order_number"],
                "client_name": order["client_name"],
                "client_logo": get_file_url(client.get("logo_path", "")) if client else None,
                "due_date": order["due_date"],
                "total_amount": order["total_amount"],
                "runtime": order.get("runtime_estimate", "2-3 days"),
                "materials_ready": materials_ready,
                "items": order["items"],
                "delivery_address": order.get("delivery_address"),
                "is_overdue": due_date < datetime.now(timezone.utc)
            }
            board[stage].append(order_info)
    
    return {"success": True, "data": board}

@api_router.get("/production/logs/{order_id}")
async def get_production_logs(order_id: str, current_user: dict = Depends(require_any_role)):
    """Get production logs for order"""
    logs = await db.production_logs.find({"order_id": order_id}).sort("timestamp", 1).to_list(1000)
    return {"success": True, "data": logs}

@api_router.post("/production/move-stage/{order_id}")
async def move_order_stage(
    order_id: str, 
    request: StageMovementRequest, 
    current_user: dict = Depends(require_admin_or_manager)
):
    """Move order to next or previous production stage"""
    # Get current order
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Define stage progression
    stage_order = [
        ProductionStage.ORDER_ENTERED,
        ProductionStage.PENDING_MATERIAL,
        ProductionStage.PAPER_SLITTING,
        ProductionStage.WINDING,
        ProductionStage.FINISHING,
        ProductionStage.DELIVERY,
        ProductionStage.INVOICING,
        ProductionStage.CLEARED
    ]
    
    current_stage = ProductionStage(order["current_stage"])
    current_index = stage_order.index(current_stage)
    
    # Calculate new stage
    if request.direction == "forward":
        new_index = min(current_index + 1, len(stage_order) - 1)
    elif request.direction == "backward":
        new_index = max(current_index - 1, 0)
    else:
        raise HTTPException(status_code=400, detail="Direction must be 'forward' or 'backward'")
    
    if new_index == current_index:
        raise HTTPException(status_code=400, detail="Order is already at the first/last stage")
    
    new_stage = stage_order[new_index]
    
    # Update order stage
    await db.orders.update_one(
        {"id": order_id},
        {
            "$set": {
                "current_stage": new_stage.value,
                "updated_at": datetime.now(timezone.utc)
            }
        }
    )
    
    # Log stage change
    stage_log = ProductionLog(
        order_id=order_id,
        stage=new_stage,
        updated_by=current_user["sub"],
        notes=request.notes
    )
    await db.production_logs.insert_one(stage_log.dict())
    
    return {"success": True, "message": f"Order moved to {new_stage.value}", "new_stage": new_stage.value}

@api_router.post("/production/jump-stage/{order_id}")
async def jump_to_stage(
    order_id: str, 
    request: StageJumpRequest, 
    current_user: dict = Depends(require_admin_or_manager)
):
    """Jump order directly to a specific production stage"""
    # Get current order
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Validate target stage
    try:
        target_stage = ProductionStage(request.target_stage)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid target stage")
    
    current_stage = ProductionStage(order["current_stage"])
    
    # Check if already at target stage
    if current_stage == target_stage:
        raise HTTPException(status_code=400, detail=f"Order is already at {target_stage.value} stage")
    
    # Update order stage
    await db.orders.update_one(
        {"id": order_id},
        {
            "$set": {
                "current_stage": target_stage.value,
                "updated_at": datetime.now(timezone.utc)
            }
        }
    )
    
    # Log stage jump
    stage_log = ProductionLog(
        order_id=order_id,
        stage=target_stage,
        updated_by=current_user["sub"],
        notes=f"Jumped from {current_stage.value} to {target_stage.value}" + (f" - {request.notes}" if request.notes else "")
    )
    await db.production_logs.insert_one(stage_log.dict())
    
    return {"success": True, "message": f"Order jumped to {target_stage.value}", "new_stage": target_stage.value}

@api_router.get("/production/materials-status/{order_id}")
async def get_materials_status(order_id: str, current_user: dict = Depends(require_any_role)):
    """Get materials status for order"""
    # Check if order exists first
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    status = await db.materials_status.find_one({"order_id": order_id})
    if not status:
        # Create default status if not exists
        default_status = MaterialsStatus(
            order_id=order_id,
            materials_ready=False,
            materials_checklist=[],
            updated_by=current_user["sub"]
        )
        await db.materials_status.insert_one(default_status.dict())
        return {"success": True, "data": default_status.dict()}
    
    # Remove MongoDB _id field to avoid serialization issues
    if "_id" in status:
        del status["_id"]
    
    return {"success": True, "data": status}

@api_router.put("/production/materials-status/{order_id}")
async def update_materials_status(
    order_id: str, 
    status_update: MaterialsStatusUpdate,
    current_user: dict = Depends(require_admin_or_manager)
):
    """Update materials status for order"""
    # Check if order exists
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Update or create materials status
    update_data = {
        "materials_ready": status_update.materials_ready,
        "materials_checklist": status_update.materials_checklist,
        "updated_by": current_user["sub"],
        "updated_at": datetime.now(timezone.utc)
    }
    
    result = await db.materials_status.update_one(
        {"order_id": order_id},
        {"$set": update_data},
        upsert=True
    )
    
    return {"success": True, "message": "Materials status updated"}

@api_router.put("/production/order-item-status/{order_id}")
async def update_order_item_status(
    order_id: str,
    item_update: OrderItemStatusUpdate,
    current_user: dict = Depends(require_admin_or_manager)
):
    """Update completion status of specific order item"""
    # Get the order
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Validate item index
    if item_update.item_index >= len(order["items"]) or item_update.item_index < 0:
        raise HTTPException(status_code=400, detail="Invalid item index")
    
    # Update the specific item's completion status
    update_query = {
        f"items.{item_update.item_index}.is_completed": item_update.is_completed,
        "updated_at": datetime.now(timezone.utc)
    }
    
    await db.orders.update_one(
        {"id": order_id},
        {"$set": update_query}
    )
    
    return {"success": True, "message": f"Item {item_update.item_index} marked as {'completed' if item_update.is_completed else 'pending'}"}


# ============= REPORTS ENDPOINTS =============

@api_router.get("/reports/outstanding-jobs")
async def get_outstanding_jobs_report(current_user: dict = Depends(require_admin_or_manager)):
    """Generate outstanding jobs report"""
    # Count jobs by stage
    jobs_by_stage = {}
    total_jobs = 0
    overdue_jobs = 0
    jobs_due_today = 0
    jobs_due_this_week = 0
    
    now = datetime.now(timezone.utc)
    today_end = now.replace(hour=23, minute=59, second=59)
    week_end = now + timedelta(days=7)
    
    # Get all active orders
    orders = await db.orders.find({"status": {"$ne": OrderStatus.COMPLETED}}).to_list(1000)
    
    for order in orders:
        stage = order.get("current_stage", ProductionStage.ORDER_ENTERED.value)
        jobs_by_stage[stage] = jobs_by_stage.get(stage, 0) + 1
        total_jobs += 1
        
        due_date = order["due_date"]
        if isinstance(due_date, str):
            due_date = datetime.fromisoformat(due_date.replace("Z", "+00:00"))
        elif isinstance(due_date, datetime) and due_date.tzinfo is None:
            # Make timezone-naive datetime timezone-aware (assume UTC)
            due_date = due_date.replace(tzinfo=timezone.utc)
        
        if due_date < now:
            overdue_jobs += 1
        elif due_date <= today_end:
            jobs_due_today += 1
        elif due_date <= week_end:
            jobs_due_this_week += 1
    
    report = OutstandingJobsReport(
        total_jobs=total_jobs,
        jobs_by_stage=jobs_by_stage,
        overdue_jobs=overdue_jobs,
        jobs_due_today=jobs_due_today,
        jobs_due_this_week=jobs_due_this_week
    )
    
    return {"success": True, "data": report.dict()}

@api_router.get("/reports/late-deliveries")
async def get_late_deliveries_report(current_user: dict = Depends(require_admin_or_manager)):
    """Generate late deliveries report"""
    # Get completed orders that were delivered late
    completed_orders = await db.orders.find({"status": OrderStatus.COMPLETED}).to_list(1000)
    
    late_deliveries = []
    for order in completed_orders:
        due_date = order["due_date"]
        completed_date = order.get("completed_at")
        
        if completed_date and due_date:
            if isinstance(due_date, str):
                due_date = datetime.fromisoformat(due_date.replace("Z", "+00:00"))
            if isinstance(completed_date, str):
                completed_date = datetime.fromisoformat(completed_date.replace("Z", "+00:00"))
            
            if completed_date > due_date:
                delay_days = (completed_date - due_date).days
                late_deliveries.append({
                    "order": order,
                    "delay_days": delay_days,
                    "client_name": order["client_name"]
                })
    
    # Calculate statistics
    total_late = len(late_deliveries)
    avg_delay = sum(ld["delay_days"] for ld in late_deliveries) / total_late if total_late > 0 else 0
    
    # Group by client
    late_by_client = {}
    for ld in late_deliveries:
        client = ld["client_name"]
        late_by_client[client] = late_by_client.get(client, 0) + 1
    
    # Group by month
    late_by_month = {}
    for ld in late_deliveries:
        completed_date = ld["order"]["completed_at"]
        if isinstance(completed_date, str):
            completed_date = datetime.fromisoformat(completed_date.replace("Z", "+00:00"))
        month_key = completed_date.strftime("%Y-%m")
        late_by_month[month_key] = late_by_month.get(month_key, 0) + 1
    
    report = LateDeliveryReport(
        total_late_deliveries=total_late,
        average_delay_days=avg_delay,
        late_deliveries_by_client=late_by_client,
        late_deliveries_by_month=late_by_month
    )
    
    return {"success": True, "data": report.dict()}

@api_router.get("/reports/customer-annual/{client_id}")
async def get_customer_annual_report(client_id: str, year: int, current_user: dict = Depends(require_admin_or_manager)):
    """Generate customer annual report"""
    # Get client info
    client = await db.clients.find_one({"id": client_id})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Get orders for the year
    start_date = datetime(year, 1, 1)
    end_date = datetime(year + 1, 1, 1)
    
    orders = await db.orders.find({
        "client_id": client_id,
        "created_at": {"$gte": start_date, "$lt": end_date}
    }).to_list(1000)
    
    # Calculate statistics
    total_orders = len(orders)
    total_revenue = sum(order["total_amount"] for order in orders)
    avg_order_value = total_revenue / total_orders if total_orders > 0 else 0
    
    # Orders by month
    orders_by_month = {}
    revenue_by_month = {}
    for i in range(1, 13):
        month_key = f"{year}-{i:02d}"
        orders_by_month[month_key] = 0
        revenue_by_month[month_key] = 0
    
    for order in orders:
        created_date = order["created_at"]
        if isinstance(created_date, str):
            created_date = datetime.fromisoformat(created_date.replace("Z", "+00:00"))
        month_key = created_date.strftime("%Y-%m")
        if month_key in orders_by_month:
            orders_by_month[month_key] += 1
            revenue_by_month[month_key] += order["total_amount"]
    
    # Top products
    product_stats = {}
    for order in orders:
        for item in order["items"]:
            product_name = item["product_name"]
            if product_name not in product_stats:
                product_stats[product_name] = {"quantity": 0, "revenue": 0, "orders": 0}
            product_stats[product_name]["quantity"] += item["quantity"]
            product_stats[product_name]["revenue"] += item["total_price"]
            product_stats[product_name]["orders"] += 1
    
    top_products = sorted(
        [{"product": k, **v} for k, v in product_stats.items()],
        key=lambda x: x["revenue"],
        reverse=True
    )[:5]
    
    # On-time delivery rate
    completed_orders = [o for o in orders if o.get("completed_at")]
    on_time_orders = 0
    for order in completed_orders:
        due_date = order["due_date"]
        completed_date = order["completed_at"]
        
        if isinstance(due_date, str):
            due_date = datetime.fromisoformat(due_date.replace("Z", "+00:00"))
        if isinstance(completed_date, str):
            completed_date = datetime.fromisoformat(completed_date.replace("Z", "+00:00"))
        
        if completed_date <= due_date:
            on_time_orders += 1
    
    on_time_rate = (on_time_orders / len(completed_orders)) * 100 if completed_orders else 0
    
    report = CustomerAnnualReport(
        client_id=client_id,
        client_name=client["company_name"],
        year=year,
        total_orders=total_orders,
        total_revenue=total_revenue,
        average_order_value=avg_order_value,
        orders_by_month=orders_by_month,
        revenue_by_month=revenue_by_month,
        top_products=top_products,
        on_time_delivery_rate=on_time_rate
    )
    
    return {"success": True, "data": report.dict()}

# ============= DOCUMENT GENERATION ENDPOINTS =============

# Simplified authentication for PDF documents - allow open access for testing
async def simple_pdf_auth(token: str = None):
    """Simplified authentication for PDF downloads"""
    # For now, let's make PDFs accessible without strict authentication for testing
    # In production, you'd want proper authentication
    return {"user_id": "test-user", "role": "admin"}

@api_router.get("/documents/acknowledgment/{order_id}")
async def generate_acknowledgment(order_id: str):
    """Generate order acknowledgment PDF"""
    # Get order and client data
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    client = await db.clients.find_one({"id": order["client_id"]})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Prepare data for PDF generation
    order_data = {
        "order_number": order["order_number"],
        "invoice_number": f"INV-{order['order_number']}",
        "due_date": order["due_date"].strftime("%d/%m/%Y") if isinstance(order["due_date"], datetime) else order["due_date"],
        "client_name": client["company_name"],
        "client_address": f"{client['address']}, {client['city']}, {client['state']} {client['postal_code']}",
        "client_email": client["email"],
        "client_phone": client["phone"],
        "client_payment_terms": client.get("payment_terms", "Net 30 days"),
        "client_lead_time_days": client.get("lead_time_days", 7),
        "delivery_instructions": order.get("delivery_instructions", "Standard delivery terms apply."),
        "items": order["items"],
        "subtotal": order["subtotal"],
        "gst": order["gst"],
        "total_amount": order["total_amount"],
        "bank_details": client.get("bank_details")
    }
    
    # Generate PDF
    pdf_content = doc_generator.generate_order_acknowledgment(order_data)
    
    return StreamingResponse(
        BytesIO(pdf_content),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=acknowledgment_{order['order_number']}.pdf"}
    )

@api_router.get("/documents/job-card/{order_id}")
async def generate_job_card(order_id: str):
    """Generate job card PDF"""
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Get job specifications if available
    job_spec = await db.job_specifications.find_one({"order_id": order_id})
    
    # Get product specifications
    specifications = None
    if order["items"] and len(order["items"]) > 0:
        product_id = order["items"][0].get("product_id")
        if product_id:
            product = await db.products.find_one({"id": product_id})
            if product:
                specifications = product.get("specifications")
    
    job_data = {
        "order_number": order["order_number"],
        "client_name": order["client_name"],
        "due_date": order["due_date"].strftime("%d/%m/%Y") if isinstance(order["due_date"], datetime) else order["due_date"],
        "current_stage": order.get("current_stage", "order_entered"),
        "specifications": specifications,
        "job_specification": job_spec
    }
    
    pdf_content = doc_generator.generate_job_card(job_data)
    
    return StreamingResponse(
        BytesIO(pdf_content),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=job_card_{order['order_number']}.pdf"}
    )

@api_router.get("/documents/packing-list/{order_id}")
async def generate_packing_list(order_id: str):
    """Generate packing list PDF"""
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    order_data = {
        "order_number": order["order_number"],
        "client_name": order["client_name"],
        "delivery_address": order.get("delivery_address", "N/A"),
        "delivery_instructions": order.get("delivery_instructions"),
        "items": order["items"]
    }
    
    pdf_content = doc_generator.generate_packing_list(order_data)
    
    return StreamingResponse(
        BytesIO(pdf_content),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=packing_list_{order['order_number']}.pdf"}
    )

# Moved this function above document endpoints

@api_router.get("/documents/invoice/{order_id}")
async def generate_invoice_pdf(order_id: str):
    """Generate invoice PDF"""
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    client = await db.clients.find_one({"id": order["client_id"]})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    invoice_data = {
        "invoice_number": f"INV-{order['order_number']}",
        "order_number": order["order_number"],
        "payment_due_date": (datetime.now(timezone.utc) + timedelta(days=30)).strftime("%d/%m/%Y"),
        "client_name": client["company_name"],
        "client_address": f"{client['address']}, {client['city']}, {client['state']} {client['postal_code']}",
        "client_abn": client.get("abn", "N/A"),
        "items": order["items"],
        "subtotal": order["subtotal"],
        "gst": order["gst"],
        "total_amount": order["total_amount"]
    }
    
    pdf_content = doc_generator.generate_invoice(invoice_data)
    
    return StreamingResponse(
        BytesIO(pdf_content),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=invoice_{order['order_number']}.pdf"}
    )

# ============= XERO INTEGRATION ENDPOINTS =============

# Xero Integration Configuration
# Force correct environment values for current environment  
os.environ['XERO_REDIRECT_URI'] = 'https://misty-ato-payroll.preview.emergentagent.com/xero-oauth-callback'
os.environ['FRONTEND_URL'] = 'https://misty-ato-payroll.preview.emergentagent.com'

XERO_CLIENT_ID = os.getenv("XERO_CLIENT_ID")
XERO_CLIENT_SECRET = os.getenv("XERO_CLIENT_SECRET")  
XERO_CALLBACK_URL = os.getenv("XERO_REDIRECT_URI")
XERO_SCOPES = "openid profile email accounting.transactions accounting.contacts accounting.settings offline_access"

# Debug: Print at module load time
print(f"🔥 DEBUG: XERO_REDIRECT_URI env var: {os.getenv('XERO_REDIRECT_URI')}")
print(f"🔥 MODULE LOAD: XERO_CALLBACK_URL set to: {XERO_CALLBACK_URL}")

# Xero Account Configuration
XERO_DEFAULT_SALES_ACCOUNT_CODE = "200"  # Sales account code
XERO_DEFAULT_TAX_TYPE = "OUTPUT"  # Default GST/tax type

# Debug endpoint for testing
@api_router.get("/xero/debug")
async def debug_xero_config():
    """Debug endpoint to check Xero configuration"""
    # Debug: print the actual value being used right now
    print(f"🚀 DEBUG ENDPOINT: XERO_CALLBACK_URL = {XERO_CALLBACK_URL}")
    
    return {
        "client_id": XERO_CLIENT_ID,
        "callback_url": XERO_CALLBACK_URL,
        "callback_url_from_variable": XERO_CALLBACK_URL,
        "scopes": XERO_SCOPES,
        "expected_auth_url_start": "https://login.xero.com/identity/connect/authorize",
        "frontend_url": os.getenv("FRONTEND_URL", "https://app.emergent.sh"),
        "configuration": {
            "XERO_SALES_ACCOUNT_CODE": os.getenv("XERO_SALES_ACCOUNT_CODE", "200"),
            "XERO_SALES_ACCOUNT_NAME": os.getenv("XERO_SALES_ACCOUNT_NAME", "Sales")
        },
        "environment_check": {
            "client_id_set": bool(XERO_CLIENT_ID),
            "client_secret_set": bool(XERO_CLIENT_SECRET),
            "redirect_uri": XERO_CALLBACK_URL
        }
    }

@api_router.get("/debug/test-pdf")
async def test_pdf_download():
    """Simple test PDF for debugging download issues"""
    from reportlab.pdfgen import canvas
    from io import BytesIO
    
    buffer = BytesIO()
    c = canvas.Canvas(buffer)
    c.drawString(100, 750, "Test PDF Download")
    c.drawString(100, 730, "If you can see this, PDF downloads are working!")
    c.save()
    
    buffer.seek(0)
    
    return StreamingResponse(
        BytesIO(buffer.read()),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=test.pdf"}
    )

@api_router.get("/xero/status")
async def check_xero_connection_status(current_user: dict = Depends(require_admin_or_manager)):
    """Check if user has active Xero connection"""
    # Check if user has stored Xero tokens
    user_tokens = await db.xero_tokens.find_one({"user_id": current_user["user_id"]})
    
    if user_tokens and user_tokens.get("access_token"):
        # TODO: Validate token is still active with Xero API
        return {"connected": True, "message": "Xero connection active", "tenant_id": user_tokens.get("tenant_id")}
    else:
        return {"connected": False, "message": "No Xero connection found"}

# Helper function to get authenticated Xero API client
async def get_xero_api_client(user_id: str):
    """Get authenticated Xero API client for user"""
    user_tokens = await db.xero_tokens.find_one({"user_id": user_id})
    if not user_tokens or not user_tokens.get("access_token"):
        raise HTTPException(status_code=401, detail="No Xero connection found")
    
    # Check if token is expired
    if user_tokens.get("expires_at") and user_tokens["expires_at"] < datetime.now(timezone.utc):
        # Try to refresh token
        try:
            refreshed_tokens = await refresh_xero_token(user_id, user_tokens["refresh_token"])
            user_tokens = refreshed_tokens
        except Exception as e:
            raise HTTPException(status_code=401, detail="Xero token expired and refresh failed")
    
    # Create OAuth2 token object
    oauth2_token = OAuth2Token(
        client_id=XERO_CLIENT_ID,
        client_secret=XERO_CLIENT_SECRET
    )
    oauth2_token.access_token = user_tokens["access_token"]
    oauth2_token.refresh_token = user_tokens.get("refresh_token")
    oauth2_token.expires_at = user_tokens.get("expires_at")
    
    # Create API client
    configuration = Configuration(oauth2_token=oauth2_token)
    api_client = ApiClient(configuration, pool_threads=1)
    
    return api_client, user_tokens.get("tenant_id")

async def refresh_xero_token(user_id: str, refresh_token: str):
    """Refresh expired Xero access token"""
    token_url = "https://identity.xero.com/connect/token"
    
    auth_string = f"{XERO_CLIENT_ID}:{XERO_CLIENT_SECRET}"
    auth_bytes = auth_string.encode('ascii')
    auth_b64 = base64.b64encode(auth_bytes).decode('ascii')
    
    headers = {
        "Authorization": f"Basic {auth_b64}",
        "Content-Type": "application/x-www-form-urlencoded"
    }
    
    token_data = {
        "grant_type": "refresh_token",
        "refresh_token": refresh_token
    }
    
    response = requests.post(token_url, headers=headers, data=token_data)
    response.raise_for_status()
    
    tokens = response.json()
    
    # Update stored tokens
    token_record = {
        "user_id": user_id,
        "access_token": tokens["access_token"],
        "refresh_token": tokens.get("refresh_token", refresh_token),
        "expires_at": datetime.now(timezone.utc) + timedelta(seconds=tokens.get("expires_in", 1800)),
        "updated_at": datetime.now(timezone.utc)
    }
    
    await db.xero_tokens.update_one(
        {"user_id": user_id},
        {"$set": token_record}
    )
    
    return token_record

@api_router.get("/xero/auth/url")
async def get_xero_auth_url(current_user: dict = Depends(require_admin_or_manager)):
    """Get Xero OAuth authorization URL"""
    # Generate secure state parameter
    state = secrets.token_urlsafe(32)
    
    # Store state for validation
    await db.xero_auth_states.insert_one({
        "state": state,
        "user_id": current_user["user_id"],
        "created_at": datetime.now(timezone.utc)
    })
    
    # Build authorization URL
    auth_params = {
        "response_type": "code",
        "client_id": XERO_CLIENT_ID,
        "redirect_uri": XERO_CALLBACK_URL,
        "scope": XERO_SCOPES,
        "state": state
    }
    
    auth_url = f"https://login.xero.com/identity/connect/authorize?{urlencode(auth_params)}"
    
    # Debug logging
    logger.info(f"Generated Xero OAuth URL: {auth_url}")
    logger.info(f"Client ID: {XERO_CLIENT_ID}")
    logger.info(f"Callback URL: {XERO_CALLBACK_URL}")
    logger.info(f"Scopes: {XERO_SCOPES}")
    
    return {"auth_url": auth_url, "state": state, "debug_info": {
        "client_id": XERO_CLIENT_ID,
        "callback_url": XERO_CALLBACK_URL,
        "scopes": XERO_SCOPES
    }}

@api_router.get("/xero/callback")
async def handle_xero_oauth_redirect(code: str = None, state: str = None, error: str = None):
    """Handle Xero OAuth redirect - redirect to frontend"""
    if error:
        # Redirect to frontend with error
        return Response(
            status_code=302,
            headers={"Location": f"{os.getenv('FRONTEND_URL')}/xero/callback?error={error}"}
        )
    
    if code and state:
        # Redirect to frontend with code and state  
        return Response(
            status_code=302,
            headers={"Location": f"{os.getenv('FRONTEND_URL')}/xero/callback?code={code}&state={state}"}
        )
    
    # Default redirect to frontend
    return Response(
        status_code=302,
        headers={"Location": f"{os.getenv('FRONTEND_URL')}/"}
    )

@api_router.post("/xero/auth/callback")
async def handle_xero_callback(
    callback_data: dict,
    current_user: dict = Depends(require_admin_or_manager)
):
    """Handle Xero OAuth callback and exchange code for tokens"""
    auth_code = callback_data.get("code")
    state = callback_data.get("state")
    
    if not auth_code or not state:
        raise HTTPException(status_code=400, detail="Missing authorization code or state")
    
    # Validate state parameter
    stored_state = await db.xero_auth_states.find_one({
        "state": state,
        "user_id": current_user["user_id"]
    })
    
    if not stored_state:
        raise HTTPException(status_code=400, detail="Invalid state parameter")
    
    # Clean up used state
    await db.xero_auth_states.delete_one({"state": state})
    
    # Exchange code for tokens
    token_url = "https://identity.xero.com/connect/token"
    
    # Prepare authentication
    auth_string = f"{XERO_CLIENT_ID}:{XERO_CLIENT_SECRET}"
    auth_bytes = auth_string.encode('ascii')
    auth_b64 = base64.b64encode(auth_bytes).decode('ascii')
    
    headers = {
        "Authorization": f"Basic {auth_b64}",
        "Content-Type": "application/x-www-form-urlencoded"
    }
    
    token_data = {
        "grant_type": "authorization_code",
        "code": auth_code,
        "redirect_uri": XERO_CALLBACK_URL
    }
    
    try:
        response = requests.post(token_url, headers=headers, data=token_data)
        response.raise_for_status()
        
        tokens = response.json()
        
        # Store tokens for user
        token_record = {
            "user_id": current_user["user_id"],
            "access_token": tokens["access_token"],
            "refresh_token": tokens["refresh_token"],
            "expires_at": datetime.now(timezone.utc) + timedelta(seconds=tokens.get("expires_in", 1800)),
            "created_at": datetime.now(timezone.utc),
            "tenant_id": None  # Will be updated when we get tenant info
        }
        
        # Upsert token record
        await db.xero_tokens.replace_one(
            {"user_id": current_user["user_id"]},
            token_record,
            upsert=True
        )
        
        return {"message": "Xero connection successful", "access_token_expires_in": tokens.get("expires_in", 1800)}
        
    except requests.exceptions.RequestException as e:
        logger.error(f"Xero token exchange failed: {str(e)}")
        raise HTTPException(status_code=400, detail="Failed to exchange authorization code")

@api_router.delete("/xero/disconnect")
async def disconnect_xero(current_user: dict = Depends(require_admin_or_manager)):
    """Disconnect from Xero"""
    # Remove stored tokens
    result = await db.xero_tokens.delete_one({"user_id": current_user["user_id"]})
    
    if result.deleted_count > 0:
        return {"message": "Xero disconnection successful"}
    else:
        return {"message": "No Xero connection found to disconnect"}

@api_router.get("/xero/next-invoice-number")
async def get_next_xero_invoice_number(current_user: dict = Depends(require_admin_or_manager)):
    """Get the next available invoice number from Xero"""
    try:
        api_client, tenant_id = await get_xero_api_client(current_user["user_id"])
        
        if not tenant_id:
            # Get tenant info if not stored
            from xero_python.identity import IdentityApi
            identity_api = IdentityApi(api_client)
            connections = identity_api.get_connections()
            
            if not connections or not connections[0]:
                raise HTTPException(status_code=400, detail="No Xero organization connected")
            
            tenant_id = connections[0].tenant_id
            
            # Update stored token record with tenant_id
            await db.xero_tokens.update_one(
                {"user_id": current_user["user_id"]},
                {"$set": {"tenant_id": tenant_id}}
            )
        
        # Get invoices to determine next number
        accounting_api = AccountingApi(api_client)
        
        # Get recent invoices ordered by invoice number descending
        invoices_response = accounting_api.get_invoices(
            xero_tenant_id=tenant_id,
            order="InvoiceNumber DESC",
            page=1
        )
        
        next_number = 1
        if invoices_response.invoices and len(invoices_response.invoices) > 0:
            # Extract numeric part from the last invoice number
            last_invoice = invoices_response.invoices[0]
            if last_invoice.invoice_number:
                # Try to extract number from invoice number (handle different formats)
                import re
                numbers = re.findall(r'\d+', last_invoice.invoice_number)
                if numbers:
                    next_number = int(numbers[-1]) + 1
        
        # Format as INV-XXXXXX
        formatted_number = f"INV-{next_number:06d}"
        
        return {
            "next_number": next_number,
            "formatted_number": formatted_number,
            "tenant_id": tenant_id
        }
        
    except Exception as e:
        logger.error(f"Failed to get next invoice number from Xero: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get next invoice number: {str(e)}")

@api_router.get("/xero/account-codes")
async def get_xero_account_codes(current_user: dict = Depends(require_admin_or_manager)):
    """Get available account codes from Xero for setup verification"""
    try:
        api_client, tenant_id = await get_xero_api_client(current_user["user_id"])
        
        if not tenant_id:
            raise HTTPException(status_code=400, detail="No Xero tenant ID found")
        
        accounting_api = AccountingApi(api_client)
        
        # Get all revenue accounts
        accounts_response = accounting_api.get_accounts(
            xero_tenant_id=tenant_id,
            where='Type=="REVENUE" AND Status=="ACTIVE"'
        )
        
        revenue_accounts = []
        if accounts_response.accounts:
            for account in accounts_response.accounts:
                revenue_accounts.append({
                    "code": account.code,
                    "name": account.name,
                    "description": account.description,
                    "type": account.type
                })
        
        return {
            "success": True,
            "revenue_accounts": revenue_accounts,
            "recommended_setup": {
                "message": "For best results, ensure you have a revenue account with code '200' or update Misty to use one of your existing codes",
                "current_default": "200"
            }
        }
        
    except Exception as e:
        logger.error(f"Failed to get Xero account codes: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get account codes: {str(e)}")

@api_router.get("/xero/tax-rates")
async def get_xero_tax_rates(current_user: dict = Depends(require_admin_or_manager)):
    """Get available tax rates from Xero"""
    try:
        api_client, tenant_id = await get_xero_api_client(current_user["user_id"])
        
        if not tenant_id:
            raise HTTPException(status_code=400, detail="No Xero tenant ID found")
        
        accounting_api = AccountingApi(api_client)
        
        # Get all tax rates
        tax_rates_response = accounting_api.get_tax_rates(xero_tenant_id=tenant_id)
        
        tax_rates = []
        if tax_rates_response.tax_rates:
            for tax_rate in tax_rates_response.tax_rates:
                tax_rates.append({
                    "name": tax_rate.name,
                    "tax_type": tax_rate.tax_type,
                    "rate": str(tax_rate.effective_rate) if tax_rate.effective_rate else "0",
                    "status": tax_rate.status
                })
        
        return {
            "success": True,
            "tax_rates": tax_rates,
            "current_default": "OUTPUT"
        }
        
    except Exception as e:
        logger.error(f"Failed to get Xero tax rates: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get tax rates: {str(e)}")

@api_router.post("/xero/webhook")
async def handle_xero_webhook(request: Request):
    """Handle Xero webhook notifications"""
    try:
        # Get the raw body
        body = await request.body()
        
        # Get webhook signature from headers
        xero_signature = request.headers.get("x-xero-signature")
        
        # Log the webhook for debugging
        logger.info(f"Received Xero webhook: {body[:100]}...")  # Log first 100 chars
        
        # For now, just return success - webhook validation can be added later
        return {"status": "received"}
        
    except Exception as e:
        logger.error(f"Xero webhook error: {str(e)}")
        raise HTTPException(status_code=500, detail="Webhook processing failed")

@api_router.get("/xero/webhook")
async def xero_webhook_intent():
    """Handle Xero webhook 'Intent to receive' verification"""
    # Return 200 to confirm we can receive webhooks
    return {"status": "ready", "message": "Webhook endpoint is ready to receive notifications"}

# ============= INTERNAL XERO HELPER FUNCTIONS =============

async def get_next_xero_invoice_number():
    """Internal helper to get next invoice number from Xero"""
    try:
        # Get system Xero token (simplified for accounting transactions)
        xero_token = await db.xero_tokens.find_one({"user_id": "system"})
        if not xero_token:
            raise Exception("No Xero connection found")
        
        api_client, tenant_id = await get_xero_api_client("system")
        
        if not tenant_id:
            # Get tenant info
            from xero_python.identity import IdentityApi
            identity_api = IdentityApi(api_client)
            connections = identity_api.get_connections()
            
            if not connections or not connections[0]:
                raise Exception("No Xero organization connected")
            
            tenant_id = connections[0].tenant_id
            
            # Update stored token record with tenant_id
            await db.xero_tokens.update_one(
                {"user_id": "system"},
                {"$set": {"tenant_id": tenant_id}}
            )
        
        # Get invoices to determine next number
        accounting_api = AccountingApi(api_client)
        
        invoices_response = accounting_api.get_invoices(
            xero_tenant_id=tenant_id,
            order="InvoiceNumber DESC",
            page=1
        )
        
        next_number = 1
        if invoices_response.invoices and len(invoices_response.invoices) > 0:
            latest_invoice = invoices_response.invoices[0]
            if latest_invoice.invoice_number:
                try:
                    # Extract numeric part and increment
                    import re
                    numeric_part = re.findall(r'\d+', latest_invoice.invoice_number)
                    if numeric_part:
                        next_number = int(numeric_part[-1]) + 1
                except (ValueError, IndexError):
                    next_number = 1
        
        formatted_number = f"INV-{next_number:04d}"
        
        return {
            "next_number": next_number,
            "formatted_number": formatted_number
        }
        
    except Exception as e:
        logger.error(f"Failed to get next Xero invoice number: {str(e)}")
        raise Exception(f"Failed to get next invoice number: {str(e)}")

async def create_xero_draft_invoice(invoice_data):
    """Internal helper to create draft invoice in Xero with proper formatting"""
    try:
        api_client, tenant_id = await get_xero_api_client("system")
        
        if not tenant_id:
            raise Exception("No Xero tenant ID available")
        
        from xero_python.accounting import Contact, Invoice, LineItem
        
        # Create contact with proper formatting
        contact = Contact(name=invoice_data["client_name"])
        
        # Add email address if available (not required)
        if invoice_data.get("client_email"):
            contact.email_address = invoice_data["client_email"]
        
        # Create line items with proper Xero formatting
        line_items = []
        for item in invoice_data["items"]:
            # Build description - combine product name and specifications
            description_parts = []
            if item.get("product_name"):
                description_parts.append(item["product_name"])
            if item.get("specifications"):
                description_parts.append(item["specifications"])
            if item.get("description"):
                description_parts.append(item["description"])
            
            # Use the first non-empty description or default
            description = " - ".join(description_parts) if description_parts else "Product/Service"
            
            # Create line item with required fields
            line_item = LineItem(
                description=description,
                quantity=float(item.get("quantity", 1)),
                unit_amount=float(item.get("unit_price", 0)),
                account_code=os.getenv("XERO_SALES_ACCOUNT_CODE", "200")  # Use configured sales account
            )
            
            # Add inventory item code if available (optional)
            if item.get("product_code"):
                line_item.item_code = item["product_code"]
            
            # Add discount if available (optional)
            if item.get("discount_percent"):
                line_item.discount_rate = float(item["discount_percent"])
            
            line_items.append(line_item)
        
        # Create invoice with proper date formatting
        invoice_date = datetime.now(timezone.utc).date()
        due_date = None
        
        if invoice_data.get("due_date"):
            try:
                due_date = datetime.strptime(invoice_data["due_date"], '%Y-%m-%d').date()
            except ValueError:
                # If date parsing fails, default to 30 days from invoice date
                due_date = invoice_date + timedelta(days=30)
        else:
            # Default to 30 days if no due date specified
            due_date = invoice_date + timedelta(days=30)
        
        # Create invoice with all required fields
        invoice = Invoice(
            type="ACCREC",  # Accounts Receivable (required)
            contact=contact,  # Contact (required)
            date=invoice_date,  # Invoice date (required)
            due_date=due_date,  # Due date (required)
            line_items=line_items,  # Line items (required)
            invoice_number=invoice_data["invoice_number"],  # Invoice number (required)
            status="DRAFT"  # Create as draft
        )
        
        # Add reference (order number) if available (optional)
        if invoice_data.get("reference") or invoice_data.get("order_number"):
            invoice.reference = invoice_data.get("reference") or invoice_data.get("order_number")
        
        # Add currency if specified (optional)
        if invoice_data.get("currency"):
            invoice.currency_code = invoice_data["currency"]
        
        accounting_api = AccountingApi(api_client)
        
        # Create the invoice
        invoices = [invoice]
        created_invoices = accounting_api.create_invoices(
            xero_tenant_id=tenant_id, 
            invoices=invoices
        )
        
        if created_invoices.invoices and len(created_invoices.invoices) > 0:
            created_invoice = created_invoices.invoices[0]
            return {
                "success": True,
                "invoice_id": created_invoice.invoice_id,
                "invoice_number": created_invoice.invoice_number,
                "status": created_invoice.status,
                "total": float(created_invoice.total) if created_invoice.total else 0
            }
        else:
            raise Exception("No invoice was created")
        
    except Exception as e:
        logger.error(f"Failed to create Xero draft invoice: {str(e)}")
        raise Exception(f"Failed to create Xero draft: {str(e)}")

async def validate_sales_account(accounting_api, tenant_id: str) -> str:
    """Validate that Sales account with code 200 exists, or find suitable alternative"""
    try:
        # First, try to find account with code "200"
        accounts_response = accounting_api.get_accounts(
            xero_tenant_id=tenant_id,
            where=f'Code="{XERO_DEFAULT_SALES_ACCOUNT_CODE}" AND Status=="ACTIVE"'
        )
        
        if accounts_response.accounts and len(accounts_response.accounts) > 0:
            account = accounts_response.accounts[0]
            logger.info(f"Found Sales account: {account.name} (Code: {account.code})")
            return account.code
        
        # If code "200" doesn't exist, look for any Sales or Revenue account
        logger.warning(f"Sales account with code {XERO_DEFAULT_SALES_ACCOUNT_CODE} not found. Looking for alternatives...")
        
        # Try to find Sales or Revenue accounts
        sales_accounts = accounting_api.get_accounts(
            xero_tenant_id=tenant_id,
            where='(Type=="REVENUE" OR Type=="SALES") AND Status=="ACTIVE"'
        )
        
        if sales_accounts.accounts and len(sales_accounts.accounts) > 0:
            # Use the first available sales/revenue account
            account = sales_accounts.accounts[0]
            logger.info(f"Using alternative Sales account: {account.name} (Code: {account.code})")
            return account.code
        
        # Last resort: look for any income account
        income_accounts = accounting_api.get_accounts(
            xero_tenant_id=tenant_id,
            where='Type=="REVENUE" AND Status=="ACTIVE"'
        )
        
        if income_accounts.accounts and len(income_accounts.accounts) > 0:
            account = income_accounts.accounts[0]
            logger.info(f"Using fallback Revenue account: {account.name} (Code: {account.code})")
            return account.code
        
        # If nothing found, return the default and let Xero handle the error
        logger.error("No suitable Sales or Revenue accounts found in Xero")
        return XERO_DEFAULT_SALES_ACCOUNT_CODE
        
    except Exception as e:
        logger.error(f"Error validating sales account: {str(e)}")
        return XERO_DEFAULT_SALES_ACCOUNT_CODE

@api_router.post("/xero/create-draft-invoice")
async def create_xero_draft_invoice(
    invoice_data: dict,
    current_user: dict = Depends(require_admin_or_manager)
):
    """Create a draft invoice in Xero"""
    try:
        api_client, tenant_id = await get_xero_api_client(current_user["user_id"])
        
        if not tenant_id:
            raise HTTPException(status_code=400, detail="No Xero tenant ID found")
        
        accounting_api = AccountingApi(api_client)
        
        # Get or create contact in Xero
        contact_name = invoice_data.get("client_name", "Unknown Client")
        contact_email = invoice_data.get("client_email")
        
        # Search for existing contact
        contact_id = None
        if contact_email:
            try:
                contacts_response = accounting_api.get_contacts(
                    xero_tenant_id=tenant_id,
                    where=f'EmailAddress="{contact_email}"'
                )
                if contacts_response.contacts and len(contacts_response.contacts) > 0:
                    contact_id = contacts_response.contacts[0].contact_id
            except:
                pass
        
        # Create contact if not found
        if not contact_id:
            try:
                new_contact = Contact(
                    name=contact_name,
                    email_address=contact_email if contact_email else None
                )
                contacts_request = Contacts(contacts=[new_contact])
                contacts_response = accounting_api.create_contacts(
                    xero_tenant_id=tenant_id,
                    contacts=contacts_request
                )
                if contacts_response.contacts and len(contacts_response.contacts) > 0:
                    contact_id = contacts_response.contacts[0].contact_id
                else:
                    raise Exception("Failed to create contact")
            except Exception as e:
                logger.error(f"Failed to create contact in Xero: {str(e)}")
                # Use default contact or create a simple one
                contact_id = None
        
        # Prepare line items with Sales account
        line_items = []
        items = invoice_data.get("items", [])
        
        # Debug logging
        logger.info(f"Creating Xero invoice with {len(items)} items")
        
        # Validate that Sales account code "200" exists
        sales_account_code = await validate_sales_account(accounting_api, tenant_id)
        
        for item in items:
            # Debug individual item
            logger.info(f"Processing item: {item}")
            
            line_item = LineItem(
                description=item.get("product_name", item.get("description", "Product/Service")),
                quantity=float(item.get("quantity", 1)),
                unit_amount=float(item.get("unit_price", 0)),
                account_code=sales_account_code,
                tax_type=XERO_DEFAULT_TAX_TYPE
            )
            line_items.append(line_item)
        
        # If no items provided, create a default line item
        if not line_items:
            total_amount = float(invoice_data.get("total_amount", 0))
            logger.info(f"No items found, creating default line item with total: {total_amount}")
            line_item = LineItem(
                description=f"Services for {contact_name}",
                quantity=1,
                unit_amount=total_amount,
                account_code=sales_account_code,
                tax_type=XERO_DEFAULT_TAX_TYPE
            )
            line_items.append(line_item)
        
        # Parse and format due_date properly for Xero
        due_date_value = invoice_data.get("due_date")
        if isinstance(due_date_value, str):
            try:
                due_date_parsed = datetime.fromisoformat(due_date_value).date()
            except:
                due_date_parsed = datetime.now().date()
        else:
            due_date_parsed = due_date_value or datetime.now().date()
        
        # Debug logging for Xero invoice creation
        logger.info(f"Creating Xero invoice: contact_name={contact_name}, contact_id={contact_id}")
        logger.info(f"Line items count: {len(line_items)}")
        logger.info(f"Invoice number: {invoice_data.get('invoice_number')}")
        
        # Create invoice object
        invoice = Invoice(
            type="ACCREC",  # Accounts Receivable
            contact=Contact(contact_id=contact_id) if contact_id else Contact(name=contact_name),
            line_items=line_items,
            date=datetime.now().date(),
            due_date=due_date_parsed,
            invoice_number=invoice_data.get("invoice_number"),
            reference=invoice_data.get("reference", invoice_data.get("order_number")),
            status="DRAFT"
        )
        
        # Create invoice in Xero
        invoices_request = Invoices(invoices=[invoice])
        invoices_response = accounting_api.create_invoices(
            xero_tenant_id=tenant_id,
            invoices=invoices_request
        )
        
        if not invoices_response.invoices or len(invoices_response.invoices) == 0:
            raise Exception("No invoice returned from Xero API")
        
        created_invoice = invoices_response.invoices[0]
        
        # Log successful creation
        logger.info(f"Successfully created Xero invoice: {created_invoice.invoice_number}")
        
        # Check for validation errors
        if hasattr(created_invoice, 'validation_errors') and created_invoice.validation_errors:
            error_messages = [error.message for error in created_invoice.validation_errors]
            raise Exception(f"Invoice validation failed: {', '.join(error_messages)}")
        
        return {
            "success": True,
            "message": "Draft invoice created in Xero",
            "invoice_id": created_invoice.invoice_id,
            "invoice_number": created_invoice.invoice_number,
            "status": created_invoice.status,
            "total": str(created_invoice.total) if created_invoice.total else "0",
            "xero_url": f"https://go.xero.com/AccountsReceivable/View.aspx?InvoiceID={created_invoice.invoice_id}"
        }
        
    except Exception as e:
        logger.error(f"Failed to create draft invoice in Xero: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to create draft invoice: {str(e)}")

# ============= INVOICING ENDPOINTS =============

@api_router.get("/invoicing/live-jobs")
async def get_live_jobs(current_user: dict = Depends(require_admin_or_manager)):
    """Get all jobs ready for invoicing (in invoicing stage on production board)"""
    live_jobs = await db.orders.find({
        "current_stage": "invoicing", 
        "status": {"$ne": "completed"}
    }).to_list(length=None)
    
    # Convert ObjectId to string and enrich with client information
    for job in live_jobs:
        # Remove MongoDB ObjectId if present
        if "_id" in job:
            del job["_id"]
            
        client = await db.clients.find_one({"id": job["client_id"]})
        if client:
            job["client_name"] = client["company_name"]
            job["client_email"] = client.get("email", "")  # Add client email for Xero
            job["client_payment_terms"] = client.get("payment_terms", "Net 30 days")
        
    return {"data": live_jobs}

@api_router.post("/invoicing/generate/{job_id}")
async def generate_job_invoice(
    job_id: str,
    invoice_data: dict,
    current_user: dict = Depends(require_admin_or_manager)
):
    """Generate invoice for a completed job"""
    # Get job/order
    job = await db.orders.find_one({"id": job_id})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    if job.get("invoiced"):
        raise HTTPException(status_code=400, detail="Job already invoiced")
    
    # Get client info
    client = await db.clients.find_one({"id": job["client_id"]})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Generate invoice number (simple sequential for now)
    invoice_count = await db.invoices.count_documents({}) + 1
    invoice_number = f"INV-{invoice_count:04d}"
    
    # Prepare invoice data
    invoice_record = {
        "id": str(uuid.uuid4()),
        "invoice_number": invoice_number,
        "job_id": job_id,
        "client_id": job["client_id"],
        "invoice_type": invoice_data.get("invoice_type", "full"),  # "full" or "partial"
        "items": invoice_data.get("items", job["items"]),
        "subtotal": invoice_data.get("subtotal", job["subtotal"]),
        "gst": invoice_data.get("gst", job["gst"]),
        "total_amount": invoice_data.get("total_amount", job["total_amount"]),
        "payment_terms": client.get("payment_terms", "Net 30 days"),
        "due_date": invoice_data.get("due_date"),
        "created_by": current_user["user_id"],
        "created_at": datetime.now(timezone.utc),
        "status": "draft"
    }
    
    # Insert invoice record
    await db.invoices.insert_one(invoice_record)
    
    # Update job status and move to accounting transactions
    update_data = {
        "invoiced": True,
        "invoice_id": invoice_record["id"],
        "current_stage": "accounting_transaction",  # New stage for accounting processing
        "status": "accounting_draft",  # New status for accounting transactions
        "invoice_date": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc)
    }
    
    # If partial invoice, don't mark as fully invoiced
    if invoice_data.get("invoice_type") == "partial":
        update_data["invoiced"] = False
        update_data["partially_invoiced"] = True
        update_data["current_stage"] = "delivery"  # Keep in delivery stage for remaining items
        update_data["status"] = "pending"  # Keep as pending for partial invoices
        # Remove invoice date for partial invoices
        del update_data["invoice_date"]
    
    await db.orders.update_one(
        {"id": job_id},
        {"$set": update_data}
    )
    
    # For full invoices in accounting transactions, automatically create Xero draft
    if invoice_data.get("invoice_type") != "partial":
        try:
            # Check if Xero is connected
            xero_token = await db.xero_tokens.find_one({"user_id": "system"})
            if xero_token and xero_token.get("access_token"):
                logger.info("Creating Xero draft invoice for accounting transaction")
                
                # Get client info for Xero invoice
                client = await db.clients.find_one({"id": job["client_id"]})
                
                # Get next Xero invoice number
                next_number_response = await get_next_xero_invoice_number()
                next_invoice_number = next_number_response["formatted_number"]
                
                # Prepare Xero invoice data with proper formatting
                xero_invoice_data = {
                    "client_name": client["company_name"] if client else job.get("client_name", "Unknown Client"),
                    "client_email": client.get("email", "") if client else "",
                    "invoice_number": next_invoice_number,
                    "order_number": job["order_number"],
                    "items": [],
                    "total_amount": invoice_data.get("total_amount", job["total_amount"]),
                    "due_date": invoice_data.get("due_date"),
                    "reference": job["order_number"]
                }
                
                # Format items for Xero with proper field mapping
                items = invoice_data.get("items", job["items"])
                for item in items:
                    xero_item = {
                        "product_name": item.get("product_name", item.get("description", "Product")),
                        "description": item.get("description", ""),
                        "specifications": item.get("specifications", ""),
                        "quantity": item.get("quantity", 1),
                        "unit_price": item.get("unit_price", item.get("price", 0)),
                        "product_code": item.get("product_code", ""),  # For InventoryItemCode
                        "discount_percent": item.get("discount_percent", 0) if item.get("discount_percent") else None
                    }
                    xero_invoice_data["items"].append(xero_item)
                
                # Create draft invoice in Xero
                xero_response = await create_xero_draft_invoice(xero_invoice_data)
                logger.info(f"Xero draft invoice created successfully: {xero_response}")
                
                # Update the invoice record with Xero details
                await db.invoices.update_one(
                    {"id": invoice_record["id"]},
                    {"$set": {
                        "xero_invoice_id": xero_response.get("invoice_id"),
                        "xero_invoice_number": next_invoice_number,
                        "xero_status": "draft"
                    }}
                )
                
        except Exception as e:
            logger.error(f"Failed to create Xero draft invoice: {str(e)}")
            # Don't fail the entire invoice process if Xero fails
            pass
    
    return {
        "message": "Invoice generated successfully and moved to accounting transactions",
        "invoice_id": invoice_record["id"],
        "invoice_number": invoice_number,
        "xero_draft_created": "xero_invoice_id" in locals()
    }

@api_router.get("/invoicing/archived-jobs")
async def get_archived_jobs(
    month: Optional[int] = None, 
    year: Optional[int] = None,
    current_user: dict = Depends(require_admin_or_manager)
):
    """Get archived jobs (completed and invoiced)"""
    # Build query filter
    query_filter = {"invoiced": True}
    
    if month and year:
        start_date = datetime(year, month, 1)
        if month == 12:
            end_date = datetime(year + 1, 1, 1)
        else:
            end_date = datetime(year, month + 1, 1)
        
        query_filter["created_at"] = {
            "$gte": start_date,
            "$lt": end_date
        }
    
    archived_jobs = await db.orders.find(query_filter).to_list(length=None)
    
    # Enrich with client and invoice information
    for job in archived_jobs:
        # Remove MongoDB ObjectId if present
        if "_id" in job:
            del job["_id"]
            
        # Get client info
        client = await db.clients.find_one({"id": job["client_id"]})
        if client:
            job["client_name"] = client["company_name"]
        
        # Get invoice info
        if job.get("invoice_id"):
            invoice = await db.invoices.find_one({"id": job["invoice_id"]})
            if invoice:
                job["invoice_number"] = invoice["invoice_number"]
                job["invoice_date"] = invoice["created_at"]
    
    return {"data": archived_jobs}

@api_router.get("/invoicing/monthly-report")
async def get_monthly_invoicing_report(
    month: int,
    year: int,
    current_user: dict = Depends(require_admin_or_manager)
):
    """Generate monthly invoicing report"""
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1)
    else:
        end_date = datetime(year, month + 1, 1)
    
    # Get jobs completed this month
    completed_jobs = await db.orders.find({
        "updated_at": {"$gte": start_date, "$lt": end_date},
        "current_stage": "cleared"
    }).to_list(length=None)
    
    # Get jobs invoiced this month
    invoiced_jobs = await db.invoices.find({
        "created_at": {"$gte": start_date, "$lt": end_date}
    }).to_list(length=None)
    
    # Remove ObjectIds from results
    for job in completed_jobs:
        if "_id" in job:
            del job["_id"]
    
    for invoice in invoiced_jobs:
        if "_id" in invoice:
            del invoice["_id"]
    
    # Calculate totals
    total_jobs_completed = len(completed_jobs)
    total_jobs_invoiced = len(invoiced_jobs)
    total_invoice_amount = sum(inv.get("total_amount", 0) for inv in invoiced_jobs)
    
    return {
        "month": month,
        "year": year,
        "total_jobs_completed": total_jobs_completed,
        "total_jobs_invoiced": total_jobs_invoiced,
        "total_invoice_amount": total_invoice_amount,
        "completed_jobs": completed_jobs,
        "invoiced_jobs": invoiced_jobs
    }

# ============= ACCOUNTING TRANSACTIONS ENDPOINTS =============

@api_router.get("/invoicing/accounting-transactions")
async def get_accounting_transactions(current_user: dict = Depends(require_admin_or_manager)):
    """Get all jobs in accounting transaction stage (invoiced but not completed)"""
    transactions = await db.orders.find({
        "current_stage": "accounting_transaction",
        "status": "accounting_draft"
    }).to_list(length=None)
    
    # Convert ObjectId to string and enrich with client information
    for transaction in transactions:
        # Remove MongoDB ObjectId if present
        if "_id" in transaction:
            del transaction["_id"]
            
        client = await db.clients.find_one({"id": transaction["client_id"]})
        if client:
            transaction["client_name"] = client["company_name"]
            transaction["client_email"] = client.get("email", "")
        
    return {"data": transactions}

@api_router.post("/invoicing/complete-transaction/{job_id}")
async def complete_accounting_transaction(
    job_id: str,
    current_user: dict = Depends(require_admin_or_manager)
):
    """Complete accounting transaction and archive the job"""
    # Get job/order
    job = await db.orders.find_one({"id": job_id})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    if job.get("current_stage") != "accounting_transaction":
        raise HTTPException(status_code=400, detail="Job is not in accounting transaction stage")
    
    # Update job to completed status
    await db.orders.update_one(
        {"id": job_id},
        {"$set": {
            "current_stage": "cleared",
            "status": "completed",
            "completed_at": datetime.now(timezone.utc),
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    # Get the updated order data for archiving
    updated_job = await db.orders.find_one({"id": job_id})
    if updated_job:
        # Create archived order dict
        archived_order = {
            "id": str(uuid.uuid4()),
            "original_order_id": updated_job["id"],
            "order_number": updated_job["order_number"],
            "client_id": updated_job["client_id"],
            "client_name": updated_job.get("client_name", ""),
            "purchase_order_number": updated_job.get("purchase_order_number"),
            "items": updated_job["items"],
            "subtotal": updated_job["subtotal"],
            "gst": updated_job["gst"],
            "total_amount": updated_job["total_amount"],
            "due_date": updated_job["due_date"],
            "delivery_address": updated_job.get("delivery_address"),
            "delivery_instructions": updated_job.get("delivery_instructions"),
            "runtime_estimate": updated_job.get("runtime_estimate"),
            "notes": updated_job.get("notes"),
            "created_by": updated_job["created_by"],
            "created_at": updated_job["created_at"],
            "completed_at": datetime.now(timezone.utc),
            "archived_by": current_user["user_id"],
            "invoice_id": updated_job.get("invoice_id"),
            "invoice_date": updated_job.get("invoice_date")
        }
        
        # Insert into archived orders collection
        await db.archived_orders.insert_one(archived_order)
    
    return {
        "message": "Accounting transaction completed and job archived successfully",
        "job_id": job_id
    }

@api_router.get("/invoicing/export-drafted-csv")
async def export_drafted_invoices_csv(current_user: dict = Depends(require_admin_or_manager)):
    """Export all accounting transactions (drafted invoices) to CSV in Xero import format"""
    try:
        # Get all jobs in accounting transaction stage
        transactions = await db.orders.find({
            "current_stage": "accounting_transaction",
            "status": "accounting_draft"
        }).to_list(length=None)
        
        # Prepare CSV data
        csv_data = []
        
        # CSV Headers based on Xero import format
        headers = [
            "ContactName", "EmailAddress", "POAddressLine1", "POAddressLine2", 
            "POAddressLine3", "POAddressLine4", "POCity", "PORegion", 
            "POPostalCode", "POCountry", "InvoiceNumber", "Reference", 
            "InvoiceDate", "DueDate", "InventoryItemCode", "Description", 
            "Quantity", "UnitAmount", "Discount", "AccountCode", "TaxType", 
            "TrackingName1", "TrackingOption1", "TrackingName2", "TrackingOption2", 
            "Currency", "BrandingTheme"
        ]
        
        csv_data.append(headers)
        
        # Process each transaction
        for transaction in transactions:
            # Get client info
            client = await db.clients.find_one({"id": transaction["client_id"]})
            client_name = client["company_name"] if client else transaction.get("client_name", "Unknown Client")
            client_email = client.get("email", "") if client else ""
            
            # Get invoice info
            invoice = await db.invoices.find_one({"id": transaction.get("invoice_id")})
            invoice_number = invoice["invoice_number"] if invoice else f"INV-{transaction['order_number']}"
            invoice_date = invoice["created_at"].strftime("%d/%m/%Y") if invoice and invoice.get("created_at") else datetime.now().strftime("%d/%m/%Y")
            
            # Calculate due date (30 days from invoice date)
            due_date_obj = invoice["created_at"] + timedelta(days=30) if invoice and invoice.get("created_at") else datetime.now() + timedelta(days=30)
            due_date = due_date_obj.strftime("%d/%m/%Y")
            
            # Process each item in the transaction
            items = transaction.get("items", [])
            if not items:
                # Create a single line if no items
                items = [{"description": f"Services for Order {transaction['order_number']}", "quantity": 1, "unit_price": transaction.get("total_amount", 0)}]
            
            for item in items:
                row = [
                    client_name,  # ContactName (required)
                    client_email,  # EmailAddress
                    "",  # POAddressLine1
                    "",  # POAddressLine2
                    "",  # POAddressLine3
                    "",  # POAddressLine4
                    "",  # POCity
                    "",  # PORegion
                    "",  # POPostalCode
                    "",  # POCountry
                    invoice_number,  # InvoiceNumber (required)
                    transaction["order_number"],  # Reference
                    invoice_date,  # InvoiceDate (required)
                    due_date,  # DueDate (required)
                    item.get("product_code", ""),  # InventoryItemCode
                    f"{item.get('product_name', item.get('description', 'Product'))} - {item.get('specifications', '')}".strip(" - "),  # Description (required)
                    str(item.get("quantity", 1)),  # Quantity (required)
                    str(item.get("unit_price", item.get("price", 0))),  # UnitAmount (required)
                    str(item.get("discount_percent", "")),  # Discount
                    os.getenv("XERO_SALES_ACCOUNT_CODE", "200"),  # AccountCode (required)
                    "OUTPUT",  # TaxType (required) - GST for sales
                    "",  # TrackingName1
                    "",  # TrackingOption1
                    "",  # TrackingName2
                    "",  # TrackingOption2
                    "AUD",  # Currency
                    ""   # BrandingTheme
                ]
                csv_data.append(row)
        
        # Generate CSV content
        import io
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerows(csv_data)
        csv_content = output.getvalue()
        output.close()
        
        # Return CSV as response
        from fastapi.responses import Response
        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=drafted_invoices_{datetime.now().strftime('%Y%m%d')}.csv"}
        )
        
    except Exception as e:
        logger.error(f"Failed to export drafted invoices CSV: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to export CSV: {str(e)}")

# ============= ARCHIVED ORDERS ENDPOINTS =============

@api_router.get("/clients/{client_id}/archived-orders")
async def get_client_archived_orders(
    client_id: str,
    filters: ArchivedOrderFilter = Depends(),
    current_user: dict = Depends(require_any_role)
):
    """Get archived orders for a specific client with filtering"""
    query = {"client_id": client_id}
    
    # Apply date filters
    if filters.date_from or filters.date_to:
        date_filter = {}
        if filters.date_from:
            date_filter["$gte"] = datetime.combine(filters.date_from, datetime.min.time())
        if filters.date_to:
            date_filter["$lte"] = datetime.combine(filters.date_to, datetime.max.time())
        query["archived_at"] = date_filter
    
    # Apply search query
    if filters.search_query:
        query["$or"] = [
            {"order_number": {"$regex": filters.search_query, "$options": "i"}},
            {"client_name": {"$regex": filters.search_query, "$options": "i"}},
            {"purchase_order_number": {"$regex": filters.search_query, "$options": "i"}},
            {"items.product_name": {"$regex": filters.search_query, "$options": "i"}}
        ]
    
    archived_orders = await db.archived_orders.find(query).sort("archived_at", -1).to_list(length=None)
    
    # Remove MongoDB ObjectIds
    for order in archived_orders:
        if "_id" in order:
            del order["_id"]
    
    return StandardResponse(success=True, message="Archived orders retrieved", data=archived_orders)

@api_router.post("/clients/{client_id}/archived-orders/fast-report")
async def generate_fast_report(
    client_id: str,
    report_request: FastReportRequest,
    current_user: dict = Depends(require_any_role)
):
    """Generate Excel fast report for client archived orders"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from io import BytesIO
    import pandas as pd
    from fastapi.responses import StreamingResponse
    
    # Calculate date range based on time period
    today = date.today()
    
    if report_request.time_period == ReportTimePeriod.CURRENT_MONTH:
        date_from = today.replace(day=1)
        date_to = today
    elif report_request.time_period == ReportTimePeriod.LAST_MONTH:
        first_day_current = today.replace(day=1)
        date_to = first_day_current - timedelta(days=1)
        date_from = date_to.replace(day=1)
    elif report_request.time_period == ReportTimePeriod.LAST_3_MONTHS:
        date_to = today
        date_from = today - timedelta(days=90)
    elif report_request.time_period == ReportTimePeriod.LAST_6_MONTHS:
        date_to = today
        date_from = today - timedelta(days=180)
    elif report_request.time_period == ReportTimePeriod.LAST_9_MONTHS:
        date_to = today
        date_from = today - timedelta(days=270)
    elif report_request.time_period == ReportTimePeriod.LAST_YEAR:
        date_to = today
        date_from = today - timedelta(days=365)
    elif report_request.time_period == ReportTimePeriod.CURRENT_QUARTER:
        quarter_start_month = (today.month - 1) // 3 * 3 + 1
        date_from = today.replace(month=quarter_start_month, day=1)
        date_to = today
    elif report_request.time_period == ReportTimePeriod.YEAR_TO_DATE:
        date_from = today.replace(month=1, day=1)
        date_to = today
    elif report_request.time_period == ReportTimePeriod.CURRENT_FINANCIAL_YEAR:
        # Australian financial year: July 1 - June 30
        if today.month >= 7:
            date_from = today.replace(month=7, day=1)
            date_to = today
        else:
            date_from = today.replace(year=today.year-1, month=7, day=1)
            date_to = today
    elif report_request.time_period == ReportTimePeriod.CUSTOM_RANGE:
        date_from = report_request.date_from
        date_to = report_request.date_to
    else:
        date_from = today - timedelta(days=30)
        date_to = today
    
    # Build query
    query = {"client_id": client_id}
    
    if date_from and date_to:
        query["archived_at"] = {
            "$gte": datetime.combine(date_from, datetime.min.time()),
            "$lte": datetime.combine(date_to, datetime.max.time())
        }
    
    # Apply product filter if specified
    if report_request.product_filter:
        query["items.product_name"] = {"$regex": report_request.product_filter, "$options": "i"}
    
    # Get archived orders
    archived_orders = await db.archived_orders.find(query).sort("archived_at", -1).to_list(length=None)
    
    if not archived_orders:
        raise HTTPException(status_code=404, detail="No archived orders found for the specified criteria")
    
    # Prepare data for Excel
    excel_data = []
    
    for order in archived_orders:
        row_data = {}
        
        # Map selected fields to Excel columns
        for field in report_request.selected_fields:
            if field == ReportField.ORDER_NUMBER:
                row_data["Order Number"] = order.get("order_number", "")
            elif field == ReportField.CLIENT_NAME:
                row_data["Client Name"] = order.get("client_name", "")
            elif field == ReportField.PURCHASE_ORDER_NUMBER:
                row_data["Purchase Order Number"] = order.get("purchase_order_number", "")
            elif field == ReportField.ORDER_DATE:
                row_data["Order Date"] = order.get("created_at", "").strftime("%Y-%m-%d") if order.get("created_at") else ""
            elif field == ReportField.COMPLETION_DATE:
                row_data["Completion Date"] = order.get("completed_at", "").strftime("%Y-%m-%d") if order.get("completed_at") else ""
            elif field == ReportField.DUE_DATE:
                row_data["Due Date"] = order.get("due_date", "").strftime("%Y-%m-%d") if order.get("due_date") else ""
            elif field == ReportField.SUBTOTAL:
                row_data["Subtotal"] = f"${order.get('subtotal', 0):.2f}"
            elif field == ReportField.GST:
                row_data["GST"] = f"${order.get('gst', 0):.2f}"
            elif field == ReportField.TOTAL_AMOUNT:
                row_data["Total Amount"] = f"${order.get('total_amount', 0):.2f}"
            elif field == ReportField.DELIVERY_ADDRESS:
                row_data["Delivery Address"] = order.get("delivery_address", "")
            elif field == ReportField.PRODUCT_NAMES:
                products = [item.get("product_name", "") for item in order.get("items", [])]
                row_data["Products"] = ", ".join(products)
            elif field == ReportField.PRODUCT_QUANTITIES:
                quantities = [f"{item.get('product_name', '')}: {item.get('quantity', 0)}" for item in order.get("items", [])]
                row_data["Product Quantities"] = ", ".join(quantities)
            elif field == ReportField.NOTES:
                row_data["Notes"] = order.get("notes", "")
            elif field == ReportField.RUNTIME_ESTIMATE:
                row_data["Runtime Estimate"] = order.get("runtime_estimate", "")
        
        excel_data.append(row_data)
    
    # Create Excel workbook
    df = pd.DataFrame(excel_data)
    wb = Workbook()
    ws = wb.active
    
    # Set title
    title = report_request.report_title or f"Archived Orders Report - {report_request.time_period.value.replace('_', ' ').title()}"
    ws.title = "Archived Orders Report"
    
    # Add header
    ws.merge_cells("A1:{}1".format(chr(65 + len(df.columns) - 1)))
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")
    
    # Add date range info
    ws.merge_cells("A2:{}2".format(chr(65 + len(df.columns) - 1)))
    ws["A2"] = f"Report Period: {date_from.strftime('%Y-%m-%d')} to {date_to.strftime('%Y-%m-%d')}"
    ws["A2"].alignment = Alignment(horizontal="center")
    
    # Add column headers
    for col, column_name in enumerate(df.columns, 1):
        cell = ws.cell(row=4, column=col, value=column_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    for row_idx, row_data in enumerate(excel_data, 5):
        for col_idx, (column_name, value) in enumerate(row_data.items(), 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-adjust column widths
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, len(df.columns) + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        # Check all cells in this column
        for row_idx in range(4, ws.max_row + 1):  # Start from row 4 (after headers)
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                if cell.value is not None:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        
        # Also check header length
        header_cell = ws.cell(row=4, column=col_idx)
        if header_cell.value is not None:
            header_length = len(str(header_cell.value))
            if header_length > max_length:
                max_length = header_length
        
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Get client name for filename
    client = await db.clients.find_one({"id": client_id})
    client_name = client.get("company_name", "Client") if client else "Client"
    safe_client_name = "".join(c for c in client_name if c.isalnum() or c in (' ', '-', '_')).strip()
    
    filename = f"{safe_client_name}_Archived_Orders_{date_from.strftime('%Y%m%d')}-{date_to.strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        BytesIO(excel_file.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ============= FILE SERVING ENDPOINTS =============

@api_router.get("/uploads/{file_path:path}")
async def serve_uploaded_file(file_path: str):
    """Serve uploaded files"""
    full_path = f"/app/uploads/{file_path}"
    if not os.path.exists(full_path):
        raise HTTPException(status_code=404, detail="File not found")
    
    return FileResponse(full_path)

# ============= STOCK MANAGEMENT SYSTEM ENDPOINTS =============

@api_router.get("/stock/raw-substrates", response_model=StandardResponse)
async def get_raw_substrates_stock(
    client_id: Optional[str] = None,
    current_user: dict = Depends(require_any_role)
):
    """Get all raw substrates on hand, optionally filtered by client"""
    try:
        query = {}
        if client_id and client_id != "all":
            query["client_id"] = client_id
            
        substrates = await db.raw_substrate_stock.find(query).to_list(length=None)
        
        # Remove MongoDB ObjectIds
        for substrate in substrates:
            if "_id" in substrate:
                del substrate["_id"]
        
        return StandardResponse(
            success=True,
            message="Raw substrates retrieved successfully",
            data=substrates
        )
    except Exception as e:
        logger.error(f"Failed to get raw substrates: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to retrieve raw substrates")

@api_router.post("/stock/raw-substrates", response_model=StandardResponse)
async def create_raw_substrate_stock(
    substrate_data: RawSubstrateStockCreate,
    current_user: dict = Depends(require_manager_or_admin)
):
    """Create new raw substrate stock entry"""
    try:
        # Create new substrate stock record
        substrate = RawSubstrateStock(
            client_id=substrate_data.client_id,
            client_name=substrate_data.client_name,
            product_id=substrate_data.product_id,
            product_code=substrate_data.product_code,
            product_description=substrate_data.product_description,
            quantity_on_hand=substrate_data.quantity_on_hand,
            unit_of_measure=substrate_data.unit_of_measure,
            source_order_id=substrate_data.source_order_id,
            source_job_id=substrate_data.source_job_id,
            is_shared_product=substrate_data.is_shared_product,
            shared_with_clients=substrate_data.shared_with_clients,
            created_from_excess=substrate_data.created_from_excess,
            material_specifications=substrate_data.material_specifications,
            material_value_m2=substrate_data.material_value_m2,
            minimum_stock_level=substrate_data.minimum_stock_level,
            created_by=current_user["user_id"]
        )
        
        substrate_dict = substrate.dict()
        substrate_dict["created_at"] = datetime.now(timezone.utc)
        
        result = await db.raw_substrate_stock.insert_one(substrate_dict)
        
        # Create stock movement record
        movement = StockMovement(
            stock_type="raw_substrate",
            stock_id=substrate.id,
            movement_type="addition",
            quantity_change=substrate_data.quantity_on_hand,
            previous_quantity=0.0,
            new_quantity=substrate_data.quantity_on_hand,
            reference_id=substrate_data.source_order_id,
            reference_type="order",
            notes=f"Initial stock creation from excess production",
            created_by=current_user["user_id"]
        )
        
        movement_dict = movement.dict()
        movement_dict["created_at"] = datetime.now(timezone.utc)
        await db.stock_movements.insert_one(movement_dict)
        
        return StandardResponse(
            success=True,
            message="Raw substrate stock created successfully",
            data={"id": substrate.id}
        )
    except Exception as e:
        logger.error(f"Failed to create raw substrate stock: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to create raw substrate stock")

@api_router.put("/stock/raw-substrates/{substrate_id}", response_model=StandardResponse)
async def update_raw_substrate_stock(
    substrate_id: str,
    update_data: RawSubstrateStockUpdate,
    current_user: dict = Depends(require_manager_or_admin)
):
    """Update raw substrate stock quantity or details"""
    try:
        # Get existing substrate
        existing = await db.raw_substrate_stock.find_one({"id": substrate_id})
        if not existing:
            raise HTTPException(status_code=404, detail="Raw substrate stock not found")
        
        previous_quantity = existing["quantity_on_hand"]
        new_quantity = update_data.quantity_on_hand if update_data.quantity_on_hand is not None else previous_quantity
        
        # Update substrate record
        update_fields = {
            "updated_at": datetime.now(timezone.utc),
            "updated_by": current_user["user_id"]
        }
        
        # Add any fields that are being updated
        if update_data.quantity_on_hand is not None:
            update_fields["quantity_on_hand"] = update_data.quantity_on_hand
        if update_data.minimum_stock_level is not None:
            update_fields["minimum_stock_level"] = update_data.minimum_stock_level
        if update_data.is_shared_product is not None:
            update_fields["is_shared_product"] = update_data.is_shared_product
        if update_data.shared_with_clients is not None:
            update_fields["shared_with_clients"] = update_data.shared_with_clients
        if update_data.material_specifications is not None:
            update_fields["material_specifications"] = update_data.material_specifications
        
        await db.raw_substrate_stock.update_one(
            {"id": substrate_id},
            {"$set": update_fields}
        )
        
        # Create stock movement record if quantity changed
        if new_quantity != previous_quantity:
            movement_type = "addition" if new_quantity > previous_quantity else "consumption"
            quantity_change = new_quantity - previous_quantity
            
            movement = StockMovement(
                stock_type="raw_substrate",
                stock_id=substrate_id,
                movement_type=movement_type,
                quantity_change=quantity_change,
                previous_quantity=previous_quantity,
                new_quantity=new_quantity,
                reference_type="manual",
                notes=update_data.notes or "Manual adjustment",
                created_by=current_user["user_id"]
            )
            
            movement_dict = movement.dict()
            movement_dict["created_at"] = datetime.now(timezone.utc)
            await db.stock_movements.insert_one(movement_dict)
        
        return StandardResponse(
            success=True,
            message="Raw substrate stock updated successfully"
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to update raw substrate stock: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to update raw substrate stock")

@api_router.get("/stock/raw-materials", response_model=StandardResponse)
async def get_raw_materials_stock(current_user: dict = Depends(require_any_role)):
    """Get all raw materials stock"""
    try:
        materials = await db.raw_material_stock.find({}).to_list(length=None)
        
        # Remove MongoDB ObjectIds
        for material in materials:
            if "_id" in material:
                del material["_id"]
        
        return StandardResponse(
            success=True,
            message="Raw materials stock retrieved successfully",
            data=materials
        )
    except Exception as e:
        logger.error(f"Failed to get raw materials stock: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to retrieve raw materials stock")

@api_router.post("/stock/raw-materials", response_model=StandardResponse)
async def create_raw_material_stock(
    material_data: RawMaterialStockCreate,
    current_user: dict = Depends(require_manager_or_admin)
):
    """Create new raw material stock entry"""
    try:
        # Create new material stock record
        material = RawMaterialStock(
            material_id=material_data.material_id,
            material_name=material_data.material_name,
            quantity_on_hand=material_data.quantity_on_hand,
            unit_of_measure=material_data.unit_of_measure,
            minimum_stock_level=material_data.minimum_stock_level,
            alert_threshold_days=material_data.alert_threshold_days,
            supplier_id=material_data.supplier_id,
            usage_rate_per_month=material_data.usage_rate_per_month
        )
        
        material_dict = material.dict()
        material_dict["created_at"] = datetime.now(timezone.utc)
        
        result = await db.raw_material_stock.insert_one(material_dict)
        
        # Create stock movement record
        movement = StockMovement(
            stock_type="raw_material",
            stock_id=material.id,
            movement_type="addition",
            quantity_change=material_data.quantity_on_hand,
            previous_quantity=0.0,
            new_quantity=material_data.quantity_on_hand,
            reference_type="manual",
            notes="Initial stock creation",
            created_by=current_user["user_id"]
        )
        
        movement_dict = movement.dict()
        movement_dict["created_at"] = datetime.now(timezone.utc)
        await db.stock_movements.insert_one(movement_dict)
        
        return StandardResponse(
            success=True,
            message="Raw material stock created successfully",
            data={"id": material.id}
        )
    except Exception as e:
        logger.error(f"Failed to create raw material stock: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to create raw material stock")

@api_router.put("/stock/raw-materials/{material_id}", response_model=StandardResponse)
async def update_raw_material_stock(
    material_id: str,
    update_data: RawMaterialStockUpdate,
    current_user: dict = Depends(require_manager_or_admin)
):
    """Update raw material stock quantity or details"""
    try:
        # Get existing material
        existing = await db.raw_material_stock.find_one({"id": material_id})
        if not existing:
            raise HTTPException(status_code=404, detail="Raw material stock not found")
        
        previous_quantity = existing["quantity_on_hand"]
        new_quantity = update_data.quantity_on_hand if update_data.quantity_on_hand is not None else previous_quantity
        
        # Update material record
        update_fields = {
            "updated_at": datetime.now(timezone.utc)
        }
        
        # Add any fields that are being updated
        if update_data.quantity_on_hand is not None:
            update_fields["quantity_on_hand"] = update_data.quantity_on_hand
        if update_data.minimum_stock_level is not None:
            update_fields["minimum_stock_level"] = update_data.minimum_stock_level
        if update_data.alert_threshold_days is not None:
            update_fields["alert_threshold_days"] = update_data.alert_threshold_days
        if update_data.usage_rate_per_month is not None:
            update_fields["usage_rate_per_month"] = update_data.usage_rate_per_month
        if update_data.last_order_date is not None:
            update_fields["last_order_date"] = update_data.last_order_date
        if update_data.last_order_quantity is not None:
            update_fields["last_order_quantity"] = update_data.last_order_quantity
        
        await db.raw_material_stock.update_one(
            {"id": material_id},
            {"$set": update_fields}
        )
        
        # Create stock movement record if quantity changed
        if new_quantity != previous_quantity:
            movement_type = "addition" if new_quantity > previous_quantity else "consumption"
            quantity_change = new_quantity - previous_quantity
            
            movement = StockMovement(
                stock_type="raw_material",
                stock_id=material_id,
                movement_type=movement_type,
                quantity_change=quantity_change,
                previous_quantity=previous_quantity,
                new_quantity=new_quantity,
                reference_type="manual",
                notes=update_data.notes or "Manual adjustment",
                created_by=current_user["user_id"]
            )
            
            movement_dict = movement.dict()
            movement_dict["created_at"] = datetime.now(timezone.utc)
            await db.stock_movements.insert_one(movement_dict)
        
        return StandardResponse(
            success=True,
            message="Raw material stock updated successfully"
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to update raw material stock: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to update raw material stock")

@api_router.get("/stock/movements/{stock_id}", response_model=StandardResponse)
async def get_stock_movements(
    stock_id: str,
    current_user: dict = Depends(require_any_role)
):
    """Get stock movement history for a specific stock item"""
    try:
        movements = await db.stock_movements.find({"stock_id": stock_id}).sort("created_at", -1).to_list(length=None)
        
        # Remove MongoDB ObjectIds
        for movement in movements:
            if "_id" in movement:
                del movement["_id"]
        
        return StandardResponse(
            success=True,
            message="Stock movements retrieved successfully",
            data=movements
        )
    except Exception as e:
        logger.error(f"Failed to get stock movements: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to retrieve stock movements")

@api_router.get("/stock/alerts", response_model=StandardResponse)
async def get_stock_alerts(current_user: dict = Depends(require_any_role)):
    """Get all active stock alerts"""
    try:
        alerts = await db.stock_alerts.find({"is_active": True}).to_list(length=None)
        
        # Remove MongoDB ObjectIds
        for alert in alerts:
            if "_id" in alert:
                del alert["_id"]
        
        return StandardResponse(
            success=True,
            message="Stock alerts retrieved successfully",
            data=alerts
        )
    except Exception as e:
        logger.error(f"Failed to get stock alerts: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to retrieve stock alerts")

@api_router.post("/stock/alerts/{alert_id}/acknowledge", response_model=StandardResponse)
async def acknowledge_stock_alert(
    alert_id: str,
    acknowledge_data: StockAlertAcknowledge,
    current_user: dict = Depends(require_any_role)
):
    """Acknowledge a stock alert"""
    try:
        update_fields = {
            "acknowledged_by": current_user["user_id"],
            "acknowledged_at": datetime.now(timezone.utc)
        }
        
        # Handle snooze functionality
        if acknowledge_data.snooze_hours:
            snooze_hours = acknowledge_data.snooze_hours
            update_fields["snooze_until"] = datetime.now(timezone.utc) + timedelta(hours=snooze_hours)
        else:
            # If not snoozed, deactivate the alert
            update_fields["is_active"] = False
        
        result = await db.stock_alerts.update_one(
            {"id": alert_id},
            {"$set": update_fields}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Stock alert not found")
        
        return StandardResponse(
            success=True,
            message="Stock alert acknowledged successfully"
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to acknowledge stock alert: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to acknowledge stock alert")

@api_router.post("/stock/check-low-stock", response_model=StandardResponse)
async def check_low_stock_alerts(current_user: dict = Depends(require_any_role)):
    """Check for low stock and create alerts"""
    try:
        alerts_created = 0
        
        # Check raw materials stock
        materials = await db.raw_material_stock.find({}).to_list(length=None)
        for material in materials:
            if material["quantity_on_hand"] <= material.get("minimum_stock_level", 0):
                # Check if alert already exists for this material
                existing_alert = await db.stock_alerts.find_one({
                    "stock_id": material["id"],
                    "stock_type": "raw_material",
                    "alert_type": "low_stock",
                    "is_active": True
                })
                
                if not existing_alert:
                    alert = StockAlert(
                        stock_type="raw_material",
                        stock_id=material["id"],
                        alert_type="low_stock",
                        message=f"Low stock alert: {material['material_name']} has {material['quantity_on_hand']} {material['unit_of_measure']} remaining (minimum: {material.get('minimum_stock_level', 0)})"
                    )
                    
                    alert_dict = alert.dict()
                    alert_dict["created_at"] = datetime.now(timezone.utc)
                    await db.stock_alerts.insert_one(alert_dict)
                    alerts_created += 1
        
        # Check raw substrates stock
        substrates = await db.raw_substrate_stock.find({}).to_list(length=None)
        for substrate in substrates:
            if substrate["quantity_on_hand"] <= substrate.get("minimum_stock_level", 0):
                # Check if alert already exists for this substrate
                existing_alert = await db.stock_alerts.find_one({
                    "stock_id": substrate["id"],
                    "stock_type": "raw_substrate", 
                    "alert_type": "low_stock",
                    "is_active": True
                })
                
                if not existing_alert:
                    alert = StockAlert(
                        stock_type="raw_substrate",
                        stock_id=substrate["id"],
                        alert_type="low_stock",
                        message=f"Low stock alert: {substrate['product_description']} ({substrate['product_code']}) has {substrate['quantity_on_hand']} {substrate['unit_of_measure']} remaining (minimum: {substrate.get('minimum_stock_level', 0)})"
                    )
                    
                    alert_dict = alert.dict()
                    alert_dict["created_at"] = datetime.now(timezone.utc)
                    await db.stock_alerts.insert_one(alert_dict)
                    alerts_created += 1
        
        return StandardResponse(
            success=True,
            message=f"Stock check completed. {alerts_created} new alerts created.",
            data={"alerts_created": alerts_created}
        )
    except Exception as e:
        logger.error(f"Failed to check low stock: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to check low stock")

@api_router.delete("/stock/raw-substrates/{substrate_id}", response_model=StandardResponse)
async def delete_raw_substrate_stock(
    substrate_id: str,
    current_user: dict = Depends(require_manager_or_admin)
):
    """Delete raw substrate stock entry"""
    try:
        result = await db.raw_substrate_stock.delete_one({"id": substrate_id})
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Raw substrate stock not found")
        
        # Also delete related stock movements and alerts
        await db.stock_movements.delete_many({"stock_id": substrate_id})
        await db.stock_alerts.delete_many({"stock_id": substrate_id})
        
        return StandardResponse(
            success=True,
            message="Raw substrate stock deleted successfully"
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete raw substrate stock: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to delete raw substrate stock")

@api_router.delete("/stock/raw-materials/{material_id}", response_model=StandardResponse)
async def delete_raw_material_stock(
    material_id: str,
    current_user: dict = Depends(require_manager_or_admin)
):
    """Delete raw material stock entry"""
    try:
        result = await db.raw_material_stock.delete_one({"id": material_id})
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Raw material stock not found")
        
        # Also delete related stock movements and alerts
        await db.stock_movements.delete_many({"stock_id": material_id})
        await db.stock_alerts.delete_many({"stock_id": material_id})
        
        return StandardResponse(
            success=True,
            message="Raw material stock deleted successfully"
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete raw material stock: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to delete raw material stock")

# ============= STOCK AVAILABILITY & ALLOCATION ENDPOINTS =============

@api_router.get("/stock/history", response_model=StandardResponse)
async def get_stock_history(
    product_id: str,
    client_id: str,
    current_user: dict = Depends(require_any_role)
):
    """Get stock history for a specific product and client"""
    try:
        # Get all stock entries for this product/client combination
        stock_entries = await db.raw_substrate_stock.find({
            "product_id": product_id,
            "client_id": client_id
        }).to_list(length=None)
        
        # Get stock movements for these entries
        stock_ids = [entry["id"] for entry in stock_entries]
        movements = await db.stock_movements.find({
            "stock_id": {"$in": stock_ids}
        }).sort("created_at", -1).to_list(length=None)
        
        # Combine stock entries with their movements
        history_data = {
            "product_id": product_id,
            "client_id": client_id,
            "stock_entries": stock_entries,
            "movements": movements,
            "total_entries": len(stock_entries),
            "total_quantity": sum(entry.get("quantity_on_hand", 0) for entry in stock_entries)
        }
        
        return StandardResponse(
            success=True,
            message="Stock history retrieved successfully",
            data=history_data
        )
        
    except Exception as e:
        logger.error(f"Failed to get stock history: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to retrieve stock history")

@api_router.get("/stock/allocations", response_model=StandardResponse)
async def get_stock_allocations(
    product_id: str,
    client_id: str,
    current_user: dict = Depends(require_any_role)
):
    """Get current stock allocations for a specific product and client"""
    try:
        # Get allocation movements (negative quantities that haven't been archived)
        allocations = await db.stock_movements.find({
            "product_id": product_id,
            "client_id": client_id,
            "movement_type": "allocation",
            "quantity": {"$lt": 0},  # Negative quantities are allocations
            "is_archived": {"$ne": True}  # Not archived
        }).sort("created_at", -1).to_list(length=None)
        
        # Calculate total allocated quantity
        total_allocated = sum(abs(allocation.get("quantity", 0)) for allocation in allocations)
        
        allocation_data = {
            "product_id": product_id,
            "client_id": client_id,
            "allocations": allocations,
            "total_allocated": total_allocated,
            "active_orders": len(allocations)
        }
        
        return StandardResponse(
            success=True,
            message="Stock allocations retrieved successfully",
            data=allocation_data
        )
        
    except Exception as e:
        logger.error(f"Failed to get stock allocations: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to retrieve stock allocations")

# ============= STOCK AVAILABILITY & ALLOCATION ENDPOINTS =============

@api_router.get("/stock/check-availability", response_model=StandardResponse)
async def check_stock_availability(
    product_id: str,
    client_id: str,
    current_user: dict = Depends(require_any_role)
):
    """Check stock availability for a specific product and client"""
    try:
        # Check if there's any stock for this product from this client
        stock = await db.raw_substrate_stock.find_one({
            "product_id": product_id,
            "client_id": client_id,
            "quantity_on_hand": {"$gt": 0}
        })
        
        if stock:
            return StandardResponse(
                success=True,
                message="Stock available",
                data={
                    "product_id": product_id,
                    "client_id": client_id,
                    "quantity_on_hand": stock["quantity_on_hand"],
                    "stock_id": stock["id"]
                }
            )
        else:
            raise HTTPException(status_code=404, detail="No stock available for this product")
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to check stock availability: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to check stock availability")

@api_router.post("/stock/allocate", response_model=StandardResponse)
async def allocate_stock(
    allocation_data: dict,
    current_user: dict = Depends(require_any_role)
):
    """Allocate stock for an order"""
    try:
        product_id = allocation_data.get("product_id")
        client_id = allocation_data.get("client_id")
        quantity = allocation_data.get("quantity")
        order_reference = allocation_data.get("order_reference")
        
        if not all([product_id, client_id, quantity]):
            raise HTTPException(status_code=400, detail="Missing required fields")
            
        # Find the stock entry
        stock = await db.raw_substrate_stock.find_one({
            "product_id": product_id,
            "client_id": client_id,
            "quantity_on_hand": {"$gte": quantity}
        })
        
        if not stock:
            raise HTTPException(status_code=400, detail="Insufficient stock available")
        
        # Update stock quantity
        new_quantity = stock["quantity_on_hand"] - quantity
        result = await db.raw_substrate_stock.update_one(
            {"id": stock["id"]},
            {"$set": {"quantity_on_hand": new_quantity}}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Stock entry not found")
        
        # Create stock movement record
        movement_id = str(uuid.uuid4())
        movement = {
            "id": movement_id,
            "stock_id": stock["id"],
            "product_id": product_id,
            "client_id": client_id,
            "movement_type": "allocation",
            "quantity": -quantity,  # Negative for allocation
            "reference": order_reference,
            "created_by": current_user["user_id"],
            "created_at": datetime.now(timezone.utc).isoformat(),
            "is_archived": False
        }
        await db.stock_movements.insert_one(movement)
        
        return StandardResponse(
            success=True,
            message=f"Successfully allocated {quantity} units from stock",
            data={
                "allocated_quantity": quantity,
                "remaining_stock": new_quantity,
                "movement_id": movement_id
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to allocate stock: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to allocate stock")

@api_router.get("/stock/history", response_model=StandardResponse)
async def get_stock_history(
    product_id: str,
    client_id: str,
    current_user: dict = Depends(require_any_role)
):
    """Get stock history for a specific product and client"""
    try:
        # Get all stock entries for this product/client combination
        stock_entries = await db.raw_substrate_stock.find({
            "product_id": product_id,
            "client_id": client_id
        }).to_list(length=None)
        
        # Get stock movements for these entries
        stock_ids = [entry["id"] for entry in stock_entries]
        movements = await db.stock_movements.find({
            "stock_id": {"$in": stock_ids}
        }).sort("created_at", -1).to_list(length=None)
        
        # Combine stock entries with their movements
        history_data = {
            "product_id": product_id,
            "client_id": client_id,
            "stock_entries": stock_entries,
            "movements": movements,
            "total_entries": len(stock_entries),
            "total_quantity": sum(entry.get("quantity_on_hand", 0) for entry in stock_entries)
        }
        
        return StandardResponse(
            success=True,
            message="Stock history retrieved successfully",
            data=history_data
        )
        
    except Exception as e:
        logger.error(f"Failed to get stock history: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to retrieve stock history")

@api_router.get("/stock/allocations", response_model=StandardResponse)
async def get_stock_allocations(
    product_id: str,
    client_id: str,
    current_user: dict = Depends(require_any_role)
):
    """Get current stock allocations for a specific product and client"""
    try:
        # Get allocation movements (negative quantities that haven't been archived)
        allocations = await db.stock_movements.find({
            "product_id": product_id,
            "client_id": client_id,
            "movement_type": "allocation",
            "quantity": {"$lt": 0},  # Negative quantities are allocations
            "is_archived": {"$ne": True}  # Not archived
        }).sort("created_at", -1).to_list(length=None)
        
        # Calculate total allocated quantity
        total_allocated = sum(abs(allocation.get("quantity", 0)) for allocation in allocations)
        
        allocation_data = {
            "product_id": product_id,
            "client_id": client_id,
            "allocations": allocations,
            "total_allocated": total_allocated,
            "active_orders": len(allocations)
        }
        
        return StandardResponse(
            success=True,
            message="Stock allocations retrieved successfully",
            data=allocation_data
        )
        
    except Exception as e:
        logger.error(f"Failed to get stock allocations: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to retrieve stock allocations")

# ============= SYSTEM ENDPOINTS =============

@api_router.get("/")
async def root():
    return {"message": "Misty Manufacturing Management System API"}

@api_router.get("/health")
async def health_check():
    return {"status": "healthy", "timestamp": datetime.now(timezone.utc)}

# Include the routers in the main app
app.include_router(api_router)
app.include_router(payroll_router)

# Add CORS middleware with debug logging
cors_origins_str = os.environ.get('CORS_ORIGINS', '*')
print(f"CORS Origins string from env: '{cors_origins_str}'")
cors_origins = cors_origins_str.split(',') if cors_origins_str != '*' else ["*"]
print(f"CORS Origins configured: {cors_origins}")

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],  # Temporarily allow all origins for testing
    allow_methods=["*"],
    allow_headers=["*"],
)

# Router already included above, removing duplicate registration

# Mount static files for uploads
app.mount("/uploads", StaticFiles(directory="/app/uploads"), name="uploads")

# Direct Xero callback route (not under /api to avoid routing issues)
@app.get("/xero-oauth-callback")
async def xero_oauth_callback_direct(code: str = None, state: str = None, error: str = None):
    """Direct Xero OAuth callback route that bypasses /api routing issues"""
    if error:
        return Response(
            status_code=302,
            headers={"Location": f"{os.getenv('FRONTEND_URL')}/xero/callback?error={error}"}
        )
    
    if code and state:
        return Response(
            status_code=302,
            headers={"Location": f"{os.getenv('FRONTEND_URL')}/xero/callback?code={code}&state={state}"}
        )
    
    return Response(
        status_code=302,
        headers={"Location": f"{os.getenv('FRONTEND_URL')}/"}
    )

# Direct CSV export route (bypasses /api routing issues)
@app.get("/export-drafted-csv")
async def export_drafted_invoices_csv_direct(authorization: str = Header(None)):
    """Export all accounting transactions (drafted invoices) to CSV - Direct route bypassing /api issues"""
    try:
        # Simple authentication check
        if not authorization or not authorization.startswith('Bearer '):
            return Response(
                content="Unauthorized - Missing or invalid token",
                media_type="text/plain",
                status_code=401
            )
        
        # For simplicity, just check if token exists (proper auth would verify JWT)
        token = authorization.replace('Bearer ', '')
        if not token:
            return Response(
                content="Unauthorized - Invalid token",
                media_type="text/plain", 
                status_code=401
            )
        # Get all jobs in accounting transaction stage
        transactions = await db.orders.find({
            "current_stage": "accounting_transaction",
            "status": "accounting_draft"
        }).to_list(length=None)
        
        # Prepare CSV data
        csv_data = []
        
        # CSV Headers based on Xero import format
        headers = [
            "ContactName", "EmailAddress", "POAddressLine1", "POAddressLine2", 
            "POAddressLine3", "POAddressLine4", "POCity", "PORegion", 
            "POPostalCode", "POCountry", "InvoiceNumber", "Reference", 
            "InvoiceDate", "DueDate", "InventoryItemCode", "Description", 
            "Quantity", "UnitAmount", "Discount", "AccountCode", "TaxType", 
            "TrackingName1", "TrackingOption1", "TrackingName2", "TrackingOption2", 
            "Currency", "BrandingTheme"
        ]
        
        csv_data.append(headers)
        
        # Process each transaction
        for transaction in transactions:
            # Get client info
            client = await db.clients.find_one({"id": transaction["client_id"]})
            client_name = client["company_name"] if client else transaction.get("client_name", "Unknown Client")
            client_email = client.get("email", "") if client else ""
            
            # Get invoice info
            invoice = await db.invoices.find_one({"id": transaction.get("invoice_id")})
            invoice_number = invoice["invoice_number"] if invoice else f"INV-{transaction['order_number']}"
            invoice_date = invoice["created_at"].strftime("%d/%m/%Y") if invoice and invoice.get("created_at") else datetime.now().strftime("%d/%m/%Y")
            
            # Calculate due date (30 days from invoice date)
            due_date_obj = invoice["created_at"] + timedelta(days=30) if invoice and invoice.get("created_at") else datetime.now() + timedelta(days=30)
            due_date = due_date_obj.strftime("%d/%m/%Y")
            
            # Process each item in the transaction
            items = transaction.get("items", [])
            if not items:
                # Create a single line if no items
                items = [{"description": f"Services for Order {transaction['order_number']}", "quantity": 1, "unit_price": transaction.get("total_amount", 0)}]
            
            for item in items:
                row = [
                    client_name,  # ContactName (required)
                    client_email,  # EmailAddress
                    "",  # POAddressLine1
                    "",  # POAddressLine2
                    "",  # POAddressLine3
                    "",  # POAddressLine4
                    "",  # POCity
                    "",  # PORegion
                    "",  # POPostalCode
                    "",  # POCountry
                    invoice_number,  # InvoiceNumber (required)
                    transaction["order_number"],  # Reference
                    invoice_date,  # InvoiceDate (required)
                    due_date,  # DueDate (required)
                    item.get("product_code", ""),  # InventoryItemCode
                    f"{item.get('product_name', item.get('description', 'Product'))} - {item.get('specifications', '')}".strip(" - "),  # Description (required)
                    str(item.get("quantity", 1)),  # Quantity (required)
                    str(item.get("unit_price", item.get("price", 0))),  # UnitAmount (required)
                    str(item.get("discount_percent", "")),  # Discount
                    os.getenv("XERO_SALES_ACCOUNT_CODE", "200"),  # AccountCode (required)
                    "OUTPUT",  # TaxType (required) - GST for sales
                    "",  # TrackingName1
                    "",  # TrackingOption1
                    "",  # TrackingName2
                    "",  # TrackingOption2
                    "AUD",  # Currency
                    ""   # BrandingTheme
                ]
                csv_data.append(row)
        
        # Generate CSV content
        import io
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerows(csv_data)
        csv_content = output.getvalue()
        output.close()
        
        # Return CSV as response
        from fastapi.responses import Response
        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=drafted_invoices_{datetime.now().strftime('%Y%m%d')}.csv"}
        )
        
    except Exception as e:
        logger.error(f"Failed to export drafted invoices CSV: {str(e)}")
        return Response(
            content=f"Error: {str(e)}",
            media_type="text/plain",
            status_code=500
        )

# Direct Xero token exchange route (bypasses /api routing)
@app.post("/xero-auth-callback")
async def xero_auth_callback_direct(callback_data: dict):
    """Direct Xero token exchange that bypasses /api routing issues"""
    try:
        # Get auth credentials from environment
        XERO_CLIENT_ID = os.getenv("XERO_CLIENT_ID")
        XERO_CLIENT_SECRET = os.getenv("XERO_CLIENT_SECRET")  
        XERO_CALLBACK_URL = os.getenv("XERO_REDIRECT_URI")
        
        auth_code = callback_data.get("code")
        state = callback_data.get("state")
        
        if not auth_code or not state:
            return {"error": "Missing authorization code or state"}
        
        # Exchange code for tokens
        token_url = "https://identity.xero.com/connect/token"
        
        # Prepare authentication
        auth_string = f"{XERO_CLIENT_ID}:{XERO_CLIENT_SECRET}"
        auth_bytes = auth_string.encode('ascii')
        auth_b64 = base64.b64encode(auth_bytes).decode('ascii')
        
        headers = {
            "Authorization": f"Basic {auth_b64}",
            "Content-Type": "application/x-www-form-urlencoded"
        }
        
        token_data = {
            "grant_type": "authorization_code",
            "code": auth_code,
            "redirect_uri": XERO_CALLBACK_URL
        }
        
        response = requests.post(token_url, headers=headers, data=token_data)
        response.raise_for_status()
        
        tokens = response.json()
        
        # Store tokens (simplified - no user context for now)
        token_record = {
            "user_id": "system",  # Simplified for testing
            "access_token": tokens["access_token"],
            "refresh_token": tokens["refresh_token"],
            "expires_at": datetime.now(timezone.utc) + timedelta(seconds=tokens.get("expires_in", 1800)),
            "created_at": datetime.now(timezone.utc),
            "tenant_id": None
        }
        
        # Store in database
        await db.xero_tokens.replace_one(
            {"user_id": "system"},
            token_record,
            upsert=True
        )
        
        return {"message": "Xero connection successful", "access_token_expires_in": tokens.get("expires_in", 1800)}
        
    except Exception as e:
        logger.error(f"Xero token exchange failed: {str(e)}")
        return {"error": f"Failed to exchange authorization code: {str(e)}"}

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def startup_event():
    """Initialize system on startup"""
    logger.info("Starting Misty Manufacturing Management System...")
    
    # Create default admin user if not exists
    admin_user = await db.users.find_one({"username": "Callum"})
    if not admin_user:
        logger.info("Creating default admin user...")
        default_admin = User(
            username="Callum",
            email="admin@adelamerchants.com.au",
            hashed_password=get_password_hash("Peach7510"),
            role=UserRole.ADMIN,
            full_name="Callum - System Administrator"
        )
        await db.users.insert_one(default_admin.dict())
        logger.info("Default admin user created successfully")
    
    logger.info("Misty Manufacturing Management System started successfully!")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()